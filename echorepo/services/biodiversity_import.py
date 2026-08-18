import csv
import io
import logging
import re
from datetime import date, datetime
from io import BytesIO
from pathlib import Path

from flask import abort
from openpyxl import load_workbook
from psycopg2.extras import Json, execute_values

from .db import get_pg_conn
from .storage.minio import (
    StorageError,
    StorageNotConfigured,
    archive_raw_biodiversity_upload,
    invalidate_biodiversity_charts,
)

BIODIVERSITY_SAMPLE_RE = re.compile(
    r"^[A-Za-z0-9]{4}-[A-Za-z0-9]{4,}-(16S|ITS)$",
    re.IGNORECASE,
)

BIODIVERSITY_MIN_RELATIVE_ABUNDANCE_PCT = 1.0


def _normalise_biodiversity_header(value) -> str:
    return re.sub(
        r"[^a-z0-9]+",
        "",
        str(value or "").strip().lower(),
    )


def _is_biodiversity_header(header) -> bool:
    """
    Recognise wide OTU files such as:

        OTU ID,AAAA-1111-16S,...,Kingdom,Phylum,...,Species

    or:

        OTU ID,AAAA-1111-ITS,...,Kingdom,Phylum,...,Species
    """
    if not header:
        return False

    cleaned = ["" if value is None else str(value).strip() for value in header]

    normalised = {_normalise_biodiversity_header(value) for value in cleaned}

    has_otu_id = bool(
        normalised.intersection(
            {
                "otuid",
                "otu",
            }
        )
    )

    has_sample_column = any(BIODIVERSITY_SAMPLE_RE.fullmatch(value) for value in cleaned if value)

    # Normal format:
    # Kingdom, Phylum, Class, Order, Family, Genus, Species
    has_named_taxonomy = "phylum" in normalised or "philum" in normalised

    # Older legacy format:
    # Taxonomy, A, B, C, D, E, F
    #
    # Taxonomy = Kingdom
    # A = Phylum
    # ...
    # F = Species
    has_taxonomy_a_to_f = "taxonomy" in normalised and all(
        column in normalised for column in ("a", "b", "c", "d", "e", "f")
    )

    # Current compact taxonomy layout used by some clean_phylum sheets:
    # A, B, C, D, E, F, G
    #
    # A = Kingdom
    # B = Phylum
    # C = Class
    # D = Order
    # E = Family
    # F = Genus
    # G = Species
    has_a_to_g_taxonomy = all(
        column in normalised for column in ("a", "b", "c", "d", "e", "f", "g")
    )

    return (
        has_otu_id
        and has_sample_column
        and (has_named_taxonomy or has_taxonomy_a_to_f or has_a_to_g_taxonomy)
    )


def _open_biodiversity_rows(
    file_bytes: bytes,
    filename: str,
):
    """
    Open either:

      - XLSX biodiversity workbook;
      - CSV biodiversity file;
      - TSV biodiversity file.

    Returns:

        header, rows_iterator, close_function, source_description
    """
    suffix = Path(filename or "").suffix.lower()

    # --------------------------------------------------------------
    # XLSX
    # --------------------------------------------------------------
    if suffix == ".xlsx":
        try:
            wb = load_workbook(
                BytesIO(file_bytes),
                read_only=True,
                data_only=True,
            )
        except Exception as e:
            raise ValueError(f"Cannot open biodiversity XLSX: {e}") from e

        sheet_lookup = {str(sheet_name).strip().lower(): sheet_name for sheet_name in wb.sheetnames}

        selected_sheet = sheet_lookup.get("clean_phylum")

        # Also support an XLSX whose biodiversity data are in another
        # sheet, provided the header is recognisable.
        if not selected_sheet:
            for sheet_name in wb.sheetnames:
                ws_candidate = wb[sheet_name]

                first_row = next(
                    ws_candidate.iter_rows(
                        min_row=1,
                        max_row=1,
                        values_only=True,
                    ),
                    None,
                )

                if _is_biodiversity_header(first_row):
                    selected_sheet = sheet_name
                    break

        if not selected_sheet:
            available = ", ".join(wb.sheetnames)
            wb.close()

            raise ValueError(
                "No biodiversity sheet was found. Expected "
                "'clean_phylum' or a sheet containing OTU ID, "
                "sample-marker columns and Phylum. "
                f"Available sheets: {available}"
            )

        ws = wb[selected_sheet]
        rows_iter = ws.iter_rows(values_only=True)

        try:
            header = next(rows_iter)
        except StopIteration:
            wb.close()
            raise ValueError(f"Biodiversity sheet '{selected_sheet}' is empty") from None

        return (
            list(header),
            rows_iter,
            wb.close,
            f"XLSX sheet '{selected_sheet}'",
        )

    # --------------------------------------------------------------
    # CSV / TSV / TXT
    # --------------------------------------------------------------
    if suffix not in {".csv", ".tsv", ".txt"}:
        raise ValueError("Unsupported biodiversity file type. Use .xlsx, .csv or .tsv.")

    byte_stream = BytesIO(file_bytes)

    try:
        text_stream = io.TextIOWrapper(
            byte_stream,
            encoding="utf-8-sig",
            newline="",
        )

        probe = text_stream.read(65536)
        text_stream.seek(0)

    except UnicodeDecodeError as e:
        byte_stream.close()

        raise ValueError("Cannot decode biodiversity text file as UTF-8") from e

    if suffix == ".tsv":
        dialect = csv.excel_tab
        delimiter_name = "tab"
    else:
        try:
            dialect = csv.Sniffer().sniff(
                probe,
                delimiters=",\t;",
            )
            delimiter_name = repr(dialect.delimiter)
        except csv.Error:
            # Normal CSV fallback.
            dialect = csv.excel
            delimiter_name = "comma"

    reader = csv.reader(
        text_stream,
        dialect=dialect,
    )

    try:
        header = next(reader)
    except StopIteration:
        text_stream.close()
        raise ValueError("Biodiversity CSV/TSV is empty") from None

    return (
        header,
        reader,
        text_stream.close,
        f"delimited text, delimiter={delimiter_name}",
    )


def _looks_like_biodiversity_file(
    file_bytes: bytes,
    filename: str,
) -> bool:
    close_source = None

    try:
        (
            header,
            _rows_iter,
            close_source,
            _source_description,
        ) = _open_biodiversity_rows(
            file_bytes,
            filename,
        )

        return _is_biodiversity_header(header)

    except Exception:
        return False

    finally:
        if close_source is not None:
            try:
                close_source()
            except Exception:
                pass


def _import_biodiversity_streaming(
    file_bytes: bytes,
    filename: str,
    uploader_id: str,
):
    """
    Import XLSX, CSV or TSV biodiversity data while preserving both:

      1. raw feature/OTU-level taxonomy and sparse non-zero abundances;
      2. the existing compact Phylum-level statistics.

    The original upload is also archived unchanged in MinIO. The SHA-256
    of that source file is used as the content-addressed upload_id.
    """
    log = logging.getLogger(__name__)

    level = "Phylum"

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------
    def normalize_header(value) -> str:
        return re.sub(
            r"[^a-z0-9]+",
            "",
            str(value or "").strip().lower(),
        )

    def split_sample_marker(value: str) -> tuple[str, str]:
        parts = str(value or "").strip().rsplit("-", 1)

        if len(parts) != 2:
            return "", ""

        sample_id = parts[0].strip().upper()
        marker = parts[1].strip().upper()

        if marker not in {"16S", "ITS"}:
            return "", ""

        return sample_id, marker

    def to_float_or_none(value):
        if value is None:
            return None

        text = str(value).strip().replace(",", ".")

        if not text:
            return None

        if text.lower() in {
            "nan",
            "none",
            "null",
            "na",
            "n/a",
        }:
            return None

        try:
            result = float(text)
        except (TypeError, ValueError):
            return None

        return result

    def clean_phylum_value(value) -> str | None:
        """
        Accept values such as:

            p__Ascomycota
            Ascomycota
            k__Fungi;p__Ascomycota;c__Dothideomycetes
            ...|k__Fungi;p__Ascomycota;...|...

        Returns:

            Ascomycota
        """
        if value is None:
            return None

        raw = str(value).strip()

        if not raw:
            return None

        if raw.lower() in {
            "nan",
            "none",
            "null",
            "na",
            "n/a",
            "unassigned",
            "unknown",
        }:
            return None

        # If this is a complete taxonomy string, locate the p__ token.
        tokens = re.split(r"[;|]", raw)

        phylum_token = None

        for token in tokens:
            token = token.strip()

            if re.match(r"^p__", token, flags=re.IGNORECASE):
                phylum_token = token
                break

        if phylum_token is not None:
            raw = phylum_token

        # Remove any one-letter taxonomy prefix such as p__.
        raw = re.sub(
            r"^[A-Za-z]__",
            "",
            raw,
            flags=re.IGNORECASE,
        ).strip()

        if not raw:
            return None

        if raw.lower() in {
            "unclassified",
            "unassigned",
            "unknown",
            "uncultured",
        }:
            return None

        return raw

    # ------------------------------------------------------------------
    # Open XLSX, CSV or TSV source
    # ------------------------------------------------------------------
    try:
        (
            raw_header,
            rows_iter,
            close_source,
            source_description,
        ) = _open_biodiversity_rows(
            file_bytes,
            filename,
        )

    except ValueError as e:
        abort(
            400,
            description=str(e),
        )

    header = ["" if value is None else str(value).strip() for value in raw_header]

    if not any(header):
        close_source()

        abort(
            400,
            description="Biodiversity file header is empty",
        )

    if not _is_biodiversity_header(header):
        close_source()

        abort(
            400,
            description=(
                "The file does not look like biodiversity data. "
                "Expected OTU ID, one or more columns such as "
                "ABCD-1234-16S or ABCD-1234-ITS, and a Phylum column."
            ),
        )

    log.warning(
        "BIOUPLOAD: opened %s from file=%s",
        source_description,
        filename,
    )

    if not any(header):
        close_source()
        abort(
            400,
            description="Biodiversity file header is empty",
        )

    # The first column remains the OTU ID column, but the ID will not be
    # stored. It is used only to recognize valid data rows.
    otu_col_idx = 0

    # ------------------------------------------------------------------
    # Identify sample columns
    # ------------------------------------------------------------------
    sample_cols = []

    for idx, column_name in enumerate(header):
        if idx == otu_col_idx:
            continue

        if not BIODIVERSITY_SAMPLE_RE.fullmatch(str(column_name).strip()):
            continue

        sample_id, marker = split_sample_marker(column_name)

        if not sample_id or not marker:
            continue

        sample_cols.append(
            {
                # Stable position within this uploaded biodiversity file.
                # Starts at 1 intentionally.
                "sample_index": len(sample_cols) + 1,
                # Python row index, zero-based.
                "idx": idx,
                # Original CSV/XLSX column number, one-based.
                "source_column_number": idx + 1,
                # Preserve exactly what the laboratory file called it.
                "column_name": column_name,
                "source_sample_label": column_name,
                # Parsed ECHOREPO identity.
                "sample_id": sample_id,
                "marker": marker,
            }
        )

    if not sample_cols:
        close_source()

        abort(
            400,
            description=(
                "No sample columns found. Expected column names "
                "such as ABCD-1234-16S or ABCD-1234-ITS."
            ),
        )

    sample_col_indices = {item["idx"] for item in sample_cols}

    taxonomy_cols = [
        (idx, column_name)
        for idx, column_name in enumerate(header)
        if idx != otu_col_idx and idx not in sample_col_indices
    ]

    # ------------------------------------------------------------------
    # Map source taxonomy columns to standard ranks
    # ------------------------------------------------------------------

    normalised_taxonomy_headers = {
        normalize_header(column_name): idx for idx, column_name in taxonomy_cols
    }

    # Supported compact taxonomy layouts:
    #
    # 1) Taxonomy | A | B | C | D | E | F
    #       Taxonomy = Kingdom
    #       A = Phylum
    #       B = Class
    #       C = Order
    #       D = Family
    #       E = Genus
    #       F = Species
    #
    # 2) A | B | C | D | E | F | G
    #       A = Kingdom
    #       B = Phylum
    #       C = Class
    #       D = Order
    #       E = Family
    #       F = Genus
    #       G = Species
    taxonomy_a_to_f_layout = "taxonomy" in normalised_taxonomy_headers and all(
        column in normalised_taxonomy_headers for column in ("a", "b", "c", "d", "e", "f")
    )

    a_to_g_taxonomy_layout = all(
        column in normalised_taxonomy_headers for column in ("a", "b", "c", "d", "e", "f", "g")
    )

    if taxonomy_a_to_f_layout:
        taxonomy_rank_indices = {
            "kingdom": normalised_taxonomy_headers.get("taxonomy"),
            "phylum": normalised_taxonomy_headers.get("a"),
            "class_name": normalised_taxonomy_headers.get("b"),
            "order_name": normalised_taxonomy_headers.get("c"),
            "family": normalised_taxonomy_headers.get("d"),
            "genus": normalised_taxonomy_headers.get("e"),
            "species": normalised_taxonomy_headers.get("f"),
        }
    elif a_to_g_taxonomy_layout:
        taxonomy_rank_indices = {
            "kingdom": normalised_taxonomy_headers.get("a"),
            "phylum": normalised_taxonomy_headers.get("b"),
            "class_name": normalised_taxonomy_headers.get("c"),
            "order_name": normalised_taxonomy_headers.get("d"),
            "family": normalised_taxonomy_headers.get("e"),
            "genus": normalised_taxonomy_headers.get("f"),
            "species": normalised_taxonomy_headers.get("g"),
        }
    else:
        taxonomy_rank_indices = {
            "kingdom": normalised_taxonomy_headers.get("kingdom"),
            "phylum": (
                normalised_taxonomy_headers.get("phylum")
                if "phylum" in normalised_taxonomy_headers
                else normalised_taxonomy_headers.get("philum")
            ),
            "class_name": normalised_taxonomy_headers.get("class"),
            "order_name": normalised_taxonomy_headers.get("order"),
            "family": normalised_taxonomy_headers.get("family"),
            "genus": normalised_taxonomy_headers.get("genus"),
            "species": normalised_taxonomy_headers.get("species"),
        }

    def clean_taxonomy_rank(
        value,
        expected_prefix: str,
    ) -> str | None:
        if value is None:
            return None

        # Excel may coerce taxonomy labels such as "11-24" into a
        # month/year date. For taxonomy cells, repair that representation.
        if isinstance(value, datetime):
            if (
                value.day == 1
                and value.hour == 0
                and value.minute == 0
                and value.second == 0
                and value.microsecond == 0
            ):
                raw = f"{value.month}-{value.year % 100:02d}"
            else:
                raw = str(value).strip()
        elif isinstance(value, date):
            if value.day == 1:
                raw = f"{value.month}-{value.year % 100:02d}"
            else:
                raw = str(value).strip()
        else:
            raw = str(value).strip()

        # Also repair ISO text produced by an earlier import of an
        # Excel-coerced month/year taxonomy token.
        match = re.fullmatch(
            r"(\d{4})-(\d{2})-01(?:[ T]00:00:00(?:\.0+)?)?",
            raw,
        )
        if match:
            year = int(match.group(1))
            month = int(match.group(2))
            if 1 <= month <= 12:
                raw = f"{month}-{year % 100:02d}"

        if not raw:
            return None

        if raw.lower() in {
            "nan",
            "none",
            "null",
            "na",
            "n/a",
            "unassigned",
            "unknown",
        }:
            return None

        normalized_unassigned = re.sub(
            r"[\s_-]+",
            "",
            raw.lower(),
        )
        if normalized_unassigned in {
            "nohit",
            "nohits",
            "unassigned",
            "unknown",
        }:
            return None

        # A cell might contain either one value:
        #
        #   p__Ascomycota
        #
        # or an entire taxonomy string:
        #
        #   k__Fungi;p__Ascomycota;c__Sordariomycetes
        #
        wanted = expected_prefix.lower() + "__"

        for token in re.split(r"[;|]", raw):
            token = token.strip()

            if token.lower().startswith(wanted):
                raw = token
                break

        # Remove k__, p__, c__, etc.
        raw = re.sub(
            r"^[A-Za-z]__",
            "",
            raw,
            flags=re.IGNORECASE,
        ).strip()

        if not raw:
            return None

        return raw

    def taxonomy_for_row(row) -> dict:
        prefix_by_rank = {
            "kingdom": "k",
            "phylum": "p",
            "class_name": "c",
            "order_name": "o",
            "family": "f",
            "genus": "g",
            "species": "s",
        }

        result = {}

        for rank, idx in taxonomy_rank_indices.items():
            if idx is None or idx >= len(row):
                result[rank] = None
                continue

            result[rank] = clean_taxonomy_rank(
                row[idx],
                prefix_by_rank[rank],
            )

        return result

    def taxonomy_source_for_row(row) -> dict:
        """
        Preserve the original taxonomy fields exactly as supplied.
        """
        result = {}

        for idx, column_name in taxonomy_cols:
            value = row[idx] if idx < len(row) else None

            result[str(column_name)] = None if value is None else str(value)

        return result

    def taxonomy_raw_for_row(row) -> str | None:
        """
        Human-readable representation of the original taxonomy fields.
        The exact source fields are also retained in taxonomy_source JSONB.
        """
        values = []

        for idx, _column_name in taxonomy_cols:
            if idx >= len(row):
                continue

            value = row[idx]

            if value is None:
                continue

            text = str(value).strip()

            if text:
                values.append(text)

        return ";".join(values) or None

    # ------------------------------------------------------------------
    # Locate an explicit Phylum column
    # ------------------------------------------------------------------
    phylum_col_idx = None

    explicit_phylum_headers = {
        "phylum",
        "philum",  # tolerate the common misspelling
        "p",
    }

    for idx, column_name in taxonomy_cols:
        if normalize_header(column_name) in explicit_phylum_headers:
            phylum_col_idx = idx
            break

    # Support compact taxonomy layouts.
    if phylum_col_idx is None:
        normalized_headers = {
            normalize_header(column_name): idx for idx, column_name in taxonomy_cols
        }

        # Taxonomy | A | ... | F
        # Taxonomy = Kingdom, A = Phylum
        if "taxonomy" in normalized_headers and all(
            column in normalized_headers for column in ("a", "b", "c", "d", "e", "f")
        ):
            phylum_col_idx = normalized_headers["a"]

        # A | B | ... | G
        # A = Kingdom, B = Phylum
        elif all(column in normalized_headers for column in ("a", "b", "c", "d", "e", "f", "g")):
            phylum_col_idx = normalized_headers["b"]

    def extract_phylum(row) -> str:
        # Preferred route: a dedicated Phylum column.
        if phylum_col_idx is not None:
            value = row[phylum_col_idx] if phylum_col_idx < len(row) else None

            cleaned = clean_phylum_value(value)

            if cleaned:
                return cleaned

        # Fallback: inspect all taxonomy columns for a p__ token.
        for idx, _column_name in taxonomy_cols:
            value = row[idx] if idx < len(row) else None

            if value is None:
                continue

            raw = str(value).strip()

            if not raw:
                continue

            if re.search(
                r"(^|[;|])\s*p__",
                raw,
                flags=re.IGNORECASE,
            ):
                cleaned = clean_phylum_value(raw)

                if cleaned:
                    return cleaned

        return "Unclassified"

    # ------------------------------------------------------------------
    # Archive original source and determine content-addressed upload ID
    # ------------------------------------------------------------------
    #
    # We do this before streaming the OTU rows because upload_id is the
    # SHA-256 of the original source file and is needed by all raw child rows.
    #
    try:
        raw_archive = archive_raw_biodiversity_upload(
            file_bytes=file_bytes,
            filename=filename,
            uploader_id=uploader_id,
            aggregation_level=level,
        )
    except StorageNotConfigured as exc:
        close_source()
        abort(
            503,
            description=str(exc),
        )
    except StorageError as exc:
        close_source()
        abort(
            503,
            description=str(exc),
        )

    raw_archive_object = raw_archive.object_name
    source_sha256 = raw_archive.sha256
    upload_id = source_sha256

    # ------------------------------------------------------------------
    # Source sample-column metadata
    # ------------------------------------------------------------------
    #
    # sample_count is the number of distinct parsed ECHOREPO sample IDs
    # represented by columns in this source file. marker_count is the
    # number of markers (normally one per file, but the importer supports
    # both 16S and ITS columns if present).
    #
    sample_count = len({item["sample_id"] for item in sample_cols})

    marker_count = len({item["marker"] for item in sample_cols})

    raw_sample_rows = [
        (
            upload_id,
            item["sample_index"],
            item["source_column_number"],
            item["source_sample_label"],
            item["sample_id"],
            item["marker"],
        )
        for item in sample_cols
    ]

    # All sample/marker pairs in the source are considered affected,
    # including columns whose abundance is entirely zero. This prevents
    # stale aggregate rows from surviving a replacement import.
    affected_sample_markers = sorted(
        {
            (
                item["sample_id"],
                item["marker"],
            )
            for item in sample_cols
        }
    )

    # ------------------------------------------------------------------
    # Stream raw OTUs + raw abundances + existing Phylum aggregation
    # ------------------------------------------------------------------
    aggregates: dict[
        tuple[str, str],
        dict[str, float],
    ] = {}

    source_rows = 0
    nonzero_values = 0
    feature_index = 0
    excluded_unassigned_features = 0

    FEATURE_BATCH_SIZE = 5_000
    ABUNDANCE_BATCH_SIZE = 10_000

    feature_batch = []
    abundance_batch = []
    aggregate_rows = []

    try:
        with get_pg_conn() as conn, conn.cursor() as cur:
            # ----------------------------------------------------------
            # 1. Upsert the parent provenance record
            # ----------------------------------------------------------
            #
            # source_row_count and nonzero_value_count are set to zero
            # temporarily and finalized after the source has been streamed.
            #
            cur.execute(
                """
                INSERT INTO biodiversity_uploads (
                    upload_id,
                    original_filename,
                    archive_object_name,
                    sha256,
                    aggregation_level,
                    sample_count,
                    marker_count,
                    source_row_count,
                    nonzero_value_count,
                    uploaded_at,
                    uploaded_by
                )
                VALUES (
                    %s,
                    %s,
                    %s,
                    %s,
                    %s,
                    %s,
                    %s,
                    0,
                    0,
                    now(),
                    %s
                )
                ON CONFLICT (upload_id) DO UPDATE SET
                    original_filename =
                        EXCLUDED.original_filename,
                    archive_object_name =
                        EXCLUDED.archive_object_name,
                    aggregation_level =
                        EXCLUDED.aggregation_level,
                    sample_count =
                        EXCLUDED.sample_count,
                    marker_count =
                        EXCLUDED.marker_count,
                    uploaded_at =
                        now(),
                    uploaded_by =
                        EXCLUDED.uploaded_by
                """,
                (
                    upload_id,
                    filename,
                    raw_archive_object,
                    source_sha256,
                    level,
                    sample_count,
                    marker_count,
                    uploader_id,
                ),
            )

            # ----------------------------------------------------------
            # 2. Safe re-import of this exact source file
            # ----------------------------------------------------------
            #
            # upload_id is content-addressed (SHA-256). If this exact file
            # was imported earlier when only aggregates were stored, these
            # deletes let us populate/rebuild its raw representation safely.
            #
            # Abundance must be deleted first because it references both
            # raw feature and raw sample rows.
            #
            cur.execute(
                """
                DELETE FROM biodiversity_raw_abundance
                WHERE upload_id = %s
                """,
                (upload_id,),
            )

            cur.execute(
                """
                DELETE FROM biodiversity_raw_features
                WHERE upload_id = %s
                """,
                (upload_id,),
            )

            cur.execute(
                """
                DELETE FROM biodiversity_raw_samples
                WHERE upload_id = %s
                """,
                (upload_id,),
            )

            # ----------------------------------------------------------
            # 3. Preserve the source sample-column definitions
            # ----------------------------------------------------------
            execute_values(
                cur,
                """
                INSERT INTO biodiversity_raw_samples (
                    upload_id,
                    sample_index,
                    source_column_number,
                    source_sample_label,
                    sample_id,
                    marker
                )
                VALUES %s
                """,
                raw_sample_rows,
                page_size=1_000,
            )

            # ----------------------------------------------------------
            # Batch flush helpers
            # ----------------------------------------------------------
            def flush_features():
                if not feature_batch:
                    return

                execute_values(
                    cur,
                    """
                    INSERT INTO biodiversity_raw_features (
                        upload_id,
                        feature_index,
                        source_row_number,
                        source_feature_id,
                        taxonomy_raw,
                        kingdom,
                        phylum,
                        class_name,
                        order_name,
                        family,
                        genus,
                        species,
                        taxonomy_source
                    )
                    VALUES %s
                    """,
                    feature_batch,
                    page_size=FEATURE_BATCH_SIZE,
                )

                feature_batch.clear()

            def flush_abundances():
                if not abundance_batch:
                    return

                # The abundance rows have FKs to raw features. Ensure every
                # referenced feature has reached PostgreSQL before abundance
                # rows are flushed. This matters because one OTU can have many
                # non-zero sample values and fill this batch before the feature
                # batch reaches its own size threshold.
                flush_features()

                execute_values(
                    cur,
                    """
                    INSERT INTO biodiversity_raw_abundance (
                        upload_id,
                        feature_index,
                        sample_index,
                        read_count
                    )
                    VALUES %s
                    """,
                    abundance_batch,
                    page_size=ABUNDANCE_BATCH_SIZE,
                )

                abundance_batch.clear()

            # ----------------------------------------------------------
            # 4. Stream every OTU / feature row once
            # ----------------------------------------------------------
            for source_row_number, row in enumerate(
                rows_iter,
                start=2,
            ):
                otu_id = "" if row[otu_col_idx] is None else str(row[otu_col_idx]).strip()

                if not otu_id:
                    continue

                source_rows += 1

                taxonomy = taxonomy_for_row(row)

                # Completely unassigned/no-hit OTUs are not part of the
                # structured biodiversity dataset. The untouched original
                # source remains archived in MinIO for provenance.
                if not any(taxonomy.values()):
                    excluded_unassigned_features += 1
                    continue

                feature_index += 1

                taxonomy_source = taxonomy_source_for_row(row)
                taxonomy_raw = taxonomy_raw_for_row(row)

                # Preserve every retained feature, including features that
                # happen to have zero abundance in all sample columns.
                feature_batch.append(
                    (
                        upload_id,
                        feature_index,
                        source_row_number,
                        otu_id,
                        taxonomy_raw,
                        taxonomy.get("kingdom"),
                        taxonomy.get("phylum"),
                        taxonomy.get("class_name"),
                        taxonomy.get("order_name"),
                        taxonomy.get("family"),
                        taxonomy.get("genus"),
                        taxonomy.get("species"),
                        Json(taxonomy_source),
                    )
                )

                if len(feature_batch) >= FEATURE_BATCH_SIZE:
                    flush_features()

                # Keep using the tolerant existing extractor for the compact
                # Phylum-level derivative.
                phylum = extract_phylum(row)

                for sample_info in sample_cols:
                    idx = sample_info["idx"]

                    value = row[idx] if idx < len(row) else None

                    count = to_float_or_none(value)

                    if count is None or count == 0:
                        continue

                    if count < 0:
                        raise ValueError(
                            "Negative abundance at source row "
                            f"{source_row_number}, column "
                            f"{sample_info['column_name']}: "
                            f"{count}"
                        )

                    # Raw sparse abundance: zeroes are intentionally omitted.
                    abundance_batch.append(
                        (
                            upload_id,
                            feature_index,
                            sample_info["sample_index"],
                            float(count),
                        )
                    )

                    if len(abundance_batch) >= ABUNDANCE_BATCH_SIZE:
                        flush_abundances()

                    nonzero_values += 1

                    # Existing compact Phylum aggregation.
                    key = (
                        sample_info["sample_id"],
                        sample_info["marker"],
                    )

                    taxon_counts = aggregates.setdefault(
                        key,
                        {},
                    )

                    taxon_counts[phylum] = (
                        taxon_counts.get(
                            phylum,
                            0.0,
                        )
                        + count
                    )

            # Flush whatever remains after the final source row.
            flush_features()
            flush_abundances()

            if not aggregates:
                raise ValueError(
                    "No non-zero biodiversity abundances were found in the uploaded file."
                )

            # ----------------------------------------------------------
            # 5. Build the existing compact Phylum rows
            # ----------------------------------------------------------
            for (
                sample_id,
                marker,
            ), taxon_counts in sorted(aggregates.items()):
                total_count = sum(taxon_counts.values())

                if total_count <= 0:
                    continue

                for taxon, read_count in sorted(
                    taxon_counts.items(),
                    key=lambda item: item[1],
                    reverse=True,
                ):
                    relative_abundance_pct = read_count / total_count * 100.0

                    aggregate_rows.append(
                        (
                            sample_id,
                            marker,
                            level,
                            taxon,
                            float(read_count),
                            float(relative_abundance_pct),
                            upload_id,
                            uploader_id,
                            filename,
                        )
                    )

            if not aggregate_rows:
                raise ValueError("No Phylum-level statistics could be produced.")

            # ----------------------------------------------------------
            # 6. Replace the complete aggregate result for every
            #    sample/marker represented by the uploaded source
            # ----------------------------------------------------------
            cur.executemany(
                """
                DELETE FROM sample_taxon_abundance
                WHERE sample_id = %s
                  AND marker = %s
                  AND level = %s
                """,
                [
                    (
                        sample_id,
                        marker,
                        level,
                    )
                    for (
                        sample_id,
                        marker,
                    ) in affected_sample_markers
                ],
            )

            execute_values(
                cur,
                """
                INSERT INTO sample_taxon_abundance (
                    sample_id,
                    marker,
                    level,
                    taxon,
                    read_count,
                    relative_abundance_pct,
                    source_upload_id,
                    uploaded_by,
                    source_file
                )
                VALUES %s
                ON CONFLICT (
                    sample_id,
                    marker,
                    level,
                    taxon
                )
                DO UPDATE SET
                    read_count =
                        EXCLUDED.read_count,
                    relative_abundance_pct =
                        EXCLUDED.relative_abundance_pct,
                    source_upload_id =
                        EXCLUDED.source_upload_id,
                    uploaded_at =
                        now(),
                    uploaded_by =
                        EXCLUDED.uploaded_by,
                    source_file =
                        EXCLUDED.source_file
                """,
                aggregate_rows,
                page_size=5_000,
            )

            # ----------------------------------------------------------
            # 7. Finalize provenance counts
            # ----------------------------------------------------------
            cur.execute(
                """
                UPDATE biodiversity_uploads
                SET
                    source_row_count = %s,
                    nonzero_value_count = (
                        SELECT COUNT(*)
                        FROM biodiversity_raw_abundance
                        WHERE upload_id = %s
                    ),
                    sample_count = %s,
                    marker_count = %s,
                    uploaded_at = now(),
                    uploaded_by = %s
                WHERE upload_id = %s
                """,
                (
                    source_rows,
                    upload_id,
                    sample_count,
                    marker_count,
                    uploader_id,
                    upload_id,
                ),
            )

            conn.commit()

    except ValueError as exc:
        abort(
            400,
            description=str(exc),
        )

    except Exception as exc:
        log.exception("Raw + Phylum biodiversity import failed")

        abort(
            500,
            description=(f"Postgres biodiversity import failed: {exc}"),
        )

    finally:
        close_source()

    # ------------------------------------------------------------------
    # Invalidate cached chart images
    # ------------------------------------------------------------------
    invalidate_biodiversity_charts(
        affected_sample_markers,
        level,
    )

    inserted = len(aggregate_rows)

    log.warning(
        (
            "BIOUPLOAD: source_rows=%d, stored_raw_features=%d, "
            "excluded_unassigned_features=%d, "
            "nonzero_values=%d, aggregate_rows=%d, "
            "samples=%d, markers=%s, raw_archive=%s"
        ),
        source_rows,
        feature_index,
        excluded_unassigned_features,
        nonzero_values,
        inserted,
        sample_count,
        sorted({marker for _sample_id, marker in affected_sample_markers}),
        raw_archive_object,
    )

    return inserted


def finalize_biodiversity_global_filter(
    min_relative_abundance_pct: float = (BIODIVERSITY_MIN_RELATIVE_ABUNDANCE_PCT),
) -> dict[str, int | float]:
    """Finalize the current biodiversity dataset after a batch of imports.

    The filter is GLOBAL across all current sample columns:

      * completely unassigned/no-hit OTUs are removed;
      * an OTU is removed only when it stays strictly below the threshold in
        every current sample;
      * an OTU reaching the threshold in at least one current sample is kept;
      * current Phylum aggregates and percentages are rebuilt afterwards.

    This must be called after the complete batch of source files has been
    imported. Calling it after every individual plate/source would make the
    cross-source rule depend on import order.

    Original archived source files and biodiversity_uploads provenance remain
    untouched, so the structured representation can be rebuilt if necessary.
    """
    if not 0 < min_relative_abundance_pct < 100:
        raise ValueError("min_relative_abundance_pct must be between 0 and 100")

    log = logging.getLogger(__name__)

    with get_pg_conn() as conn, conn.cursor() as cur:
        # A current sample/marker must resolve to exactly one source upload.
        cur.execute(
            """
            SELECT COUNT(*)
            FROM (
                SELECT
                    UPPER(sample_id) AS sample_id,
                    UPPER(marker) AS marker
                FROM sample_taxon_abundance
                WHERE source_upload_id IS NOT NULL
                GROUP BY UPPER(sample_id), UPPER(marker)
                HAVING COUNT(DISTINCT source_upload_id) > 1
            ) AS conflicts
            """
        )
        if int(cur.fetchone()[0]) != 0:
            raise RuntimeError(
                "Some current sample/marker pairs point to more than one "
                "source_upload_id; global biodiversity cleanup aborted."
            )

        # Temporary tables are transaction-local and disappear on commit.
        cur.execute(
            """
            CREATE TEMP TABLE _bio_current_sources ON COMMIT DROP AS
            SELECT
                UPPER(sample_id) AS sample_id,
                UPPER(marker) AS marker,
                MIN(source_upload_id) AS source_upload_id
            FROM sample_taxon_abundance
            WHERE source_upload_id IS NOT NULL
            GROUP BY UPPER(sample_id), UPPER(marker)
            """
        )
        cur.execute(
            """
            CREATE UNIQUE INDEX ON _bio_current_sources (
                sample_id,
                marker
            )
            """
        )

        cur.execute(
            """
            CREATE TEMP TABLE _bio_current_sample_columns ON COMMIT DROP AS
            SELECT
                current.sample_id,
                current.marker,
                current.source_upload_id AS upload_id,
                rs.sample_index
            FROM _bio_current_sources AS current
            JOIN biodiversity_raw_samples AS rs
              ON rs.upload_id = current.source_upload_id
             AND UPPER(rs.sample_id) = current.sample_id
             AND UPPER(rs.marker) = current.marker
            """
        )
        cur.execute(
            """
            CREATE UNIQUE INDEX ON _bio_current_sample_columns (
                sample_id,
                marker
            )
            """
        )

        cur.execute("SELECT COUNT(*) FROM _bio_current_sources")
        expected_samples = int(cur.fetchone()[0])
        cur.execute("SELECT COUNT(*) FROM _bio_current_sample_columns")
        resolved_samples = int(cur.fetchone()[0])
        if expected_samples != resolved_samples:
            raise RuntimeError(
                "Could not resolve every current sample/marker to exactly "
                f"one raw sample column: expected={expected_samples}, "
                f"resolved={resolved_samples}"
            )

        # Identify completely unassigned/no-hit feature rows that predate the
        # importer-side filter.
        cur.execute(
            """
            CREATE TEMP TABLE _bio_unassigned_features ON COMMIT DROP AS
            SELECT
                f.upload_id,
                f.feature_index,
                f.source_feature_id
            FROM biodiversity_raw_features AS f
            WHERE f.upload_id IN (
                SELECT DISTINCT upload_id
                FROM _bio_current_sample_columns
            )
            AND regexp_replace(
                    LOWER(BTRIM(COALESCE(f.kingdom, ''))),
                    '[^a-z0-9]+', '', 'g'
                ) IN (
                    '', 'nohit', 'nohits', 'unassigned', 'unknown',
                    'nan', 'none', 'null', 'na'
                )
            AND regexp_replace(
                    LOWER(BTRIM(COALESCE(f.phylum, ''))),
                    '[^a-z0-9]+', '', 'g'
                ) IN (
                    '', 'nohit', 'nohits', 'unassigned', 'unknown',
                    'nan', 'none', 'null', 'na'
                )
            AND regexp_replace(
                    LOWER(BTRIM(COALESCE(f.class_name, ''))),
                    '[^a-z0-9]+', '', 'g'
                ) IN (
                    '', 'nohit', 'nohits', 'unassigned', 'unknown',
                    'nan', 'none', 'null', 'na'
                )
            AND regexp_replace(
                    LOWER(BTRIM(COALESCE(f.order_name, ''))),
                    '[^a-z0-9]+', '', 'g'
                ) IN (
                    '', 'nohit', 'nohits', 'unassigned', 'unknown',
                    'nan', 'none', 'null', 'na'
                )
            AND regexp_replace(
                    LOWER(BTRIM(COALESCE(f.family, ''))),
                    '[^a-z0-9]+', '', 'g'
                ) IN (
                    '', 'nohit', 'nohits', 'unassigned', 'unknown',
                    'nan', 'none', 'null', 'na'
                )
            AND regexp_replace(
                    LOWER(BTRIM(COALESCE(f.genus, ''))),
                    '[^a-z0-9]+', '', 'g'
                ) IN (
                    '', 'nohit', 'nohits', 'unassigned', 'unknown',
                    'nan', 'none', 'null', 'na'
                )
            AND regexp_replace(
                    LOWER(BTRIM(COALESCE(f.species, ''))),
                    '[^a-z0-9]+', '', 'g'
                ) IN (
                    '', 'nohit', 'nohits', 'unassigned', 'unknown',
                    'nan', 'none', 'null', 'na'
                )
            """
        )
        cur.execute(
            """
            CREATE UNIQUE INDEX ON _bio_unassigned_features (
                upload_id,
                feature_index
            )
            """
        )

        # Denominator for the 1% rule: total reads in each current sample after
        # excluding completely unassigned/no-hit OTUs, but before the 1% filter.
        cur.execute(
            """
            CREATE TEMP TABLE _bio_sample_totals ON COMMIT DROP AS
            SELECT
                c.sample_id,
                c.marker,
                c.upload_id,
                c.sample_index,
                SUM(a.read_count)::double precision AS total_reads
            FROM _bio_current_sample_columns AS c
            JOIN biodiversity_raw_abundance AS a
              ON a.upload_id = c.upload_id
             AND a.sample_index = c.sample_index
            LEFT JOIN _bio_unassigned_features AS bad
              ON bad.upload_id = a.upload_id
             AND bad.feature_index = a.feature_index
            WHERE bad.feature_index IS NULL
            GROUP BY
                c.sample_id,
                c.marker,
                c.upload_id,
                c.sample_index
            """
        )

        cur.execute(
            """
            SELECT COUNT(*)
            FROM _bio_current_sample_columns AS c
            LEFT JOIN _bio_sample_totals AS t
              ON t.sample_id = c.sample_id
             AND t.marker = c.marker
            WHERE COALESCE(t.total_reads, 0) <= 0
            """
        )
        if int(cur.fetchone()[0]) != 0:
            raise RuntimeError(
                "At least one current sample/marker has zero reads after "
                "removing completely unassigned OTUs; cleanup aborted."
            )

        # Maximum relative abundance for each OTU ID across every current
        # sample. Sparse zeroes need no rows: their contribution is 0%.
        cur.execute(
            """
            CREATE TEMP TABLE _bio_feature_max_pct ON COMMIT DROP AS
            SELECT
                f.source_feature_id,
                MAX(
                    100.0 * a.read_count / t.total_reads
                )::double precision AS max_relative_abundance_pct
            FROM _bio_current_sample_columns AS c
            JOIN _bio_sample_totals AS t
              ON t.sample_id = c.sample_id
             AND t.marker = c.marker
             AND t.upload_id = c.upload_id
             AND t.sample_index = c.sample_index
            JOIN biodiversity_raw_abundance AS a
              ON a.upload_id = c.upload_id
             AND a.sample_index = c.sample_index
            JOIN biodiversity_raw_features AS f
              ON f.upload_id = a.upload_id
             AND f.feature_index = a.feature_index
            LEFT JOIN _bio_unassigned_features AS bad
              ON bad.upload_id = f.upload_id
             AND bad.feature_index = f.feature_index
            WHERE bad.feature_index IS NULL
            GROUP BY f.source_feature_id
            """
        )
        cur.execute(
            """
            CREATE UNIQUE INDEX ON _bio_feature_max_pct (
                source_feature_id
            )
            """
        )

        # If an OTU reaches the threshold in ANY current sample, retain every
        # occurrence of that OTU in the current structured dataset.
        cur.execute(
            """
            CREATE TEMP TABLE _bio_low_features ON COMMIT DROP AS
            SELECT
                f.upload_id,
                f.feature_index,
                f.source_feature_id
            FROM biodiversity_raw_features AS f
            LEFT JOIN _bio_feature_max_pct AS pct
              ON pct.source_feature_id = f.source_feature_id
            LEFT JOIN _bio_unassigned_features AS unassigned
              ON unassigned.upload_id = f.upload_id
             AND unassigned.feature_index = f.feature_index
            WHERE f.upload_id IN (
                SELECT DISTINCT upload_id
                FROM _bio_current_sample_columns
            )
              AND unassigned.feature_index IS NULL
              AND COALESCE(
                    pct.max_relative_abundance_pct,
                    0.0
                  ) < %s
            """,
            (float(min_relative_abundance_pct),),
        )
        cur.execute(
            """
            CREATE UNIQUE INDEX ON _bio_low_features (
                upload_id,
                feature_index
            )
            """
        )

        cur.execute(
            """
            CREATE TEMP TABLE _bio_features_to_remove ON COMMIT DROP AS
            SELECT
                upload_id,
                feature_index,
                source_feature_id,
                'unassigned'::text AS reason
            FROM _bio_unassigned_features

            UNION ALL

            SELECT
                upload_id,
                feature_index,
                source_feature_id,
                'below_threshold_everywhere'::text AS reason
            FROM _bio_low_features
            """
        )
        cur.execute(
            """
            CREATE UNIQUE INDEX ON _bio_features_to_remove (
                upload_id,
                feature_index
            )
            """
        )

        cur.execute(
            """
            SELECT COUNT(*)
            FROM _bio_features_to_remove
            WHERE reason = 'unassigned'
            """
        )
        unassigned_features = int(cur.fetchone()[0])

        cur.execute(
            """
            SELECT COUNT(*)
            FROM _bio_features_to_remove
            WHERE reason = 'below_threshold_everywhere'
            """
        )
        low_features = int(cur.fetchone()[0])

        cur.execute(
            """
            SELECT COUNT(*)
            FROM biodiversity_raw_abundance AS a
            JOIN _bio_features_to_remove AS bad
              ON bad.upload_id = a.upload_id
             AND bad.feature_index = a.feature_index
            """
        )
        abundance_rows = int(cur.fetchone()[0])

        # Never allow the filter to erase an entire current sample.
        cur.execute(
            """
            CREATE TEMP TABLE _bio_retained_sample_totals ON COMMIT DROP AS
            SELECT
                c.sample_id,
                c.marker,
                SUM(a.read_count)::double precision AS retained_reads
            FROM _bio_current_sample_columns AS c
            JOIN biodiversity_raw_abundance AS a
              ON a.upload_id = c.upload_id
             AND a.sample_index = c.sample_index
            LEFT JOIN _bio_features_to_remove AS bad
              ON bad.upload_id = a.upload_id
             AND bad.feature_index = a.feature_index
            WHERE bad.feature_index IS NULL
            GROUP BY c.sample_id, c.marker
            """
        )
        cur.execute(
            """
            SELECT COUNT(*)
            FROM _bio_current_sample_columns AS c
            LEFT JOIN _bio_retained_sample_totals AS retained
              ON retained.sample_id = c.sample_id
             AND retained.marker = c.marker
            WHERE COALESCE(retained.retained_reads, 0) <= 0
            """
        )
        if int(cur.fetchone()[0]) != 0:
            raise RuntimeError(
                "The global biodiversity threshold would leave at least "
                "one current sample/marker with zero retained reads; "
                "cleanup aborted."
            )

        # Delete FK children before raw feature rows.
        cur.execute(
            """
            DELETE FROM biodiversity_raw_abundance AS a
            USING _bio_features_to_remove AS bad
            WHERE a.upload_id = bad.upload_id
              AND a.feature_index = bad.feature_index
            """
        )
        cur.execute(
            """
            DELETE FROM biodiversity_raw_features AS f
            USING _bio_features_to_remove AS bad
            WHERE f.upload_id = bad.upload_id
              AND f.feature_index = bad.feature_index
            """
        )

        # Refresh current upload non-zero counts. source_row_count intentionally
        # remains the original source-row count for provenance.
        cur.execute(
            """
            UPDATE biodiversity_uploads AS bu
            SET nonzero_value_count = counts.nonzero_value_count
            FROM (
                SELECT
                    current_upload.upload_id,
                    COUNT(a.feature_index)::integer
                        AS nonzero_value_count
                FROM (
                    SELECT DISTINCT upload_id
                    FROM _bio_current_sample_columns
                ) AS current_upload
                LEFT JOIN biodiversity_raw_abundance AS a
                  ON a.upload_id = current_upload.upload_id
                GROUP BY current_upload.upload_id
            ) AS counts
            WHERE counts.upload_id = bu.upload_id
            """
        )

        # Rebuild all current Phylum aggregates so the percentages match the
        # retained OTU-level data.
        cur.execute(
            """
            DELETE FROM sample_taxon_abundance AS sta
            USING _bio_current_sources AS current
            WHERE UPPER(sta.sample_id) = current.sample_id
              AND UPPER(sta.marker) = current.marker
              AND LOWER(sta.level) = 'phylum'
            """
        )

        cur.execute(
            """
            WITH phylum_counts AS (
                SELECT
                    current.sample_id,
                    current.marker,
                    current.source_upload_id,
                    CASE
                        WHEN regexp_replace(
                            LOWER(BTRIM(COALESCE(f.phylum, ''))),
                            '[^a-z0-9]+', '', 'g'
                        ) IN (
                            '', 'nohit', 'nohits', 'unassigned',
                            'unknown', 'unclassified', 'uncultured',
                            'nan', 'none', 'null', 'na'
                        )
                            THEN 'Unclassified'
                        ELSE BTRIM(f.phylum)
                    END AS taxon,
                    SUM(a.read_count)::double precision AS read_count
                FROM _bio_current_sources AS current
                JOIN biodiversity_raw_samples AS rs
                  ON rs.upload_id = current.source_upload_id
                 AND UPPER(rs.sample_id) = current.sample_id
                 AND UPPER(rs.marker) = current.marker
                JOIN biodiversity_raw_abundance AS a
                  ON a.upload_id = rs.upload_id
                 AND a.sample_index = rs.sample_index
                JOIN biodiversity_raw_features AS f
                  ON f.upload_id = a.upload_id
                 AND f.feature_index = a.feature_index
                GROUP BY
                    current.sample_id,
                    current.marker,
                    current.source_upload_id,
                    CASE
                        WHEN regexp_replace(
                            LOWER(BTRIM(COALESCE(f.phylum, ''))),
                            '[^a-z0-9]+', '', 'g'
                        ) IN (
                            '', 'nohit', 'nohits', 'unassigned',
                            'unknown', 'unclassified', 'uncultured',
                            'nan', 'none', 'null', 'na'
                        )
                            THEN 'Unclassified'
                        ELSE BTRIM(f.phylum)
                    END
            ),
            with_totals AS (
                SELECT
                    counts.*,
                    SUM(counts.read_count) OVER (
                        PARTITION BY counts.sample_id, counts.marker
                    ) AS total_count
                FROM phylum_counts AS counts
            )
            INSERT INTO sample_taxon_abundance (
                sample_id,
                marker,
                level,
                taxon,
                read_count,
                relative_abundance_pct,
                source_upload_id,
                uploaded_by,
                source_file
            )
            SELECT
                totals.sample_id,
                totals.marker,
                'Phylum',
                totals.taxon,
                totals.read_count,
                totals.read_count / totals.total_count * 100.0,
                totals.source_upload_id,
                bu.uploaded_by,
                bu.original_filename
            FROM with_totals AS totals
            JOIN biodiversity_uploads AS bu
              ON bu.upload_id = totals.source_upload_id
            WHERE totals.read_count > 0
              AND totals.total_count > 0
            ON CONFLICT (
                sample_id,
                marker,
                level,
                taxon
            )
            DO UPDATE SET
                read_count = EXCLUDED.read_count,
                relative_abundance_pct =
                    EXCLUDED.relative_abundance_pct,
                source_upload_id = EXCLUDED.source_upload_id,
                uploaded_at = now(),
                uploaded_by = EXCLUDED.uploaded_by,
                source_file = EXCLUDED.source_file
            """
        )

        cur.execute(
            """
            SELECT COUNT(*)
            FROM sample_taxon_abundance
            WHERE LOWER(level) = 'phylum'
            """
        )
        aggregate_rows = int(cur.fetchone()[0])

    log.warning(
        (
            "BIODIVERSITY GLOBAL FILTER: threshold=%.4f%%, "
            "unassigned_features=%d, low_features=%d, "
            "abundance_rows_removed=%d, current_phylum_rows=%d"
        ),
        min_relative_abundance_pct,
        unassigned_features,
        low_features,
        abundance_rows,
        aggregate_rows,
    )

    return {
        "threshold_pct": float(min_relative_abundance_pct),
        "unassigned_features_removed": unassigned_features,
        "below_threshold_features_removed": low_features,
        "abundance_rows_removed": abundance_rows,
        "current_phylum_rows": aggregate_rows,
    }


# Public aliases for callers outside this module. The underscored names are
# retained so existing web.py imports/calls can be migrated gradually.
looks_like_biodiversity_file = _looks_like_biodiversity_file
import_biodiversity_file = _import_biodiversity_streaming
