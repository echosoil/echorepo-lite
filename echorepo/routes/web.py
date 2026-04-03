import csv
import io
import json
import logging
import os
import pathlib
import re
import sqlite3
import zipfile  # kept in case you later want to build ZIPs locally
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path

import pandas as pd
from flask import (
    Blueprint,
    abort,
    g,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from flask_babel import get_locale
from flask_babel import gettext as _
from openpyxl import load_workbook
from psycopg2.extras import RealDictCursor

from echorepo.i18n import build_i18n_labels
from echorepo.services.i18n_labels import make_labels
from echorepo.services.i18n_overrides import (
    _canon_locale,
    get_overrides_msgid,
)

from ..auth.decorators import login_required
from ..config import settings
from ..services.db import _ensure_lab_enrichment, get_pg_conn, query_sample, query_user_df
from ..services.lab_permissions import can_upload_lab_data
from ..services.validation import find_default_coord_rows, select_country_mismatches
from ..utils.table import make_table_html, strip_orig_cols

try:
    from echorepo.routes.data_api import CANONICAL_SAMPLE_COLS, _oxide_to_metal
except Exception:
    # fallback – if import fails, just no conversion
    def _oxide_to_metal(param, value):
        return None

    # fallback canonical columns (same as in data_api.py)
    CANONICAL_SAMPLE_COLS = [
        "sample_id",
        "timestamp_utc",
        "lat",
        "lon",
        "country_code",
        "location_accuracy_m",
        "ph",
        "organic_carbon_pct",
        "earthworms_count",
        "contamination_debris",
        "contamination_plastic",
        "contamination_other_orig",
        "contamination_other_en",
        "pollutants_count",
        "soil_structure_orig",
        "soil_structure_en",
        "soil_texture_orig",
        "soil_texture_en",
        "observations_orig",
        "observations_en",
        "metals_info_orig",
        "metals_info_en",
        "collected_by",
        "data_source",
        "qa_status",
        "licence",
    ]

# Zenodo sync log path (can also be set via env var)
ZENODO_LOG_DEFAULT = os.getenv("ZENODO_LOG_FILE", "/data/zenodo_sync_log.csv")

# constants for privacy acceptance
PRIVACY_VERSION = "2025-11-echo"  # bump when text changes
PRIVACY_CSV_PATH = os.getenv("PRIVACY_CSV_PATH", "/data/privacy_acceptances.csv")

# blueprint
web_bp = Blueprint("web", __name__)

# --- Remove oxides helpers ------------------------------------------------


def _looks_like_oxide(label: str) -> bool:
    """
    True for formulas like SiO2, Al2O3, FeO, K2O, P2O5, CaO, etc.
    (Any multi-token formula that contains an O-token.)
    """
    if not label:
        return False
    s = str(label).strip()
    s = re.sub(r"\(.*?\)", "", s)  # strip units like "(ppm)"
    # tokenize into element(+optional digits) chunks
    tokens = re.findall(r"[A-Z][a-z]?\d*", s)
    if len(tokens) < 2:
        return False
    # oxide if one of the tokens is 'O', 'O2', 'O3', ...
    return any(t.startswith("O") for t in tokens)


def _drop_oxide_rows(
    df: pd.DataFrame, code_col="parameter_code", name_col="parameter_name"
) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    mask = pd.Series(False, index=df.index)
    if code_col in df.columns:
        mask |= df[code_col].fillna("").map(_looks_like_oxide)
    if name_col in df.columns:
        mask |= df[name_col].fillna("").map(_looks_like_oxide)
    return df.loc[~mask].copy()


def _drop_oxide_columns_from_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    Remove columns whose names look like oxides (directly or as a suffix).
    Works for things like 'SiO2', 'lab_SiO2', 'value_Fe2O3_(%)', etc.
    """
    if df is None or df.empty:
        return df
    to_drop = []
    for col in df.columns:
        if "photo" in col.lower():
            continue
        name = str(col)
        # check whole name
        if _looks_like_oxide(name):
            to_drop.append(col)
            continue
        # also check the last token after separators
        last = re.split(r"[_\s/\-]+", name)[-1]
        if _looks_like_oxide(last):
            to_drop.append(col)
    if to_drop:
        df = df.drop(columns=to_drop, errors="ignore")
    return df


OXIDE_NAMES = {"MN2O3", "AL2O3", "CAO", "FE2O3", "MGO", "SIO2", "P2O5", "TIO2", "K2O", "SO3"}


def _strip_oxides_from_info_str(s: str) -> str:
    if not isinstance(s, str) or not s.strip():
        return ""
    parts = [p.strip() for p in s.split(";") if p.strip()]
    keep = []
    for token in parts:
        left = token.split("=", 1)[0]
        norm = re.sub(r"\s+", "", left).upper()
        if norm in OXIDE_NAMES:
            continue
        keep.append(token)
    return "; ".join(keep)


# --------------------------------------------------------------------------
# --- Privacy acceptance helpers -------------------------------------------
# --------------------------------------------------------------------------
def _env_true(name: str, default: bool = True) -> bool:
    val = os.getenv(name)
    if val is None:
        return default
    return str(val).strip().lower() in {"1", "true", "yes", "on", "y", "t"}


def _get_repo_user_id_from_db() -> str | None:
    """
    Return the userId we store in the samples table (the same we use for the
    survey r= param). Falls back to the display user_key if nothing else is found.
    """
    # this is the same way you find the "logical" user
    user_key = session.get("user") or session.get("kc", {}).get("profile", {}).get("email")
    if not user_key:
        return None

    df = query_user_df(user_key)
    if df is None or df.empty:
        return user_key  # last resort

    for col in ("userId", "user_id", "kc_user_id"):
        if col in df.columns:
            val = df[col].dropna().astype(str).iloc[0].strip()
            if val:
                return val

    return user_key


def _current_user_id() -> str | None:
    kc_profile = (session.get("kc") or {}).get("profile") or {}

    # prefer a stable internal KC id
    if kc_profile.get("id"):
        return kc_profile["id"]
    if kc_profile.get("sub"):
        return kc_profile["sub"]

    # then fall back to whatever you used before
    if session.get("user"):
        return session["user"]

    # last resort: email
    if kc_profile.get("email"):
        return kc_profile["email"]

    return None


def _has_accepted_privacy(user_id: str) -> bool:
    if not user_id:
        return False
    # fast path: we can also store in session
    if session.get("privacy_accepted_version") == PRIVACY_VERSION:
        return True

    if not os.path.exists(PRIVACY_CSV_PATH):
        return False

    try:
        with open(PRIVACY_CSV_PATH, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                if row.get("user_id") == user_id and row.get("version") == PRIVACY_VERSION:
                    # cache in session
                    session["privacy_accepted_version"] = PRIVACY_VERSION
                    return True
    except Exception:
        return False

    return False


def _append_privacy_acceptance(user_id: str):
    dir_name = os.path.dirname(PRIVACY_CSV_PATH)
    if dir_name:
        os.makedirs(dir_name, exist_ok=True)

    file_exists = os.path.exists(PRIVACY_CSV_PATH)
    with open(PRIVACY_CSV_PATH, "a", newline="", encoding="utf-8") as f:
        fieldnames = ["user_id", "accepted_at", "version"]
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        if not file_exists:
            writer.writeheader()
        writer.writerow(
            {
                "user_id": user_id,
                "accepted_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
                "version": PRIVACY_VERSION,
            }
        )
    session["privacy_accepted_version"] = PRIVACY_VERSION


# --------------------------------------------------------------------------
# SoSci survey helper
# --------------------------------------------------------------------------
def _user_has_lab_results(df_samples: pd.DataFrame) -> bool:
    """
    Return True if any of this user's samples have rows in sample_parameters.
    Works even if the samples DF doesn't contain any metals summary columns.
    """
    try:
        if df_samples is None or df_samples.empty:
            return False

        # Find a column that looks like the canonical sample id
        sid_col = None
        for cand in ("sampleId", "sample_id", "Sample", "sampleid"):
            if cand in df_samples.columns:
                sid_col = cand
                break
        if not sid_col:
            return False

        sample_ids = df_samples[sid_col].dropna().astype(str).unique().tolist()
        if not sample_ids:
            return False

        # Fast existence check in Postgres
        with get_pg_conn() as conn, conn.cursor() as cur:
            cur.execute(
                "SELECT 1 FROM sample_parameters WHERE sample_id = ANY(%s) LIMIT 1",
                (sample_ids,),
            )
            return cur.fetchone() is not None
    except Exception as e:
        logging.getLogger(__name__).warning("lab presence check failed: %s", e)
        return False


def _user_has_metals_legacy(df: pd.DataFrame) -> bool:
    if df is None or df.empty:
        return False
    candidates = [
        "METALS_info",
        "lab_METALS_info",
        "METALS",
        "metals",
        "metals_info",
        "metals_info_en",
        "metals_info_orig",
        "elemental_concentrations",
        "elemental_concentrations_en",
        "elemental_concentrations_orig",
    ]
    for c in df.columns:
        cl = str(c).lower()
        if c in candidates or "metals" in cl or ("elemental" in cl and "concentration" in cl):
            s = (
                df[c]
                .fillna("")
                .astype(str)
                .str.replace("<br>", ";", regex=False)
                .str.strip()
                .replace({"nan": "", "None": "", "NaN": "", "0": "", "0.0": ""})
            )
            if s.str.contains("=", regex=False).any():
                return True
    return False


def _build_sosci_url(user_id: str | None) -> str | None:
    if not user_id:
        return None

    sosci_map = {
        "de": "deu",
        "el": "gre",
        "en": "eng",
        "es": "spa",
        "fi": "fin",
        "it": "ita",
        "pl": "pol",
        "pt": "por",
        "ro": "rum",
    }

    current = str(get_locale() or "en")
    current_base = current.split("_", 1)[0].split("-", 1)[0].lower()

    if current_base == "en":
        browser = request.accept_languages.best_match(
            ["de", "it", "fi", "el", "es", "po", "pt", "ro", "en"]
        )
        if browser:
            current_base = browser.split("-", 1)[0].lower()

    sosci_lang = sosci_map.get(current_base, "eng")

    base = getattr(
        settings,
        "SURVEY_BASE_URL",
        "https://www.soscisurvey.de/default",
    )

    sep = "&" if "?" in base else "?"
    return f"{base}{sep}l={sosci_lang}&r={user_id}"


# --------------------------------------------------------------------------
# labels for front-end
# --------------------------------------------------------------------------
def _js_base_labels() -> dict:
    return {
        "privacyRadius": _("Privacy radius (~±{km} km)"),
        "soilPh": _("Soil pH"),
        "acid": _("Acidic (≤5.5)"),
        "slightlyAcid": _("Slightly acidic (5.5–6.5)"),
        "neutral": _("Neutral (6.5–7.5)"),
        "slightlyAlkaline": _("Slightly alkaline (7.5–8.5)"),
        "alkaline": _("Alkaline (≥8.5)"),
        "yourSamples": _("Your samples"),
        "otherSamples": _("Other samples"),
        "export": _("Export"),
        "clear": _("Clear"),
        "exportFiltered": _("Export filtered ({n})"),
        "date": _("Date"),
        "qr": _("QR code"),
        "ph": _("pH"),
        "colour": _("Colour"),
        "soilOrganicMatter": _("Soil organic matter"),
        "texture": _("Texture"),
        "structure": _("Structure"),
        "earthworms": _("Earthworms"),
        "plastic": _("Plastic"),
        "debris": _("Debris"),
        "contamination": _("Contamination"),
        "metals": _("Metals"),
        "elementalConcentrations": _("Elemental concentrations"),
        "drawRectangle": _("Draw a rectangle"),
        "cancelDrawing": _("Cancel drawing"),
        "cancel": _("Cancel"),
        "deleteLastPoint": _("Delete last point"),
        "drawRectangleHint": _("Click and drag to draw a rectangle."),
        "releaseToFinish": _("Release mouse to finish drawing."),
    }


# --------------------------------------------------------------------------
# helpers for lab upload / QR
# --------------------------------------------------------------------------
def _normalize_qr(raw: str) -> str:
    if not raw:
        return ""
    raw = str(raw).strip()
    if raw.upper().startswith("ECHO-"):
        raw = raw[5:]
    if "-" not in raw and len(raw) >= 5:
        raw = raw[:4] + "-" + raw[4:]
    return raw


def _user_has_metals(df: pd.DataFrame) -> bool:
    if df is None or df.empty:
        return False

    cols = list(df.columns)
    # Common exact names across old/new pipelines
    exact = {
        "METALS_info",
        "lab_METALS_info",
        "METALS",
        "metals",
        "metals_info",
        "metals_info_en",
        "metals_info_orig",
        "elemental_concentrations",
        "elemental_concentrations_en",
        "elemental_concentrations_orig",
    }

    def _series_has_assignments(series: pd.Series) -> bool:
        s = (
            series.fillna("")
            .astype(str)
            .str.replace("<br>", ";", regex=False)
            .str.strip()
            .replace({"nan": "", "None": "", "NaN": "", "0": "", "0.0": ""})
        )
        # our metals blob looks like "Cu=12; Zn=5" etc.
        return s.str.contains("=", regex=False).any()

    # 1) Fast path: exact column name matches
    for c in cols:
        if c in exact and _series_has_assignments(df[c]):
            return True

    # 2) Fallback: any column whose name suggests metals/elemental info
    for c in cols:
        name = str(c).lower()
        if ("metals" in name) or ("elemental" in name and "concentration" in name):
            if _series_has_assignments(df[c]):
                return True
    return False


# --------------------------------------------------------------------------
# MinIO helpers + canonical download routes (proxy through Flask)
# --------------------------------------------------------------------------
try:
    from minio import Minio
except ImportError:
    Minio = None


def _get_minio_client():
    """
    Build a MinIO client from env / app config.
    We use it server-side to fetch objects and stream them to the user.
    """
    if Minio is None:
        return None

    endpoint = os.getenv("MINIO_ENDPOINT", "minio:9000")
    access_key = os.getenv("MINIO_ACCESS_KEY") or os.getenv("MINIO_ROOT_USER") or ""
    secret_key = os.getenv("MINIO_SECRET_KEY") or os.getenv("MINIO_ROOT_PASSWORD") or ""
    secure = False
    if endpoint.startswith("https://"):
        secure = True
        endpoint = endpoint[len("https://") :]
    elif endpoint.startswith("http://"):
        secure = False
        endpoint = endpoint[len("http://") :]

    if not access_key or not secret_key:
        return None

    return Minio(
        endpoint,
        access_key=access_key,
        secret_key=secret_key,
        secure=secure,
    )


def _stream_minio_canonical(obj_name: str):
    """
    Instead of redirecting the browser to MinIO (which you don't expose),
    we, the Flask app, download the object from MinIO and send it to the client.
    """
    client = _get_minio_client()
    bucket = os.getenv("MINIO_BUCKET", "echorepo-uploads")

    if client is None:
        abort(503, description="MinIO not configured on this instance")

    key = f"canonical/{obj_name}"
    try:
        resp = client.get_object(bucket, key)
    except Exception as e:
        abort(404, description=f"object not found in MinIO: {key}, error: {e}")

    # read into memory, then close the MinIO response
    data = resp.read()
    resp.close()
    resp.release_conn()

    mimetype = "application/zip" if obj_name.endswith(".zip") else "text/csv"
    return send_file(
        io.BytesIO(data),
        mimetype=mimetype,
        as_attachment=True,
        download_name=obj_name,
    )


def _upload_canonical_csvs_to_minio(csv_dict: dict[str, str], version_date: str):
    """
    Upload canonical CSVs (given as text) to MinIO, under:
      canonical/<version_date>/samples.csv
      canonical/<version_date>/sample_images.csv
      canonical/<version_date>/sample_parameters.csv
      canonical/latest/<same>

    csv_dict keys should be exactly "samples.csv", "sample_images.csv",
    "sample_parameters.csv".
    """
    client = _get_minio_client()
    if client is None:
        # MinIO not configured, silently skip
        return

    bucket = os.getenv("MINIO_BUCKET", "echorepo-uploads")

    try:
        if not client.bucket_exists(bucket):
            client.make_bucket(bucket)
    except Exception as e:
        logging.getLogger(__name__).error(f"Error ensuring MinIO bucket {bucket}: {e}")
        return

    for filename, csv_text in csv_dict.items():
        if not csv_text:
            continue

        data = csv_text.encode("utf-8")
        for prefix in (f"canonical/{version_date}/", "canonical/latest/"):
            obj_name = prefix + filename
            try:
                client.put_object(
                    bucket,
                    obj_name,
                    BytesIO(data),
                    length=len(data),
                    content_type="text/csv",
                )
            except Exception as e:
                logging.getLogger(__name__).error(f"Error uploading {obj_name} to MinIO: {e}")


def _upload_canonical_all_zip_to_minio(zip_bytes: bytes, version_date: str):
    """
    Upload all.zip snapshot to MinIO under:
      canonical/<version_date>/all.zip
      canonical/latest/all.zip
    """
    client = _get_minio_client()
    if client is None:
        logging.getLogger(__name__).warning("MinIO not configured; skipping all.zip upload")
        return

    bucket = os.getenv("MINIO_BUCKET", "echorepo-uploads")

    try:
        if not client.bucket_exists(bucket):
            client.make_bucket(bucket)
    except Exception as e:
        logging.getLogger(__name__).error(f"Error ensuring MinIO bucket {bucket}: {e}")
        return

    from io import BytesIO

    for prefix in (f"canonical/{version_date}/", "canonical/latest/"):
        obj_name = prefix + "all.zip"
        try:
            client.put_object(
                bucket,
                obj_name,
                BytesIO(zip_bytes),
                length=len(zip_bytes),
                content_type="application/zip",
            )
        except Exception as e:
            logging.getLogger(__name__).error(f"Error uploading {obj_name} to MinIO: {e}")


# --------- Search helpers
def _parse_sample_id_list(raw: str) -> list[str]:
    """
    Accept:
      - single: "AAAA-1111"
      - comma-separated: "AAAA-1111,BBBB-2222"
      - whitespace/newlines: "AAAA-1111  BBBB-2222"
    Returns cleaned non-empty tokens.
    """
    if not raw:
        return []
    # split on commas OR whitespace
    parts = re.split(r"[;,\s]+", raw.strip())
    return [p.strip() for p in parts if p.strip()]


def _get_latest_canonical_snapshot_date() -> str | None:
    """
    Return the latest YYYY-MM-DD for which canonical/<date>/all.zip exists
    in MinIO. Used by search export headers.
    """
    client = _get_minio_client()
    if client is None:
        return None

    bucket = os.getenv("MINIO_BUCKET", "echorepo-uploads")

    try:
        dates: set[str] = set()
        # Expect keys like: canonical/2025-12-08/all.zip
        for obj in client.list_objects(bucket, prefix="canonical/", recursive=True):
            parts = obj.object_name.split("/")
            if len(parts) == 3 and parts[0] == "canonical" and parts[2] == "all.zip":
                dates.add(parts[1])

        if not dates:
            return None
        # Lexicographically max works for ISO dates YYYY-MM-DD
        return max(dates)
    except Exception as e:
        logging.getLogger(__name__).error(f"Error listing canonical snapshots from MinIO: {e}")
        return None


# -------------- Endpoints for privacy acceptance ----------------


@web_bp.post("/privacy/accept")
@login_required
def privacy_accept():
    if not _env_true("PRIVACY_GATE", False):  # short-circuit the accept route when off
        return redirect(url_for("web.home"))
    repo_user_id = _get_repo_user_id_from_db()
    if not repo_user_id:
        abort(400, description="Cannot determine userId from database")

    # make sure the dir exists (handle the case when path has no dir part)
    dir_name = os.path.dirname(PRIVACY_CSV_PATH)
    if dir_name:
        os.makedirs(dir_name, exist_ok=True)

    _append_privacy_acceptance(repo_user_id)
    return redirect(url_for("web.home"))


# --------------------------------------------------------------------------
# Canonical downloads (now built from Postgres on demand)
# --------------------------------------------------------------------------


@web_bp.route("/download/canonical/samples.csv")
@login_required
def download_canonical_samples():
    g._analytics_extra = {
        "dataset": "canonical_samples",
        "file_name": "samples.csv",
        "kind": "canonical_export",
    }
    sql = """
        SELECT
            sample_id,
            timestamp_utc,
            lat,
            lon,
            country_code,
            location_accuracy_m,
            ph,
            organic_carbon_pct,
            earthworms_count,
            contamination_debris,
            contamination_plastic,
            contamination_other_orig,
            contamination_other_en,
            pollutants_count,
            soil_structure_orig,
            soil_structure_en,
            soil_texture_orig,
            soil_texture_en,
            observations_orig,
            observations_en,
            metals_info_orig,
            metals_info_en,
            collected_by,
            data_source,
            qa_status,
            licence
        FROM samples
        ORDER BY timestamp_utc DESC, sample_id
    """
    with get_pg_conn() as conn, conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute(sql)
        rows = cur.fetchall()

    if not rows:
        abort(404, description="No canonical samples found in database")

    df = pd.DataFrame(rows)

    # build CSV body into text
    buf_txt = io.StringIO()
    df.to_csv(buf_txt, index=False)
    body = buf_txt.getvalue()

    # simple "live export" header
    generated_at = datetime.utcnow().isoformat(timespec="seconds") + "Z"
    base_url = request.url_root.rstrip("/")

    # latest citable snapshot date (may be None if not yet created)
    snapshot_date = _get_latest_canonical_snapshot_date()
    if snapshot_date:
        snapshot_zip_url = f"{base_url}/download/canonical/{snapshot_date}/all.zip"
        snapshot_csv_url = f"{base_url}/download/canonical/{snapshot_date}/samples.csv"
        snapshot_lines = [
            f"# Latest citable snapshot (all.zip): {snapshot_zip_url}",
            f"# Latest citable snapshot (samples.csv): {snapshot_csv_url}",
        ]
    else:
        snapshot_lines = [
            "# Latest citable snapshot: (not available yet on this instance)",
        ]

    header = [
        "# ECHOrepo Canonical Dataset",
        "# File: samples.csv",
        f"# Generated at: {generated_at}",
        f"# Downloaded from: {base_url}/download/canonical/samples.csv",
        "# Note: This is a live export. For a fixed, citable snapshot, use the full canonical ZIP export below.",
        *snapshot_lines,
        "",
    ]
    csv_text = "\n".join(header) + body

    data = csv_text.encode("utf-8")
    buf = BytesIO(data)
    return send_file(
        buf,
        as_attachment=True,
        download_name="samples.csv",
        mimetype="text/csv",
    )


@web_bp.route("/download/canonical/sample_images.csv")
@login_required
def download_canonical_sample_images():
    g._analytics_extra = {
        "dataset": "canonical_sample_images",
        "file_name": "sample_images.csv",
        "kind": "canonical_export",
    }
    sql = """
        SELECT
            sample_id,
            country_code,
            image_id,
            image_url,
            image_description_orig,
            image_description_en,
            collected_by,
            timestamp_utc,
            licence
        FROM sample_images
        ORDER BY sample_id, image_id
    """
    with get_pg_conn() as conn, conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute(sql)
        rows = cur.fetchall()

    if not rows:
        abort(404, description="No canonical sample_images found in database")

    df = pd.DataFrame(rows)

    # build CSV body into text
    buf_txt = io.StringIO()
    df.to_csv(buf_txt, index=False)
    body = buf_txt.getvalue()

    # simple "live export" header
    generated_at = datetime.utcnow().isoformat(timespec="seconds") + "Z"
    base_url = request.url_root.rstrip("/")

    snapshot_date = _get_latest_canonical_snapshot_date()
    if snapshot_date:
        snapshot_zip_url = f"{base_url}/download/canonical/{snapshot_date}/all.zip"
        snapshot_csv_url = f"{base_url}/download/canonical/{snapshot_date}/sample_images.csv"
        snapshot_lines = [
            f"# Latest citable snapshot (all.zip): {snapshot_zip_url}",
            f"# Latest citable snapshot (sample_images.csv): {snapshot_csv_url}",
        ]
    else:
        snapshot_lines = [
            "# Latest citable snapshot: (not available yet on this instance)",
        ]

    header = [
        "# ECHOrepo Canonical Dataset",
        "# File: sample_images.csv",
        f"# Generated at: {generated_at}",
        f"# Downloaded from: {base_url}/download/canonical/sample_images.csv",
        "# Note: This is a live export. For a fixed, citable snapshot, use the full canonical ZIP export below.",
        *snapshot_lines,
        "",
    ]
    csv_text = "\n".join(header) + body

    data = csv_text.encode("utf-8")
    buf = BytesIO(data)
    return send_file(
        buf,
        as_attachment=True,
        download_name="sample_images.csv",
        mimetype="text/csv",
    )


@web_bp.route("/download/canonical/sample_parameters.csv")
@login_required
def download_canonical_sample_parameters():
    g._analytics_extra = {
        "dataset": "canonical_sample_parameters",
        "file_name": "sample_parameters.csv",
        "kind": "canonical_export",
    }
    sql = """
        SELECT
            sample_id,
            country_code,
            parameter_code,
            parameter_name,
            value,
            uom,
            analysis_method,
            analysis_date,
            lab_id,
            created_by,
            licence,
            parameter_uri
        FROM sample_parameters
        ORDER BY sample_id, parameter_code
    """
    with get_pg_conn() as conn, conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute(sql)
        rows = cur.fetchall()

    if not rows:
        abort(404, description="No canonical sample_parameters found in database")

    df = pd.DataFrame(rows)
    df = _drop_oxide_rows(df)

    # build CSV body into text
    buf_txt = io.StringIO()
    df.to_csv(buf_txt, index=False)
    body = buf_txt.getvalue()

    # simple "live export" header
    generated_at = datetime.utcnow().isoformat(timespec="seconds") + "Z"
    base_url = request.url_root.rstrip("/")

    snapshot_date = _get_latest_canonical_snapshot_date()
    if snapshot_date:
        snapshot_zip_url = f"{base_url}/download/canonical/{snapshot_date}/all.zip"
        snapshot_csv_url = f"{base_url}/download/canonical/{snapshot_date}/sample_parameters.csv"
        snapshot_lines = [
            f"# Latest citable snapshot (all.zip): {snapshot_zip_url}",
            f"# Latest citable snapshot (sample_parameters.csv): {snapshot_csv_url}",
        ]
    else:
        snapshot_lines = [
            "# Latest citable snapshot: (not available yet on this instance)",
        ]

    header = [
        "# ECHOrepo Canonical Dataset",
        "# File: sample_parameters.csv",
        f"# Generated at: {generated_at}",
        f"# Downloaded from: {base_url}/download/canonical/sample_parameters.csv",
        "# Note: This is a live export. For a fixed, citable snapshot, use the full canonical ZIP export below.",
        *snapshot_lines,
        "",
    ]
    csv_text = "\n".join(header) + body

    data = csv_text.encode("utf-8")
    buf = BytesIO(data)
    return send_file(
        buf,
        as_attachment=True,
        download_name="sample_parameters.csv",
        mimetype="text/csv",
    )


@web_bp.route("/download/canonical/sample_biodiversity.csv")
@login_required
def download_canonical_sample_biodiversity():
    g._analytics_extra = {
        "dataset": "canonical_sample_biodiversity",
        "file_name": "sample_biodiversity.csv",
        "kind": "canonical_export",
    }

    sql = """
        SELECT
            sample_id,
            marker,
            otu_id,
            count,
            taxa,
            uploaded_at,
            uploaded_by,
            source_file
        FROM sample_otu_counts
        ORDER BY sample_id, marker, otu_id
    """

    with get_pg_conn() as conn, conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute(sql)
        rows = cur.fetchall()

    if not rows:
        abort(404, description="No biodiversity rows found in database")

    df = pd.DataFrame(rows)

    buf_txt = io.StringIO()
    df.to_csv(buf_txt, index=False)
    body = buf_txt.getvalue()

    generated_at = datetime.utcnow().isoformat(timespec="seconds") + "Z"
    base_url = request.url_root.rstrip("/")

    header = [
        "# ECHOrepo Canonical Dataset",
        "# File: sample_biodiversity.csv",
        f"# Generated at: {generated_at}",
        f"# Downloaded from: {base_url}/download/canonical/sample_biodiversity.csv",
        "# Description: Biodiversity OTU abundance data per sample.",
        "",
    ]
    csv_text = "\n".join(header) + body

    data = csv_text.encode("utf-8")
    buf = BytesIO(data)
    return send_file(
        buf,
        as_attachment=True,
        download_name="sample_biodiversity.csv",
        mimetype="text/csv",
    )


@web_bp.route("/download/canonical/all.zip")
@login_required
def download_canonical_zip():
    g._analytics_extra = {
        "dataset": "canonical_all",
        "file_name": "all_canonical.zip",
        "kind": "canonical_export",
    }

    # Version date for this canonical snapshot
    version_date = datetime.utcnow().date().isoformat()
    # Base URL for building reference links in headers
    base_url = request.url_root.rstrip("/")

    # We'll collect the final CSV text for MinIO upload
    csv_contents: dict[str, str] = {}

    mem = BytesIO()
    with zipfile.ZipFile(mem, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        # 1) samples.csv
        with get_pg_conn() as conn, conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("SELECT * FROM samples ORDER BY timestamp_utc DESC, sample_id")
            df_samples = pd.DataFrame(cur.fetchall())

        buf1 = io.StringIO()
        df_samples.to_csv(buf1, index=False)
        body_samples = buf1.getvalue()

        header_samples = [
            "# ECHOrepo Canonical Dataset",
            "# File: samples.csv",
            f"# Version date: {version_date}",
            f"# Version URL: {base_url}/download/canonical/{version_date}/samples.csv",
            f"# Latest canonical: {base_url}/download/canonical/samples.csv",
            "# Description: Canonical sample-level data (locations, pH, texture, structure, etc.).",
            "",
        ]
        csv_samples = "\n".join(header_samples) + body_samples

        csv_contents["samples.csv"] = csv_samples
        zf.writestr("samples.csv", csv_samples)

        # 2) sample_images.csv
        with get_pg_conn() as conn, conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("SELECT * FROM sample_images ORDER BY sample_id, image_id")
            df_imgs = pd.DataFrame(cur.fetchall())

        buf2 = io.StringIO()
        df_imgs.to_csv(buf2, index=False)
        body_imgs = buf2.getvalue()

        header_imgs = [
            "# ECHOrepo Canonical Dataset",
            "# File: sample_images.csv",
            f"# Version date: {version_date}",
            f"# Version URL: {base_url}/download/canonical/{version_date}/sample_images.csv",
            f"# Latest canonical: {base_url}/download/canonical/sample_images.csv",
            "# Description: Canonical image metadata linked to samples (IDs, URLs, descriptions).",
            "",
        ]
        csv_imgs = "\n".join(header_imgs) + body_imgs

        csv_contents["sample_images.csv"] = csv_imgs
        zf.writestr("sample_images.csv", csv_imgs)

        # 3) sample_parameters.csv
        with get_pg_conn() as conn, conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("SELECT * FROM sample_parameters ORDER BY sample_id, parameter_code")
            df_params = pd.DataFrame(cur.fetchall())
            df_params = _drop_oxide_rows(df_params)

        buf3 = io.StringIO()
        df_params.to_csv(buf3, index=False)
        body_params = buf3.getvalue()

        header_params = [
            "# ECHOrepo Canonical Dataset",
            "# File: sample_parameters.csv",
            f"# Version date: {version_date}",
            f"# Version URL: {base_url}/download/canonical/{version_date}/sample_parameters.csv",
            f"# Latest canonical: {base_url}/download/canonical/sample_parameters.csv",
            "# Description: Canonical laboratory parameters (metals, nutrients, etc.) per sample.",
            "",
        ]
        csv_params = "\n".join(header_params) + body_params

        csv_contents["sample_parameters.csv"] = csv_params
        zf.writestr("sample_parameters.csv", csv_params)

        with get_pg_conn() as conn, conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                SELECT
                    sample_id,
                    marker,
                    otu_id,
                    count,
                    taxa,
                    uploaded_at,
                    uploaded_by,
                    source_file
                FROM sample_otu_counts
                ORDER BY sample_id, marker, otu_id
            """)
            df_biodiv = pd.DataFrame(cur.fetchall())

        buf4 = io.StringIO()
        df_biodiv.to_csv(buf4, index=False)
        body_biodiv = buf4.getvalue()

        header_biodiv = [
            "# ECHOrepo Canonical Dataset",
            "# File: sample_biodiversity.csv",
            f"# Version date: {version_date}",
            f"# Version URL: {base_url}/download/canonical/{version_date}/sample_biodiversity.csv",
            f"# Latest canonical: {base_url}/download/canonical/sample_biodiversity.csv",
            "# Description: Biodiversity OTU abundance data per sample.",
            "",
        ]
        csv_biodiv = "\n".join(header_biodiv) + body_biodiv

        csv_contents["sample_biodiversity.csv"] = csv_biodiv
        zf.writestr("sample_biodiversity.csv", csv_biodiv)

    _upload_canonical_csvs_to_minio(csv_contents, version_date)

    # Prepare ZIP bytes once
    mem.seek(0)
    zip_bytes = mem.getvalue()

    # Also upload all.zip snapshot to MinIO
    _upload_canonical_all_zip_to_minio(zip_bytes, version_date)

    # Send to user
    return send_file(
        BytesIO(zip_bytes),
        as_attachment=True,
        download_name="all_canonical.zip",
        mimetype="application/zip",
    )


# --------------------------------------------------------------------------
# main UI
# --------------------------------------------------------------------------
@web_bp.get("/", endpoint="landing")
def landing():
    logged_in = bool(session.get("user") or (session.get("kc") or {}).get("profile"))
    if logged_in:
        return redirect(url_for("web.home"))  # now /my
    return redirect(url_for("web.explore"))


@web_bp.get("/my", endpoint="home")
@login_required
def home():
    # who is this
    user_key = session.get("user") or session.get("kc", {}).get("profile", {}).get("email")
    if not user_key:
        return redirect(url_for("auth.login"))

    # NEW: lab upload permission flag
    can_upload = can_upload_lab_data(user_key)

    # privacy gate
    privacy_user_id = _get_repo_user_id_from_db()
    # 1) default from env (now default is OFF)
    privacy_gate_default = _env_true("PRIVACY_GATE", False)

    # 2) enable/disable URL override via .env
    privacy_override_enabled = _env_true("PRIVACY_DEBUG_OVERRIDE", False)

    if privacy_override_enabled:
        # optional per-request override: ?privacy=on / ?privacy=off
        override = (request.args.get("privacy") or "").strip().lower()
        if override in {"on", "1", "true", "yes", "y"}:
            privacy_gate_on = True
        elif override in {"off", "0", "false", "no", "n"}:
            privacy_gate_on = False
        else:
            privacy_gate_on = privacy_gate_default
    else:
        # ignore ?privacy=... when override is disabled
        privacy_gate_on = privacy_gate_default

    needs_privacy = privacy_gate_on and not _has_accepted_privacy(privacy_user_id or "")

    # user data (samples)
    df = query_user_df(user_key)

    # decide survey visibility from canonical lab rows, not from samples DF
    has_lab_results = _user_has_lab_results(df) or _user_has_metals_legacy(df)

    # only for display, now drop oxide-like columns (harmless on samples DF)
    df = _drop_oxide_columns_from_df(df)

    i18n = {"labels": build_i18n_labels(_js_base_labels())}

    # try to get "real" internal user id from data
    kc_user_id = None
    if not df.empty:
        for col in ("userId", "user_id", "kc_user_id"):
            if col in df.columns:
                kc_user_id = df[col].dropna().astype(str).iloc[0].strip()
                break

    # survey url
    survey_user_id = kc_user_id or user_key
    survey_url = _build_sosci_url(survey_user_id)

    # only show survey if user actually has metals-like data
    has_metals = has_lab_results
    show_survey = bool(survey_url) and has_metals

    # --- survey override via URL (?survey=on/off), for testing ---
    survey_override_enabled = _env_true("SURVEY_DEBUG_OVERRIDE", False)

    if survey_override_enabled:
        survey_override = (request.args.get("survey") or "").strip().lower()
        if survey_override in {"on", "1", "true", "yes", "y"}:
            show_survey = bool(survey_url)  # ignore has_metals here
        elif survey_override in {"off", "0", "false", "no", "n"}:
            show_survey = False
    # else: ignore any ?survey=... and keep show_survey as computed

    # EMPTY CASE ------------------------------------------------------------
    if df.empty:
        return render_template(
            "results.html",
            issue_count=0,
            user_key=user_key,
            kc_user_id=kc_user_id,
            columns=[],
            table_html="<p>No data available for this user.</p>",
            jitter_m=int(settings.MAX_JITTER_METERS),
            lat_col=settings.LAT_COL,
            lon_col=settings.LON_COL,
            I18N=i18n,
            survey_url=survey_url,
            show_survey=show_survey,
            # privacy
            needs_privacy=needs_privacy,
            privacy_version=PRIVACY_VERSION,
            # pass flag to template
            can_upload_lab_data=can_upload,
            current_locale=str(get_locale() or "en"),
        )

    # NON-EMPTY: data issues ------------------------------------------------
    defaults = find_default_coord_rows(df)
    mism = select_country_mismatches(df)
    issue_count = len(defaults) + len(mism)

    # HTML copy — prettify timestamp column
    df_html = df.copy()
    if "fs_createdAt" in df_html.columns:
        df_html["fs_createdAt"] = (
            df_html["fs_createdAt"]
            .fillna("")
            .astype(str)
            .str.split(".")
            .str[0]
            .str.replace("T", " ", regex=False)
            .str.replace("Z", "", regex=False)
        )
        cols = list(df_html.columns)
        if "fs_createdAt" in cols:
            if "sampleId" in cols:
                cols.insert(cols.index("sampleId") + 1, cols.pop(cols.index("fs_createdAt")))
            else:
                cols.insert(0, cols.pop(cols.index("fs_createdAt")))
            df_html = df_html[cols]

    table_html = make_table_html(df_html)

    return render_template(
        "results.html",
        issue_count=issue_count,
        user_key=user_key,
        kc_user_id=kc_user_id,
        columns=list(df.columns),
        table_html=table_html,
        jitter_m=int(settings.MAX_JITTER_METERS),
        lat_col=settings.LAT_COL,
        lon_col=settings.LON_COL,
        I18N=i18n,
        survey_url=survey_url,
        show_survey=show_survey,
        # privacy flags for the modal
        needs_privacy=needs_privacy,
        privacy_version=PRIVACY_VERSION,
        # NEW: pass flag to template
        can_upload_lab_data=can_upload,
        current_locale=str(get_locale() or "en"),
    )


# --------------------------------------------------------------------------
# misc routes
# --------------------------------------------------------------------------
@web_bp.get("/i18n/labels")
@login_required
def i18n_labels():
    return jsonify({"labels": build_i18n_labels(_js_base_labels())})


@web_bp.get("/labels")
@login_required
def labels_json():
    loc_raw = request.args.get("locale") or str(get_locale() or "en")
    loc = _canon_locale(loc_raw)
    payload = {
        "labels": make_labels(loc),
        "by_msgid": get_overrides_msgid(loc) or {},
    }
    return jsonify(payload)


@web_bp.post("/download/csv")
@login_required
def download_csv():
    g._analytics_extra = {
        "dataset": "user_data",
        "file_name": "user_data.csv",
        "kind": "user_export",
    }
    user_key = (request.form.get("user_key") or "").strip()
    if not user_key:
        abort(400)
    df = query_user_df(user_key)
    if df.empty:
        abort(404)
    df = strip_orig_cols(df)
    df = _drop_oxide_columns_from_df(df)

    buf = BytesIO()
    df.to_csv(buf, index=False)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"{user_key}_data.csv",
        mimetype="text/csv",
    )


@web_bp.post("/download/xlsx")
@login_required
def download_xlsx():
    g._analytics_extra = {
        "dataset": "user_data",
        "file_name": "user_data.xlsx",
        "kind": "user_export",
    }
    user_key = (request.form.get("user_key") or "").strip()
    if not user_key:
        abort(400)
    df = query_user_df(user_key)
    if df.empty:
        abort(404)
    df = strip_orig_cols(df)
    df = _drop_oxide_columns_from_df(df)

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="xlsxwriter") as w:
        df.to_excel(w, index=False, sheet_name="data")
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"{user_key}_data.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@web_bp.get("/download/all_csv")
@login_required
def download_all_csv():
    g._analytics_extra = {
        "dataset": "all_data",
        "file_name": "all_data.csv",
        "kind": "user_export",
    }
    p = pathlib.Path(settings.INPUT_CSV)
    if not p.exists() or not p.is_file():
        abort(404, description="Full CSV not found on server.")

    pii_cols = {
        "userId",
        "user_id",
        "email",
        getattr(settings, "USER_KEY_COLUMN", None),
    }
    pii_cols = {c.lower() for c in pii_cols if c}

    try:
        df_all = pd.read_csv(p, dtype=str, keep_default_na=False, low_memory=False)
        df_all = strip_orig_cols(df_all)
        df_all = _drop_oxide_columns_from_df(df_all)

        drop_these = [col for col in df_all.columns if col.lower() in pii_cols]
        if drop_these:
            df_all = df_all.drop(columns=drop_these, errors="ignore")

        buf = BytesIO()
        df_all.to_csv(buf, index=False)
        buf.seek(0)
        return send_file(
            buf,
            as_attachment=True,
            download_name="echorepo_all_samples.csv",
            mimetype="text/csv",
        )
    except Exception:
        import csv

        buf = BytesIO()
        with p.open("r", encoding="utf-8", newline="") as f_in:
            reader = csv.reader(f_in)
            rows = list(reader)
            if not rows:
                abort(404, description="CSV is empty")

            header = rows[0]
            keep_idx = []
            for i, name in enumerate(header):
                if name.lower() in pii_cols:
                    continue
                if _looks_like_oxide(name) or _looks_like_oxide(re.split(r"[_\s/\-]+", name)[-1]):
                    continue
                keep_idx.append(i)

            writer = csv.writer(buf)
            writer.writerow([header[i] for i in keep_idx])
            for row in rows[1:]:
                writer.writerow([row[i] for i in keep_idx])

        buf.seek(0)
        return send_file(
            buf,
            as_attachment=True,
            download_name="echorepo_all_samples.csv",
            mimetype="text/csv",
        )


@web_bp.get("/download/sample_csv")
@login_required
def download_sample_csv():
    g._analytics_extra = {
        "dataset": "user_data",
        "file_name": "sample_data.csv",
        "kind": "user_export",
    }
    sample_id = (request.args.get("sampleId") or "").strip()
    if not sample_id:
        abort(400, description="sampleId is required")
    df = query_sample(sample_id)
    if df.empty:
        abort(404, description="Sample not found")

    user_key = session.get("user")
    is_owner = False
    try:
        if (
            settings.USER_KEY_COLUMN in df.columns
            and (df[settings.USER_KEY_COLUMN] == user_key).any()
        ):
            is_owner = True
        if "userId" in df.columns and (df["userId"] == user_key).any():
            is_owner = True
    except Exception:
        pass

    if not is_owner:
        for pii in ("email", "userId"):
            if pii in df.columns:
                df = df.drop(columns=[pii])

    buf = BytesIO()
    df.to_csv(buf, index=False)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"sample_{sample_id}.csv",
        mimetype="text/csv",
    )


# --------------------------------------------------------------------------
# lab upload
# --------------------------------------------------------------------------
@web_bp.post("/lab-import")
@login_required
def lab_import():
    user_key = session.get("user") or session.get("kc", {}).get("profile", {}).get("email")
    if not can_upload_lab_data(user_key):
        abort(403, description="Not authorised to upload lab data")

    file = request.files.get("file")
    if not file:
        abort(400, description="No file uploaded")

    kc_profile = (session.get("kc") or {}).get("profile") or {}
    uploader_id = kc_profile.get("id") or kc_profile.get("sub") or session.get("user") or "unknown"

    filename = file.filename or ""
    try:
        if filename.lower().endswith(".xlsx"):
            df = pd.read_excel(file)
        else:
            try:
                df = pd.read_csv(file, sep="\t")
            except Exception:
                file.stream.seek(0)
                df = pd.read_csv(file)
    except Exception as e:
        abort(400, description=f"Cannot read file: {e}")

    if df.empty:
        abort(400, description="Uploaded file has no rows")

    db_path = settings.SQLITE_PATH
    if not os.path.exists(db_path):
        abort(500, description=f"SQLite database not found at {db_path}")

    conn = sqlite3.connect(db_path)
    _ensure_lab_enrichment(conn)
    cur = conn.cursor()
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS lab_enrichment (
          qr_code    TEXT NOT NULL,
          param      TEXT NOT NULL,
          value      TEXT,
          unit       TEXT,
          user_id    TEXT,
          raw_row    TEXT,
          updated_at TEXT DEFAULT (datetime('now')),
          PRIMARY KEY (qr_code, param)
        )
        """
    )

    fieldnames = list(df.columns)

    for __idx, row in df.iterrows():
        raw_dict = row.to_dict()
        qr = _normalize_qr(raw_dict.get("ID") or raw_dict.get("id") or "")
        if not qr:
            continue

        clean_raw = {k: ("" if pd.isna(v) else v) for k, v in raw_dict.items()}
        raw_json = json.dumps(clean_raw, ensure_ascii=False)

        for idx, col in enumerate(fieldnames):
            if col in ("ID", "id"):
                continue
            val = row.get(col)
            if pd.isna(val) or val == "":
                continue
            if str(col).lower().startswith("unit"):
                continue

            param = str(col).strip()
            unit = ""

            if idx + 1 < len(fieldnames):
                maybe_unit_col = fieldnames[idx + 1]
                if str(maybe_unit_col).lower().startswith("unit"):
                    uval = row.get(maybe_unit_col)
                    if not pd.isna(uval):
                        unit = str(uval).strip()

            # 1) store the raw value (oxide or otherwise)
            cur.execute(
                """
                INSERT INTO lab_enrichment (qr_code, param, value, unit, user_id, raw_row, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, datetime('now'))
                ON CONFLICT(qr_code, param) DO UPDATE SET
                  value=excluded.value,
                  unit=excluded.unit,
                  user_id=excluded.user_id,
                  raw_row=excluded.raw_row,
                  updated_at=datetime('now')
                """,
                (qr, param, str(val), unit, uploader_id, raw_json),
            )

            # 2) if this is an oxide like 'K2O', also store elemental 'K'
            conv = _oxide_to_metal(param, val)

            if conv is not None:
                metal_param, metal_val = conv
                cur.execute(
                    """
                    INSERT INTO lab_enrichment (qr_code, param, value, unit, user_id, raw_row, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, datetime('now'))
                    ON CONFLICT(qr_code, param) DO UPDATE SET
                      value=excluded.value,
                      unit=excluded.unit,
                      user_id=excluded.user_id,
                      raw_row=excluded.raw_row,
                      updated_at=datetime('now')
                    """,
                    (qr, metal_param, str(metal_val), unit, uploader_id, raw_json),
                )

    conn.commit()
    conn.close()

    return redirect(url_for("web.home"))


@web_bp.get("/lab-upload")
@login_required
def lab_upload():
    user_key = session.get("user") or session.get("kc", {}).get("profile", {}).get("email")
    if not can_upload_lab_data(user_key):
        abort(403, description="Not authorised to upload lab data")
    return render_template("lab_upload.html")


@web_bp.post("/lab-upload")
@login_required
def lab_upload_post():
    user_key = session.get("user") or session.get("kc", {}).get("profile", {}).get("email")
    if not can_upload_lab_data(user_key):
        abort(403, description="Not authorised to upload lab data")

    file = request.files.get("file")
    if not file:
        abort(400, "No file")
    else:
        g._analytics_extra = {
            "upload_type": "lab_csv",
            "filename": file.filename,
            "content_length": request.content_length,
        }
    return redirect(url_for("web.home"))


@web_bp.route("/search", endpoint="search_samples", methods=["GET"])
def search_samples():
    q = request.args.get("q", "").strip()

    base = base_labels()
    i18n = {
        "labels": build_i18n_labels(base),
        "by_msgid": {},
    }
    # ----- read filters from querystring -----
    criteria = {
        "sample_id": (request.args.get("sample_id") or "").strip(),
        "country_code": (request.args.get("country_code") or "").strip().upper(),
        "ph_min": (request.args.get("ph_min") or "").strip(),
        "ph_max": (request.args.get("ph_max") or "").strip(),
        "date_from": (request.args.get("date_from") or "").strip(),
        "date_to": (request.args.get("date_to") or "").strip(),
    }
    fmt = (request.args.get("format") or "").lower()

    # pagination (for HTML)
    page = max(int(request.args.get("page", 1)), 1)
    per_page = 50
    offset = (page - 1) * per_page

    # helper to build WHERE + params
    def _build_where(criteria):
        where = ["1=1"]
        params = []
        if criteria["sample_id"]:
            tokens = _parse_sample_id_list(criteria["sample_id"])

            if len(tokens) == 1:
                # keep old behaviour: substring match
                where.append("sample_id ILIKE %s")
                params.append(f"%{tokens[0]}%")
            else:
                # multiple: match ANY of the tokens (OR)
                ors = []
                for t in tokens:
                    ors.append("sample_id ILIKE %s")
                    params.append(f"%{t}%")
                where.append("(" + " OR ".join(ors) + ")")
        if criteria["country_code"]:
            where.append("country_code = %s")
            params.append(criteria["country_code"])
        if criteria["ph_min"]:
            where.append("ph >= %s")
            params.append(float(criteria["ph_min"]))
        if criteria["ph_max"]:
            where.append("ph <= %s")
            params.append(float(criteria["ph_max"]))
        if criteria["date_from"]:
            where.append("timestamp_utc >= %s")
            params.append(criteria["date_from"])
        if criteria["date_to"]:
            where.append("timestamp_utc <= %s")
            params.append(criteria["date_to"])
        return " AND ".join(where), params

    where_sql, params = _build_where(criteria)

    def _is_logged_in() -> bool:
        return bool(session.get("user") or (session.get("kc") or {}).get("profile"))

    if fmt == "zip" and not _is_logged_in():
        # show search page as normal, but deny export
        abort(403, description="Please sign in to export ZIP files.")

    # ---- special case: export as ZIP ----
    if fmt == "zip":
        query_string = request.query_string.decode("utf-8")
        generated = datetime.utcnow().isoformat() + "Z"

        # Must match the date used when uploading all.zip in canonical_all_zip
        base_url = request.url_root.rstrip("/")
        version_date_current = date.today().isoformat()
        snapshot_date = _get_latest_canonical_snapshot_date()
        version_date = snapshot_date or version_date_current

        snapshot_url = f"{base_url}/download/canonical/{version_date}/all.zip"

        # Use the full canonical column set from data_api.py
        cols_sql = ", ".join(CANONICAL_SAMPLE_COLS)
        sql_all = f"""
            SELECT {cols_sql}
            FROM samples
            WHERE {where_sql}
            ORDER BY timestamp_utc DESC
        """

        with get_pg_conn() as conn, conn.cursor() as cur:
            # 1) get all matching samples
            cur.execute(sql_all, params)
            samples = cur.fetchall()

            # sample_id is the first canonical column
            sample_ids = [
                row[CANONICAL_SAMPLE_COLS.index("sample_id")]
                for row in samples
                if row[CANONICAL_SAMPLE_COLS.index("sample_id")]
            ]
            # to avoid "IN ()" when no results
            if not sample_ids:
                # return empty zip
                mem = io.BytesIO()
                with zipfile.ZipFile(mem, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
                    zf.writestr(
                        "samples_filtered.csv",
                        "sample_id,timestamp_utc,country_code,ph,lat,lon,collected_by\n",
                    )
                    zf.writestr(
                        "sample_images_filtered.csv",
                        "sample_id,country_code,image_id,image_url,image_description_orig,image_description_en,collected_by,timestamp_utc,licence\n",
                    )
                    zf.writestr(
                        "sample_parameters_filtered.csv",
                        "sample_id,country_code,parameter_code,parameter_name,value,uom,analysis_method,analysis_date,lab_id,created_by,licence,parameter_uri\n",
                    )
                    zf.writestr(
                        "sample_biodiversity_filtered.csv",
                        "sample_id,marker,otu_id,count,taxa,uploaded_at,uploaded_by,source_file\n",
                    )
                mem.seek(0)
                return send_file(
                    mem,
                    as_attachment=True,
                    download_name="search_export.zip",
                    mimetype="application/zip",
                )

            # 2) fetch images for those sample_ids
            sql_imgs = """
                SELECT sample_id, country_code, image_id, image_url,
                       image_description_orig, image_description_en,
                       collected_by, timestamp_utc, licence
                FROM sample_images
                WHERE sample_id = ANY(%s)
                ORDER BY sample_id, image_id
            """
            cur.execute(sql_imgs, (sample_ids,))
            images = cur.fetchall()

            # 3) fetch parameters for those sample_ids
            sql_params = """
                SELECT sample_id, country_code, parameter_code, parameter_name,
                       value, uom, analysis_method, analysis_date,
                       lab_id, created_by, licence, parameter_uri
                FROM sample_parameters
                WHERE sample_id = ANY(%s)
                ORDER BY sample_id, parameter_code
            """
            cur.execute(sql_params, (sample_ids,))
            params_rows = cur.fetchall()

            # 4) fetch biodiversity data for those sample_ids (if any)
            sql_biodiv = """
                SELECT sample_id, marker, otu_id, count, taxa, uploaded_at, uploaded_by, source_file
                FROM sample_otu_counts
                WHERE sample_id = ANY(%s)
                ORDER BY sample_id, marker, otu_id
            """
            cur.execute(sql_biodiv, (sample_ids,))
            biodiv_rows = cur.fetchall()

        # ---- drop oxides here (list-of-tuples: keep columns 2=code, 3=name) ----
        def _row_is_oxide(t):
            code = (t[2] or "").strip()
            name = (t[3] or "").strip()
            return _looks_like_oxide(code) or _looks_like_oxide(name)

        params_rows = [r for r in params_rows if not _row_is_oxide(r)]

        # build zip in-memory
        mem = io.BytesIO()
        with zipfile.ZipFile(mem, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
            # samples_filtered.csv
            out1 = io.StringIO()

            # ----- METADATA HEADER -----
            out1.write("# ECHOrepo Filtered Dataset\n")
            out1.write("# Source table: samples (filtered subset)\n")
            out1.write(f"# Download full dataset snapshot: {snapshot_url}\n")
            out1.write(f"# Generated at: {generated}\n")
            out1.write(f"# Query: {query_string}\n")
            out1.write(
                "# Note: This is a filtered export for user inspection. It is NOT a stable or citable dataset.\n"
            )
            out1.write("\n")

            w1 = csv.writer(out1)

            # header: all canonical sample columns
            w1.writerow(CANONICAL_SAMPLE_COLS)

            # rows come from SELECT {cols_sql} in the same order
            for r in samples:
                w1.writerow(r)

            zf.writestr("samples_filtered.csv", out1.getvalue())

            # sample_images_filtered.csv
            out2 = io.StringIO()

            # ----- METADATA HEADER -----
            out2.write("# ECHOrepo Filtered Dataset\n")
            out2.write("# Source table: sample_images (filtered subset)\n")
            out2.write(f"# Download full dataset snapshot: {snapshot_url}\n")
            out2.write(f"# Generated at: {generated}\n")
            out2.write(f"# Query: {query_string}\n")
            out2.write(
                "# Note: This is a filtered export for user inspection. It is NOT a stable or citable dataset.\n"
            )
            out2.write("\n")

            w2 = csv.writer(out2)
            w2.writerow(
                [
                    "sample_id",
                    "country_code",
                    "image_id",
                    "image_url",
                    "image_description_orig",
                    "image_description_en",
                    "collected_by",
                    "timestamp_utc",
                    "licence",
                ]
            )
            for r in images:
                w2.writerow(r)
            zf.writestr("sample_images_filtered.csv", out2.getvalue())

            # sample_parameters_filtered.csv
            out3 = io.StringIO()

            # ----- METADATA HEADER -----
            out3.write("# ECHOrepo Filtered Dataset\n")
            out3.write("# Source table: samples (filtered subset)\n")
            out3.write(f"# Download full dataset snapshot: {snapshot_url}\n")
            out3.write(f"# Generated at: {generated}\n")
            out3.write(f"# Query: {query_string}\n")
            out3.write(
                "# Note: This is a filtered export for user inspection. It is NOT a stable or citable dataset.\n"
            )
            out3.write("\n")

            w3 = csv.writer(out3)
            w3.writerow(
                [
                    "sample_id",
                    "country_code",
                    "parameter_code",
                    "parameter_name",
                    "value",
                    "uom",
                    "analysis_method",
                    "analysis_date",
                    "lab_id",
                    "created_by",
                    "licence",
                    "parameter_uri",
                ]
            )
            for r in params_rows:
                w3.writerow(r)
            zf.writestr("sample_parameters_filtered.csv", out3.getvalue())

            out4 = io.StringIO()

            out4.write("# ECHOrepo Filtered Dataset\n")
            out4.write("# Source table: sample_otu_counts (filtered subset)\n")
            out4.write(f"# Download full dataset snapshot: {snapshot_url}\n")
            out4.write(f"# Generated at: {generated}\n")
            out4.write(f"# Query: {query_string}\n")
            out4.write(
                "# Note: This is a filtered export for user inspection. It is NOT a stable or citable dataset.\n"
            )
            out4.write("\n")

            w4 = csv.writer(out4)
            w4.writerow(
                [
                    "sample_id",
                    "marker",
                    "otu_id",
                    "count",
                    "taxa",
                    "uploaded_at",
                    "uploaded_by",
                    "source_file",
                ]
            )
            for r in biodiv_rows:
                w4.writerow(r)

            zf.writestr("sample_biodiversity_filtered.csv", out4.getvalue())
        mem.seek(0)
        return send_file(
            mem,
            as_attachment=True,
            download_name="search_export.zip",
            mimetype="application/zip",
        )

    # ---- normal HTML search with pagination ----
    total_rows = 0
    rows = []

    with get_pg_conn() as conn, conn.cursor(cursor_factory=RealDictCursor) as cur:
        # count
        cur.execute(f"SELECT COUNT(*) FROM samples WHERE {where_sql}", params)
        total_rows = cur.fetchone()["count"]

        # page
        cur.execute(
            f"""
            SELECT sample_id, timestamp_utc, country_code, ph, lat, lon, collected_by
            FROM samples
            WHERE {where_sql}
            ORDER BY timestamp_utc DESC
            LIMIT %s OFFSET %s
            """,
            params + [per_page, offset],
        )
        rows = cur.fetchall()

    total_pages = max((total_rows + per_page - 1) // per_page, 1)

    # add analytics extras
    g._analytics_extra = {
        "search_query": q,
        # maybe:
        # "search_source": "web",
        # "filters": list(request.args.keys()),
    }

    return render_template(
        "search.html",
        criteria=criteria,
        rows=rows,
        page=page,
        total_pages=total_pages,
        total_rows=total_rows,
        I18N=i18n,
    )


@web_bp.route("/admin/usage")
@login_required  # TODO: restrict to admins if possible
def usage_dashboard():
    """
    Simple usage statistics dashboard backed by Postgres usage_events table.
    Shows:
      - summary counts for the period
      - per-day stats
      - downloads per dataset
      - events per type
    """

    # Period control via ?days=...
    try:
        days = int(request.args.get("days", 30))
    except ValueError:
        days = 30
    days = max(1, min(days, 365))  # clamp to [1, 365]

    end_ts = datetime.utcnow()
    start_ts = end_ts - timedelta(days=days)

    with get_pg_conn() as conn, conn.cursor(cursor_factory=RealDictCursor) as cur:
        # 1) Summary for the period
        cur.execute(
            """
            SELECT
              COUNT(*) AS total_requests,
              COUNT(DISTINCT user_id) AS unique_users,
              COUNT(DISTINCT ip_hash) AS unique_visitors,
              COUNT(*) FILTER (WHERE event_type = 'download') AS total_downloads
            FROM usage_events
            WHERE ts >= %s AND ts < %s
            """,
            (start_ts, end_ts),
        )
        summary = cur.fetchone() or {
            "total_requests": 0,
            "unique_users": 0,
            "unique_visitors": 0,
            "total_downloads": 0,
        }

        # 2) Per-day stats
        cur.execute(
            """
            SELECT
              date_trunc('day', ts)::date AS day,
              COUNT(*) AS requests,
              COUNT(DISTINCT user_id) AS unique_users,
              COUNT(DISTINCT ip_hash) AS unique_visitors,
              COUNT(*) FILTER (WHERE event_type = 'download') AS downloads
            FROM usage_events
            WHERE ts >= %s AND ts < %s
            GROUP BY day
            ORDER BY day;
            """,
            (start_ts, end_ts),
        )
        daily = cur.fetchall() or []

        # 3) Downloads per dataset (for this period)
        cur.execute(
            """
            SELECT
              date_trunc('month', ts)::date AS month,
              COALESCE(
                extra->>'dataset',
                extra->>'file',
                path
              ) AS dataset,
              COUNT(*) AS downloads
            FROM usage_events
            WHERE event_type = 'download'
              AND ts >= %s AND ts < %s
            GROUP BY month, dataset
            ORDER BY month DESC, downloads DESC
            LIMIT 200;
            """,
            (start_ts, end_ts),
        )
        downloads_by_dataset = cur.fetchall() or []

        # 4) Events by type
        cur.execute(
            """
            SELECT
              event_type,
              COUNT(*) AS count
            FROM usage_events
            WHERE ts >= %s AND ts < %s
            GROUP BY event_type
            ORDER BY count DESC;
            """,
            (start_ts, end_ts),
        )
        by_type = cur.fetchall() or []

    return render_template(
        "admin/usage_stats.html",
        days=days,
        start_ts=start_ts,
        end_ts=end_ts,
        summary=summary,
        daily=daily,
        downloads_by_dataset=downloads_by_dataset,
        by_type=by_type,
    )


@web_bp.get("/download/canonical/<date>/<filename>")
@login_required
def download_canonical_version(date, filename):
    """
    Serve a specific dated canonical CSV from MinIO, e.g.
    /download/canonical/2025-12-02/samples.csv
    which maps to MinIO object canonical/2025-12-02/samples.csv
    """
    obj_name = f"{date}/{filename}"
    return _stream_minio_canonical(obj_name)


from echorepo.i18n import base_labels


@web_bp.get("/explore", endpoint="explore")
def explore():
    base = base_labels()  # <-- THIS IS THE KEY LINE

    i18n = {
        "labels": build_i18n_labels(base),
        "by_msgid": {},
    }

    return render_template(
        "explore.html",
        jitter_m=int(settings.MAX_JITTER_METERS),
        lat_col=settings.LAT_COL,
        lon_col=settings.LON_COL,
        I18N=i18n,
        current_locale=str(get_locale() or "en"),
        public_mode=True,
    )


@web_bp.get("/public/others_geojson")
def public_others_geojson():
    # Keep this very lightweight: only what the map needs.
    sql = """
    SELECT
        sample_id, lat, lon, ph, country_code, timestamp_utc,
        soil_texture_en, soil_texture_orig,
        soil_structure_en, soil_structure_orig,
        earthworms_count,
        contamination_plastic, contamination_debris,
        contamination_other_en, contamination_other_orig,
        observations_en, observations_orig,
        metals_info_en, metals_info_orig,
        location_accuracy_m
    FROM samples
    WHERE lat IS NOT NULL AND lon IS NOT NULL
    ORDER BY timestamp_utc DESC
    LIMIT 20000
    """

    with get_pg_conn() as conn, conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute(sql)
        rows = cur.fetchall()

    feats = []
    for r in rows:
        try:
            lat = float(r["lat"])
            lon = float(r["lon"])
        except Exception:
            continue

        feats.append(
            {
                "type": "Feature",
                "geometry": {"type": "Point", "coordinates": [lon, lat]},
                "properties": {
                    "sampleId": r["sample_id"],
                    "PH_ph": r.get("ph"),
                    "country_code": r.get("country_code"),
                    "timestamp_utc": (
                        r["timestamp_utc"].isoformat() if r.get("timestamp_utc") else None
                    ),
                    # soil descriptors (your map.js "pick(...)" already supports these)
                    "soil_texture_en": r.get("soil_texture_en"),
                    "soil_texture_orig": r.get("soil_texture_orig"),
                    "soil_structure_en": r.get("soil_structure_en"),
                    "soil_structure_orig": r.get("soil_structure_orig"),
                    # simple counts/flags
                    "earthworms_count": r.get("earthworms_count"),
                    "contamination_plastic": r.get("contamination_plastic"),
                    "contamination_debris": r.get("contamination_debris"),
                    "contamination_other_en": r.get("contamination_other_en"),
                    "contamination_other_orig": r.get("contamination_other_orig"),
                    "observations_en": r.get("observations_en"),
                    "observations_orig": r.get("observations_orig"),
                    # “metals blob” if you have it in samples
                    "metals_info_en": r.get("metals_info_en"),
                    "metals_info_orig": r.get("metals_info_orig"),
                    # optional
                    "location_accuracy_m": r.get("location_accuracy_m"),
                },
            }
        )

    return jsonify({"type": "FeatureCollection", "features": feats})


@web_bp.get("/public/sample_image/<sample_id>")
def public_sample_image(sample_id: str):
    # Return 1 thumbnail-ish image URL (or None) for a sample_id
    sql = """
      SELECT image_url, image_description_en, image_description_orig
      FROM sample_images
      WHERE sample_id = %s
        AND image_url IS NOT NULL AND image_url <> ''
      ORDER BY image_id
      LIMIT 1
    """
    with get_pg_conn() as conn, conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute(sql, (sample_id,))
        row = cur.fetchone()

    if not row:
        return jsonify({"ok": True, "image_url": None})

    desc = row.get("image_description_en") or row.get("image_description_orig") or ""
    return jsonify({"ok": True, "image_url": row.get("image_url"), "caption": desc})


@web_bp.post("/lab-import-biodiversity")
@login_required
def lab_import_biodiversity():
    user_key = session.get("user") or session.get("kc", {}).get("profile", {}).get("email")
    if not can_upload_lab_data(user_key):
        abort(403, description="Not authorised to upload lab data")

    file = request.files.get("file")
    if not file:
        abort(400, description="No file uploaded")

    kc_profile = (session.get("kc") or {}).get("profile") or {}
    uploader_id = kc_profile.get("id") or kc_profile.get("sub") or session.get("user") or "unknown"
    filename = (file.filename or "").strip()

    log = logging.getLogger(__name__)

    # --- Read XLSX (only the sheet you want) ---
    try:
        df = pd.read_excel(
            file,
            engine="openpyxl",
            sheet_name="clean_phylum",
            dtype=str,  # keep everything as text; we'll parse counts ourselves
        )
    except Exception as e:
        abort(400, description=f"Cannot read XLSX (sheet clean_phylum): {e}")

    if df is None or df.empty:
        abort(400, description="Uploaded file has no rows")

    # Drop fully empty columns (Excel exports often have trailing blanks)
    df = df.dropna(axis=1, how="all")

    log.warning(
        "BIOUPLOAD: loaded sheet rows=%d cols=%d file=%s", len(df), len(df.columns), filename
    )

    # --- Find OTU column ---
    otu_col = None
    for cand in ("OTU ID", "OTU_ID", "OTU", "otu_id", "otu"):
        if cand in df.columns:
            otu_col = cand
            break
    if not otu_col:
        otu_col = df.columns[0]

    # Detect sample columns like "AAUU-9633-16S" / "ABYU-1769-ITS"
    sample_pat = re.compile(r"^[A-Za-z0-9]{4}-[A-Za-z0-9]{4,}-(16S|ITS)$", re.IGNORECASE)
    sample_cols = [c for c in df.columns if c != otu_col and sample_pat.match(str(c).strip())]

    if not sample_cols:
        abort(
            400,
            description="No sample columns found. Expected headers like 'AAUU-9633-16S' or 'AAUU-9633-ITS'.",
        )

    # Taxonomy columns = everything else (except OTU + sample columns)
    taxa_cols = [c for c in df.columns if c not in ([otu_col] + sample_cols)]
    # Drop taxonomy cols that are fully empty
    clean_taxa_cols = []
    for c in taxa_cols:
        s = df[c]
        if s.notna().any() and (s.astype(str).str.strip() != "").any():
            clean_taxa_cols.append(c)
    taxa_cols = clean_taxa_cols
    has_taxa = bool(taxa_cols)

    # Prepare OTU ids and taxa JSON per row index (cheap, avoids recomputing per sample)
    otu_ids = df[otu_col].fillna("").astype(str).str.strip().tolist()
    if has_taxa:
        taxa_json_by_row = []
        taxa_df = df[taxa_cols].fillna("")
        for _, rr in taxa_df.iterrows():
            d = {}
            for c in taxa_cols:
                v = str(rr.get(c) or "").strip()
                if v:
                    d[str(c)] = v
            taxa_json_by_row.append(json.dumps(d, ensure_ascii=False) if d else None)
    else:
        taxa_json_by_row = [None] * len(df)

    # --- Postgres insert ---
    upsert_sql = """
      INSERT INTO sample_otu_counts
        (sample_id, marker, otu_id, count, taxa, uploaded_at, uploaded_by, source_file)
      VALUES
        (%s, %s, %s, %s, %s::jsonb, now(), %s, %s)
      ON CONFLICT (sample_id, marker, otu_id) DO UPDATE SET
        count       = EXCLUDED.count,
        taxa        = COALESCE(EXCLUDED.taxa, sample_otu_counts.taxa),
        uploaded_at = now(),
        uploaded_by = EXCLUDED.uploaded_by,
        source_file = EXCLUDED.source_file
    """

    def _to_float(v):
        if v is None:
            return None
        s = str(v).strip()
        if not s or s.lower() in {"nan", "none", "null"}:
            return None
        # treat "0" as zero
        try:
            return float(s.replace(",", "."))
        except Exception:
            return None

    CHUNK = 20000
    total_rows = 0

    try:
        with get_pg_conn() as conn, conn.cursor() as cur:
            # ensure table exists (ok for now; move to migrations later)
            cur.execute("""
              CREATE TABLE IF NOT EXISTS sample_otu_counts (
                sample_id TEXT NOT NULL,
                marker TEXT NOT NULL,
                otu_id TEXT NOT NULL,
                count DOUBLE PRECISION,
                taxa JSONB,
                uploaded_at TIMESTAMPTZ NOT NULL DEFAULT now(),
                uploaded_by TEXT,
                source_file TEXT,
                PRIMARY KEY (sample_id, marker, otu_id)
              )
            """)
            conn.commit()

            batch = []

            # iterate column-by-column (no melt)
            for col in sample_cols:
                sample_id, marker = col.rsplit("-", 1)
                sample_id = sample_id.strip().upper()
                marker = marker.strip().upper()

                vals = df[col].tolist()
                for i, v in enumerate(vals):
                    otu_id = otu_ids[i]
                    if not otu_id:
                        continue

                    count = _to_float(v)
                    if count is None or count == 0:
                        continue

                    batch.append(
                        (
                            sample_id,
                            marker,
                            otu_id,
                            count,
                            taxa_json_by_row[i],
                            uploader_id,
                            filename,
                        )
                    )

                    if len(batch) >= CHUNK:
                        cur.executemany(upsert_sql, batch)
                        conn.commit()
                        total_rows += len(batch)
                        log.warning("BIOUPLOAD: inserted %d rows so far...", total_rows)
                        batch.clear()

            if batch:
                cur.executemany(upsert_sql, batch)
                conn.commit()
                total_rows += len(batch)
                batch.clear()

    except Exception as e:
        abort(500, description=f"Postgres import failed: {e}")

    if total_rows == 0:
        abort(400, description="No non-zero OTU counts found to import (all empty/zero?).")

    g._analytics_extra = {
        "upload_type": "biodiversity_otu_xlsx",
        "filename": filename,
        "rows_inserted": total_rows,
        "sheet": "clean_phylum",
    }

    log.warning("BIOUPLOAD: done, rows_inserted=%d", total_rows)
    return redirect(url_for("web.home"))


@web_bp.post("/lab-import-auto")
@login_required
def lab_import_auto():
    user_key = session.get("user") or session.get("kc", {}).get("profile", {}).get("email")
    if not can_upload_lab_data(user_key):
        abort(403, description="Not authorised to upload lab data")

    file = request.files.get("file")
    if not file:
        abort(400, description="No file uploaded")

    filename = (file.filename or "").strip()
    data = file.read()
    if not data:
        abort(400, description="Uploaded file is empty")

    kc_profile = (session.get("kc") or {}).get("profile") or {}
    uploader_id = kc_profile.get("id") or kc_profile.get("sub") or session.get("user") or "unknown"

    # XLSX can be either biodiversity OR metals
    if filename.lower().endswith(".xlsx"):
        if _looks_like_biodiversity_xlsx(data):
            inserted = _import_biodiversity_xlsx_streaming(data, filename, uploader_id)

            g._analytics_extra = {
                "upload_type": "biodiversity_otu_xlsx",
                "filename": filename,
                "rows_inserted": inserted,
            }
            return redirect(url_for("web.home"))

        # otherwise treat XLSX as normal lab/metals file
        _import_metals_file_bytes(data, filename, uploader_id)
        g._analytics_extra = {
            "upload_type": "lab_xlsx",
            "filename": filename,
        }
        return redirect(url_for("web.home"))

    # non-XLSX fallback: CSV/TSV metals importer
    _import_metals_file_bytes(data, filename, uploader_id)
    g._analytics_extra = {
        "upload_type": "lab_csv",
        "filename": filename,
    }
    return redirect(url_for("web.home"))


def _looks_like_biodiversity_xlsx(xlsx_bytes: bytes) -> bool:
    """
    Detect OTU XLSX by presence of sheet 'clean_phylum' (case-insensitive) OR
    by columns that look like '<QR>-16S'/'<QR>-ITS' and 'OTU ID'.
    Cheap sniff: open workbook metadata only.
    """
    try:
        xls = pd.ExcelFile(io.BytesIO(xlsx_bytes), engine="openpyxl")
        sheet_names = [s.strip().lower() for s in xls.sheet_names]
        if "clean_phylum" in sheet_names:
            return True

        # fallback: peek first sheet headers quickly
        df0 = pd.read_excel(xls, sheet_name=0, nrows=1, dtype=str)
        cols = [str(c).strip() for c in df0.columns]
        if any(c.lower() in {"otu id", "otu_id", "otu"} for c in cols):
            pat = re.compile(r"^[A-Za-z0-9]{4}-[A-Za-z0-9]{4,}-(16S|ITS)$", re.IGNORECASE)
            if any(pat.match(c) for c in cols):
                return True
        return False
    except Exception:
        return False


def _import_metals_file_bytes(data: bytes, filename: str, uploader_id: str):
    """
    Preserve your existing /lab-import behavior, but operate on bytes.
    """
    # parse to dataframe like before
    try:
        if filename.lower().endswith(".xlsx"):
            df = pd.read_excel(io.BytesIO(data))
        else:
            # try TSV then CSV
            try:
                df = pd.read_csv(io.BytesIO(data), sep="\t")
            except Exception:
                df = pd.read_csv(io.BytesIO(data))
    except Exception as e:
        abort(400, description=f"Cannot read file: {e}")

    if df.empty:
        abort(400, description="Uploaded file has no rows")

    db_path = settings.SQLITE_PATH
    if not os.path.exists(db_path):
        abort(500, description=f"SQLite database not found at {db_path}")

    conn = sqlite3.connect(db_path)
    _ensure_lab_enrichment(conn)
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS lab_enrichment (
          qr_code    TEXT NOT NULL,
          param      TEXT NOT NULL,
          value      TEXT,
          unit       TEXT,
          user_id    TEXT,
          raw_row    TEXT,
          updated_at TEXT DEFAULT (datetime('now')),
          PRIMARY KEY (qr_code, param)
        )
    """)

    fieldnames = list(df.columns)

    for __idx, row in df.iterrows():
        raw_dict = row.to_dict()
        qr = _normalize_qr(raw_dict.get("ID") or raw_dict.get("id") or "")
        if not qr:
            continue

        clean_raw = {k: ("" if pd.isna(v) else v) for k, v in raw_dict.items()}
        raw_json = json.dumps(clean_raw, ensure_ascii=False)

        for idx, col in enumerate(fieldnames):
            if col in ("ID", "id"):
                continue
            val = row.get(col)
            if pd.isna(val) or val == "":
                continue
            if str(col).lower().startswith("unit"):
                continue

            param = str(col).strip()
            unit = ""

            if idx + 1 < len(fieldnames):
                maybe_unit_col = fieldnames[idx + 1]
                if str(maybe_unit_col).lower().startswith("unit"):
                    uval = row.get(maybe_unit_col)
                    if not pd.isna(uval):
                        unit = str(uval).strip()

            cur.execute(
                """
                INSERT INTO lab_enrichment (qr_code, param, value, unit, user_id, raw_row, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, datetime('now'))
                ON CONFLICT(qr_code, param) DO UPDATE SET
                  value=excluded.value,
                  unit=excluded.unit,
                  user_id=excluded.user_id,
                  raw_row=excluded.raw_row,
                  updated_at=datetime('now')
            """,
                (qr, param, str(val), unit, uploader_id, raw_json),
            )

            conv = _oxide_to_metal(param, val)
            if conv is not None:
                metal_param, metal_val = conv
                cur.execute(
                    """
                    INSERT INTO lab_enrichment (qr_code, param, value, unit, user_id, raw_row, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, datetime('now'))
                    ON CONFLICT(qr_code, param) DO UPDATE SET
                      value=excluded.value,
                      unit=excluded.unit,
                      user_id=excluded.user_id,
                      raw_row=excluded.raw_row,
                      updated_at=datetime('now')
                """,
                    (qr, metal_param, str(metal_val), unit, uploader_id, raw_json),
                )

    conn.commit()
    conn.close()


def _import_biodiversity_xlsx_bytes(xlsx_bytes: bytes, filename: str, uploader_id: str):
    log = logging.getLogger(__name__)

    try:
        xls = pd.ExcelFile(io.BytesIO(xlsx_bytes), engine="openpyxl")
        sheets = {s.strip().lower(): s for s in xls.sheet_names}
        sheet = sheets.get("clean_phylum")
        if not sheet:
            abort(400, description="Sheet 'clean_phylum' not found in XLSX")
        df = pd.read_excel(xls, sheet_name=sheet, dtype=str)
    except Exception as e:
        abort(400, description=f"Cannot read biodiversity XLSX: {e}")

    if df is None or df.empty:
        abort(400, description="Uploaded biodiversity file has no rows")

    df = df.dropna(axis=1, how="all")
    log.warning(
        "BIOUPLOAD(auto): loaded sheet=%s rows=%d cols=%d file=%s",
        sheet,
        len(df),
        len(df.columns),
        filename,
    )

    otu_col = None
    for cand in ("OTU ID", "OTU_ID", "OTU", "otu_id", "otu"):
        if cand in df.columns:
            otu_col = cand
            break
    if not otu_col:
        otu_col = df.columns[0]

    sample_pat = re.compile(r"^[A-Za-z0-9]{4}-[A-Za-z0-9]{4,}-(16S|ITS)$", re.IGNORECASE)
    sample_cols = [c for c in df.columns if c != otu_col and sample_pat.match(str(c).strip())]

    if not sample_cols:
        abort(400, description="No OTU sample columns found (expected 'AAUU-9633-16S'/'...-ITS').")

    taxa_cols = [c for c in df.columns if c not in ([otu_col] + sample_cols)]
    clean_taxa_cols = []
    for c in taxa_cols:
        s = df[c]
        if s.notna().any() and (s.astype(str).str.strip() != "").any():
            clean_taxa_cols.append(c)
    taxa_cols = clean_taxa_cols
    has_taxa = bool(taxa_cols)

    otu_ids = df[otu_col].fillna("").astype(str).str.strip().tolist()

    if has_taxa:
        taxa_json_by_row = []
        taxa_df = df[taxa_cols].fillna("")
        for _, rr in taxa_df.iterrows():
            d = {}
            for c in taxa_cols:
                v = str(rr.get(c) or "").strip()
                if v:
                    d[str(c)] = v
            taxa_json_by_row.append(json.dumps(d, ensure_ascii=False) if d else None)
    else:
        taxa_json_by_row = [None] * len(df)

    def _to_float(v):
        if v is None:
            return None
        s = str(v).strip()
        if not s or s.lower() in {"nan", "none", "null"}:
            return None
        try:
            return float(s.replace(",", "."))
        except Exception:
            return None

    upsert_sql = """
      INSERT INTO sample_otu_counts
        (sample_id, marker, otu_id, count, taxa, uploaded_at, uploaded_by, source_file)
      VALUES
        (%s, %s, %s, %s, %s::jsonb, now(), %s, %s)
      ON CONFLICT (sample_id, marker, otu_id) DO UPDATE SET
        count       = EXCLUDED.count,
        taxa        = COALESCE(EXCLUDED.taxa, sample_otu_counts.taxa),
        uploaded_at = now(),
        uploaded_by = EXCLUDED.uploaded_by,
        source_file = EXCLUDED.source_file
    """

    chunk = 20000
    total_rows = 0
    batch = []

    try:
        with get_pg_conn() as conn, conn.cursor() as cur:
            cur.execute("""
              CREATE TABLE IF NOT EXISTS sample_otu_counts (
                sample_id TEXT NOT NULL,
                marker TEXT NOT NULL,
                otu_id TEXT NOT NULL,
                count DOUBLE PRECISION,
                taxa JSONB,
                uploaded_at TIMESTAMPTZ NOT NULL DEFAULT now(),
                uploaded_by TEXT,
                source_file TEXT,
                PRIMARY KEY (sample_id, marker, otu_id)
              )
            """)
            conn.commit()

            for col in sample_cols:
                sample_id, marker = str(col).strip().rsplit("-", 1)
                sample_id = sample_id.strip().upper()
                marker = marker.strip().upper()

                vals = df[col].tolist()
                for i, v in enumerate(vals):
                    otu_id = otu_ids[i]
                    if not otu_id:
                        continue

                    count = _to_float(v)
                    if count is None or count == 0:
                        continue

                    batch.append(
                        (
                            sample_id,
                            marker,
                            otu_id,
                            float(count),
                            taxa_json_by_row[i],
                            uploader_id,
                            filename,
                        )
                    )

                    if len(batch) >= chunk:
                        cur.executemany(upsert_sql, batch)
                        conn.commit()
                        total_rows += len(batch)
                        log.warning("BIOUPLOAD(auto): inserted %d rows so far...", total_rows)
                        batch.clear()

            if batch:
                cur.executemany(upsert_sql, batch)
                conn.commit()
                total_rows += len(batch)

    except Exception as e:
        abort(500, description=f"Postgres biodiversity import failed: {e}")

    if total_rows == 0:
        abort(400, description="No non-zero OTU counts found to import (all empty/zero?).")

    log.warning("BIOUPLOAD(auto): done, rows_inserted=%d", total_rows)


def _import_biodiversity_xlsx_streaming(xlsx_bytes: bytes, filename: str, uploader_id: str):
    sample_pat = re.compile(r"^[A-Za-z0-9]{4}-[A-Za-z0-9]{4,}-(16S|ITS)$", re.IGNORECASE)

    try:
        wb = load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
    except Exception as e:
        abort(400, description=f"Cannot open XLSX: {e}")

    if "clean_phylum" not in wb.sheetnames:
        abort(400, description="XLSX does not contain required sheet 'clean_phylum'")

    ws = wb["clean_phylum"]

    rows_iter = ws.iter_rows(values_only=True)

    try:
        header = next(rows_iter)
    except StopIteration:
        abort(400, description="Sheet 'clean_phylum' is empty")

    header = [("" if v is None else str(v).strip()) for v in header]

    if not header:
        abort(400, description="Sheet header is empty")

    otu_col_idx = 0

    sample_cols = []
    for idx, col in enumerate(header):
        if idx == otu_col_idx:
            continue
        if sample_pat.match(col):
            sample_cols.append((idx, col))

    if not sample_cols:
        abort(400, description="No sample columns found like ABCD-1234-16S or ABCD-1234-ITS")

    taxa_cols = []
    for idx, col in enumerate(header):
        if idx == otu_col_idx:
            continue
        if any(idx == sidx for sidx, _ in sample_cols):
            continue
        taxa_cols.append((idx, col))

    insert_sql = """
        INSERT INTO sample_otu_counts
            (sample_id, marker, otu_id, count, taxa, uploaded_at, uploaded_by, source_file)
        VALUES
            (%s, %s, %s, %s, %s::jsonb, now(), %s, %s)
        ON CONFLICT (sample_id, marker, otu_id) DO UPDATE SET
            count       = EXCLUDED.count,
            taxa        = COALESCE(EXCLUDED.taxa, sample_otu_counts.taxa),
            uploaded_at = now(),
            uploaded_by = EXCLUDED.uploaded_by,
            source_file = EXCLUDED.source_file
    """

    def split_sample_marker(s: str):
        parts = s.rsplit("-", 1)
        if len(parts) != 2:
            return s.strip().upper(), ""
        return parts[0].strip().upper(), parts[1].strip().upper()

    def to_float_or_none(v):
        if v is None:
            return None
        s = str(v).strip().replace(",", ".")
        if not s:
            return None
        try:
            return float(s)
        except Exception:
            return None

    batch = []
    batch_size = 5000
    inserted = 0

    try:
        with get_pg_conn() as conn, conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS sample_otu_counts (
                    sample_id TEXT NOT NULL,
                    marker TEXT NOT NULL,
                    otu_id TEXT NOT NULL,
                    count DOUBLE PRECISION,
                    taxa JSONB,
                    uploaded_at TIMESTAMPTZ NOT NULL DEFAULT now(),
                    uploaded_by TEXT,
                    source_file TEXT,
                    PRIMARY KEY (sample_id, marker, otu_id)
                )
            """)

            cur.execute("""
                SELECT 1
                FROM information_schema.columns
                WHERE table_name='sample_otu_counts' AND column_name='taxa'
            """)
            if cur.fetchone() is None:
                cur.execute("ALTER TABLE sample_otu_counts ADD COLUMN taxa JSONB")

            for row in rows_iter:
                otu_id = "" if row[otu_col_idx] is None else str(row[otu_col_idx]).strip()
                if not otu_id:
                    continue

                taxa_dict = {}
                for idx, col_name in taxa_cols:
                    val = row[idx] if idx < len(row) else None
                    if val is None:
                        continue
                    sval = str(val).strip()
                    if sval:
                        taxa_dict[col_name] = sval

                taxa_json = json.dumps(taxa_dict, ensure_ascii=False) if taxa_dict else None

                for idx, sample_col in sample_cols:
                    val = row[idx] if idx < len(row) else None
                    count = to_float_or_none(val)
                    if count is None or count == 0:
                        continue

                    sample_id, marker = split_sample_marker(sample_col)
                    if not sample_id or not marker:
                        continue

                    batch.append(
                        (sample_id, marker, otu_id, count, taxa_json, uploader_id, filename)
                    )

                    if len(batch) >= batch_size:
                        cur.executemany(insert_sql, batch)
                        inserted += len(batch)
                        batch.clear()

            if batch:
                cur.executemany(insert_sql, batch)
                inserted += len(batch)

            conn.commit()

    except Exception as e:
        abort(500, description=f"Postgres biodiversity import failed: {e}")

    return inserted


@web_bp.get("/public/sample_piechart/<sample_id>")
def public_sample_piechart(sample_id: str):
    marker = (request.args.get("marker") or "16S").strip()
    level = (request.args.get("level") or "Genus").strip()

    client = _get_minio_client()
    bucket = os.getenv("MINIO_BUCKET", "echorepo-uploads")
    object_name = f"biodiversity/piecharts/{marker}/{level}/{sample_id}.png"

    if client is None:
        return jsonify({"ok": True, "image_url": None, "caption": ""})

    try:
        client.stat_object(bucket, object_name)
    except Exception:
        return jsonify({"ok": True, "image_url": None, "caption": ""})

    image_url = f"/storage/{object_name}"
    return jsonify(
        {
            "ok": True,
            "image_url": image_url,
            "caption": f"{marker} · {level}",
        }
    )


def _read_zenodo_publications(log_path: str) -> list[dict]:
    p = Path(log_path)
    if not p.exists():
        return []

    rows: list[dict] = []
    with p.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if (row.get("status") or "").strip().lower() != "ok":
                continue
            if not (row.get("version_doi") or row.get("zenodo_html")):
                continue
            rows.append(row)

    # newest first
    rows.sort(key=lambda r: r.get("run_at_utc", ""), reverse=True)
    return rows


@web_bp.get("/publications/zenodo", endpoint="zenodo_publications")
def zenodo_publications():
    publications = _read_zenodo_publications(ZENODO_LOG_DEFAULT)

    # If you want only production records on the public page:
    # publications = [p for p in publications if p.get("sandbox") != "1"]

    return render_template(
        "zenodo_publications.html",
        publications=publications,
        current_locale=str(get_locale() or "en"),
    )
