#!/usr/bin/env python3
"""
ECHOrepo biodiversity image generator
======================================

Run commands from the repository root, where the project .env file is located.

The script skips images already present in MinIO by default. Use --force to
recreate and overwrite them.

IMAGE TYPES AND COMMANDS
------------------------

1. Bacterial taxonomic pie charts: 16S / Phylum

   python3 tools/generate_biodiversity_piecharts.py \
     --marker 16S \
     --level Phylum

   MinIO destination:

     biodiversity/piecharts/16S/Phylum/<sample_id>.png


2. Fungal taxonomic pie charts: ITS / Phylum

   python3 tools/generate_biodiversity_piecharts.py \
     --marker ITS \
     --level Phylum

   MinIO destination:

     biodiversity/piecharts/ITS/Phylum/<sample_id>.png


3. Fungal ecological guild images

   python3 tools/generate_biodiversity_piecharts.py \
     --marker ITS \
     --level Phylum \
     --fungal-guilds

   This command also generates any missing ITS Phylum pie charts.

   Fungal guild data are reconstructed from the current raw ITS archives in
   MinIO. The legacy sample_otu_counts table is used only as a fallback for
   older samples.

   The raw ITS data must contain usable genus-level taxonomy.

   MinIO destination:

     biodiversity/guildplots/fungi/<sample_id>.png


4. Build FAPROTAX input files for bacterial guild analysis

   python3 tools/generate_biodiversity_piecharts.py \
     --marker 16S \
     --level Phylum \
     --build-faprotax-inputs

   Generated files:

     data/biodiversity/faprotax_work/6_otu_clean_counts_no_blanks.csv
     data/biodiversity/faprotax_work/7_taxonomy_clean.csv

   This step reads the current structured raw OTU data from PostgreSQL:

     biodiversity_raw_samples
     biodiversity_raw_features
     biodiversity_raw_abundance

   sample_taxon_abundance.source_upload_id is used to select the current source
   upload for each sample, so historical/replaced uploads are not mixed in.

   Run the separate FAPROTAX analysis after creating these files. Its resulting
   sample-by-function file must be available at FAPROTAX_FUNCTION_CSV, whose
   default value is:

     data/biodiversity/8_faprotax_samples_x_functions.csv


5. Bacterial ecological guild images

   python3 tools/generate_biodiversity_piecharts.py \
     --marker 16S \
     --level Phylum \
     --bacterial-guilds

   This command also generates any missing 16S Phylum pie charts.

   Bacterial guild images are generated from the configured FAPROTAX
   sample-by-function CSV.

   MinIO destination:

     biodiversity/guildplots/bacteria/<sample_id>.png


GENERATE ALL IMAGE TYPES
------------------------

First build the FAPROTAX inputs:

   python3 tools/generate_biodiversity_piecharts.py \
     --marker 16S \
     --level Phylum \
     --build-faprotax-inputs

Run the external FAPROTAX processing and place its result at the path configured
by FAPROTAX_FUNCTION_CSV.

Then generate all bacterial images:

   python3 tools/generate_biodiversity_piecharts.py \
     --marker 16S \
     --level Phylum \
     --bacterial-guilds

Finally, generate all fungal images:

   python3 tools/generate_biodiversity_piecharts.py \
     --marker ITS \
     --level Phylum \
     --fungal-guilds


USEFUL OPTIONS
--------------

Generate images for one sample only:

   python3 tools/generate_biodiversity_piecharts.py \
     --marker ITS \
     --fungal-guilds \
     --sample-id CLMW-8393

Generate images for several selected samples:

   python3 tools/generate_biodiversity_piecharts.py \
     --marker ITS \
     --fungal-guilds \
     --sample-id CLMW-8393,AACW-5934

The --sample-id option may also be repeated:

   python3 tools/generate_biodiversity_piecharts.py \
     --marker ITS \
     --fungal-guilds \
     --sample-id CLMW-8393 \
     --sample-id AACW-5934

Preview missing images without generating or uploading them:

   python3 tools/generate_biodiversity_piecharts.py \
     --marker ITS \
     --fungal-guilds \
     --dry-run

Recreate all selected images, including images already present in MinIO:

   python3 tools/generate_biodiversity_piecharts.py \
     --marker ITS \
     --fungal-guilds \
     --force


ENVIRONMENT-VARIABLE ALTERNATIVES
---------------------------------

The following environment variables can be used instead of the corresponding
command-line flags:

   GENERATE_FUNGAL_GUILDS=1
   GENERATE_BACTERIAL_GUILDS=1
   BUILD_FAPROTAX_INPUTS=1
   BIODIV_FORCE_REGENERATE=1
   BIODIV_MARKER=16S
   BIODIV_LEVEL=Phylum

Example:

   GENERATE_FUNGAL_GUILDS=1 \
   BIODIV_MARKER=ITS \
   python3 tools/generate_biodiversity_piecharts.py
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import math
import os
import re
import shutil
import sys
import tempfile
import zipfile
from pathlib import Path

import matplotlib as mpl
import matplotlib.pyplot as plt
import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Load .env exactly like pull_and_enrich_samples.py
# ---------------------------------------------------------------------------
env_path = Path.cwd() / ".env"
load_dotenv(dotenv_path=env_path)
print(f"[INFO] Loaded environment from {env_path}")

# ---------------------------------------------------------------------------
# Make sure project root is importable
# ---------------------------------------------------------------------------
THIS_DIR = Path(__file__).resolve().parent
DEFAULT_ROOT = THIS_DIR.parent
PROJECT_ROOT = Path(os.getenv("PROJECT_ROOT", str(DEFAULT_ROOT)))
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))
print(f"[INFO] Using PROJECT_ROOT={PROJECT_ROOT}")

# ---------------------------------------------------------------------------
# MinIO config: same style as pull_and_enrich_samples.py
# ---------------------------------------------------------------------------
try:
    from minio import Minio
    from minio.error import S3Error
except ImportError:
    Minio = None

    class S3Error(Exception):
        pass


MINIO_ENDPOINT = (
    os.getenv("MINIO_ENDPOINT_INSIDE")
    or os.getenv("MINIO_ENDPOINT_OUTSIDE")
    or os.getenv("MINIO_ENDPOINT")
    or "localhost:9000"
)
MINIO_ACCESS_KEY = os.getenv("MINIO_ACCESS_KEY") or os.getenv("MINIO_ROOT_USER") or ""
MINIO_SECRET_KEY = os.getenv("MINIO_SECRET_KEY") or os.getenv("MINIO_ROOT_PASSWORD") or ""
MINIO_BUCKET = os.getenv("MINIO_BUCKET", "echorepo-uploads")
PUBLIC_STORAGE_BASE = os.getenv("PUBLIC_STORAGE_BASE", "/storage")
FUNGUILD_DB_JSON = os.getenv(
    "FUNGUILD_DB_JSON",
    str(PROJECT_ROOT / "data" / "biodiversity" / "FUNGuild_db.json"),
)

GENERATE_FUNGAL_GUILDS = os.getenv("GENERATE_FUNGAL_GUILDS", "0") == "1"
BUILD_FAPROTAX_INPUTS = os.getenv("BUILD_FAPROTAX_INPUTS", "0") == "1"

GENERATE_BACTERIAL_GUILDS = os.getenv("GENERATE_BACTERIAL_GUILDS", "0") == "1"

FAPROTAX_FUNCTION_CSV = os.getenv(
    "FAPROTAX_FUNCTION_CSV",
    str(PROJECT_ROOT / "data" / "biodiversity" / "8_faprotax_samples_x_functions.csv"),
)

# ---------------------------------------------------------------------------
# Plot styling
# ---------------------------------------------------------------------------
PIE_BG = "#FFFFFF"
PIE_TEXT = "#000000"
PIE_EDGE = "#FFFFFF"
PIE_GRID = "#e0e0e0"

PIE_COLORS = [
    "#f0746a",  # salmon
    "#df9600",  # orange
    "#a6a800",  # olive
    "#41c400",  # green
    "#12bf80",  # teal-green
    "#1db7be",  # cyan-teal
    "#20a7df",  # blue
    "#8a83e6",  # lavender
    "#cc62dc",  # magenta-violet
    "#eb5bb3",  # pink
    "#999999",  # grey fallback for "Other"
]

mpl.rcParams["font.family"] = "DejaVu Sans"
mpl.rcParams["text.color"] = PIE_TEXT
mpl.rcParams["axes.labelcolor"] = PIE_TEXT
mpl.rcParams["xtick.color"] = PIE_TEXT
mpl.rcParams["ytick.color"] = PIE_TEXT

# ---------------------------------------------------------------------------
# Postgres config
# ---------------------------------------------------------------------------
try:
    import psycopg2
except ImportError:
    psycopg2 = None


def get_pg_conn():
    if psycopg2 is None:
        raise RuntimeError("psycopg2 is not installed")

    host = (
        os.getenv("DB_HOST_OUTSIDE")
        or os.getenv("DB_HOST_INSIDE")
        or os.getenv("DB_HOST")
        or "localhost"
    )
    port = int(
        os.getenv("DB_PORT_OUTSIDE")
        or os.getenv("DB_PORT_INSIDE")
        or os.getenv("DB_PORT")
        or "5432"
    )

    return psycopg2.connect(
        host=host,
        port=port,
        dbname=os.getenv("DB_NAME", "echorepo"),
        user=os.getenv("DB_USER", "echorepo"),
        password=os.getenv("DB_PASSWORD", "echorepo-pass"),
    )


def init_minio():
    if Minio is None:
        print("[INFO] python-minio not installed; skipping MinIO upload.")
        return None

    secure = False
    endpoint = MINIO_ENDPOINT
    if endpoint.startswith("https://"):
        secure = True
        endpoint = endpoint[len("https://") :]
    elif endpoint.startswith("http://"):
        secure = False
        endpoint = endpoint[len("http://") :]

    if not MINIO_ACCESS_KEY or not MINIO_SECRET_KEY:
        print("[WARN] MinIO credentials not set; skipping chart upload.")
        return None

    client = Minio(
        endpoint,
        access_key=MINIO_ACCESS_KEY,
        secret_key=MINIO_SECRET_KEY,
        secure=secure,
    )

    try:
        found = client.bucket_exists(MINIO_BUCKET)
        if not found:
            client.make_bucket(MINIO_BUCKET)
            print(f"[INFO] Created MinIO bucket {MINIO_BUCKET}")
    except Exception as e:
        print(f"[WARN] Could not ensure MinIO bucket: {e}")
        return None

    print(f"[INFO] MinIO ready at {MINIO_ENDPOINT}, bucket={MINIO_BUCKET}")
    return client


def upload_file_to_minio(
    mclient, local_path: Path, object_name: str, content_type: str = "image/png"
):
    if mclient is None:
        return None

    try:
        size = local_path.stat().st_size
        with local_path.open("rb") as f:
            mclient.put_object(
                MINIO_BUCKET,
                object_name,
                data=f,
                length=size,
                content_type=content_type,
            )
        print(f"[OK] uploaded to MinIO: {object_name}")
        return f"{PUBLIC_STORAGE_BASE}/{object_name}"
    except Exception as e:
        print(f"[WARN] could not upload {local_path} to MinIO as {object_name}: {e}")
        return None


def list_existing_minio_objects(mclient, prefix: str) -> set[str]:
    """
    Return all existing object names below a MinIO prefix.

    The generator calls this once per chart family, rather than issuing one
    stat/HEAD request per sample. If MinIO is configured but listing fails,
    abort instead of accidentally regenerating the complete dataset.
    """
    if mclient is None:
        raise RuntimeError(
            "MinIO is unavailable, so existing charts cannot be checked safely. "
            "Configure MinIO or run with --force to regenerate local files intentionally."
        )

    try:
        existing = {
            obj.object_name
            for obj in mclient.list_objects(
                MINIO_BUCKET,
                prefix=prefix,
                recursive=True,
            )
        }
    except Exception as e:
        raise RuntimeError(
            f"Could not list existing MinIO objects under {prefix!r}: {e}"
        ) from e

    print(f"[INFO] Found {len(existing)} existing MinIO objects under {prefix}")
    return existing


def normalize_sample_filter(values: list[str] | None) -> set[str] | None:
    """Normalize repeated/comma-separated --sample-id arguments."""
    if not values:
        return None

    result: set[str] = set()
    for value in values:
        for token in re.split(r"[,;\s]+", str(value or "").strip()):
            if token:
                result.add(token.upper())

    return result or None

def normalize_taxonomic_level(level: str) -> str:
    """
    Normalize user/env level names.

    Accepts:
      Philum, phylum, Phylum, p, p__, p__Ascomycota

    Returns one of:
      Kingdom, Phylum, Class, Order, Family, Genus, Species
    """
    if level is None:
        return "Phylum"

    s = str(level).strip()

    if not s:
        return "Phylum"

    s_lower = s.lower().strip()

    aliases = {
        "kingdom": "Kingdom",
        "taxonomy": "Kingdom",
        "k": "Kingdom",
        "k__": "Kingdom",
        "d": "Kingdom",
        "d__": "Kingdom",

        "phylum": "Phylum",
        "philum": "Phylum",   # common typo
        "p": "Phylum",
        "p__": "Phylum",

        "class": "Class",
        "c": "Class",
        "c__": "Class",

        "order": "Order",
        "o": "Order",
        "o__": "Order",

        "family": "Family",
        "f": "Family",
        "f__": "Family",

        "genus": "Genus",
        "g": "Genus",
        "g__": "Genus",

        "species": "Species",
        "s": "Species",
        "s__": "Species",
    }

    if s_lower in aliases:
        return aliases[s_lower]

    # If someone passes something like p__Ascomycota,
    # infer the level from the prefix.
    m = re.match(r"^([dkpcofgs])__", s_lower)
    if m:
        return aliases.get(m.group(1), "Phylum")

    # If already correctly capitalized
    for valid in ("Kingdom", "Phylum", "Class", "Order", "Family", "Genus", "Species"):
        if s_lower == valid.lower():
            return valid

    print(f"[WARN] Unknown BIODIV_LEVEL={level!r}; falling back to Phylum")
    return "Phylum"

def sanitize_filename(s: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", str(s).strip())


def fetch_otu_data(marker: str = "16S") -> pd.DataFrame:
    """
    Fetch legacy OTU-level data.

    This is retained only as a legacy fallback for older fungal/FUNGuild data.
    FAPROTAX inputs are built from the structured biodiversity_raw_* tables.
    """
    sql = """
        SELECT sample_id, otu_id, count, taxa
        FROM sample_otu_counts
        WHERE marker = %s
    """
    with get_pg_conn() as conn:
        df = pd.read_sql(sql, conn, params=[marker])
    return df



BIODIVERSITY_SAMPLE_COLUMN_RE = re.compile(
    r"^(?P<sample>[A-Za-z0-9]{4}-[A-Za-z0-9]{4,})-(?P<marker>16S|ITS)$",
    re.IGNORECASE,
)


def _normalize_biodiversity_header(value) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def fetch_current_biodiversity_archives(
    marker: str,
    sample_ids: set[str] | None = None,
) -> pd.DataFrame:
    """
    Return the raw MinIO archive currently associated with each sample.

    sample_taxon_abundance is the source of truth for which upload is current.
    All Phylum rows for one sample/marker should point to the same upload ID.
    """
    sql = """
        SELECT DISTINCT
            UPPER(sta.sample_id) AS sample_id,
            sta.source_upload_id,
            bu.archive_object_name,
            bu.original_filename,
            bu.uploaded_at
        FROM sample_taxon_abundance AS sta
        JOIN biodiversity_uploads AS bu
          ON bu.upload_id = sta.source_upload_id
        WHERE UPPER(sta.marker) = UPPER(%s)
          AND sta.source_upload_id IS NOT NULL
    """
    params: list[object] = [marker]

    if sample_ids:
        sql += " AND UPPER(sta.sample_id) = ANY(%s)"
        params.append(sorted(sample_ids))

    sql += " ORDER BY sample_id, bu.uploaded_at DESC"

    with get_pg_conn() as conn:
        df = pd.read_sql(sql, conn, params=params)

    if df.empty:
        return df

    # Defensive: if inconsistent rows exist, retain the newest mapping.
    return (
        df.sort_values("uploaded_at", ascending=False)
        .drop_duplicates(subset=["sample_id"], keep="first")
        .reset_index(drop=True)
    )


def _download_minio_object(mclient, object_name: str, destination: Path) -> None:
    if mclient is None:
        raise RuntimeError("MinIO is required to read raw biodiversity archives")

    response = None
    try:
        response = mclient.get_object(MINIO_BUCKET, object_name)
        with destination.open("wb") as out:
            shutil.copyfileobj(response, out, length=1024 * 1024)
    finally:
        if response is not None:
            response.close()
            response.release_conn()


def _extract_biodiversity_source(zip_path: Path, output_dir: Path) -> Path:
    """Extract the original CSV/TSV/XLSX from one raw-upload ZIP."""
    supported = {".csv", ".tsv", ".txt", ".xlsx"}

    with zipfile.ZipFile(zip_path) as zf:
        members = [
            info
            for info in zf.infolist()
            if not info.is_dir()
            and Path(info.filename).suffix.lower() in supported
        ]

        if not members:
            raise ValueError(f"No CSV/TSV/XLSX source found in {zip_path.name}")

        preferred_name = None
        try:
            manifest = json.loads(zf.read("manifest.json").decode("utf-8"))
            preferred_name = Path(
                str(manifest.get("original_filename") or "")
            ).name
        except Exception:
            preferred_name = None

        selected = members[0]
        if preferred_name:
            for member in members:
                if Path(member.filename).name == preferred_name:
                    selected = member
                    break

        safe_name = Path(selected.filename).name
        destination = output_dir / safe_name
        with zf.open(selected) as source, destination.open("wb") as target:
            shutil.copyfileobj(source, target, length=1024 * 1024)

    return destination


def _aggregate_fungal_rows(
    header,
    rows,
    target_samples: set[str],
) -> pd.DataFrame:
    """
    Reconstruct the information needed by FUNGuild from a raw OTU table.

    Rows are aggregated by sample and genus to keep memory use small. Counts
    without a usable genus are retained so chart percentages still use the
    full fungal-community read total as their denominator.
    """
    cleaned_header = [str(value or "").strip() for value in header]
    normalized = [_normalize_biodiversity_header(value) for value in cleaned_header]

    otu_idx = 0
    for idx, value in enumerate(normalized):
        if value in {"otuid", "otu"}:
            otu_idx = idx
            break

    sample_columns: list[tuple[int, str]] = []
    for idx, value in enumerate(cleaned_header):
        match = BIODIVERSITY_SAMPLE_COLUMN_RE.fullmatch(value)
        if not match or match.group("marker").upper() != "ITS":
            continue
        sample_id = match.group("sample").upper()
        if sample_id in target_samples:
            sample_columns.append((idx, sample_id))

    if not sample_columns:
        return pd.DataFrame(columns=["sample_id", "otu_id", "count", "taxa"])

    by_name = {value: idx for idx, value in enumerate(normalized) if value}
    genus_idx = by_name.get("genus")
    taxonomy_idx = by_name.get("taxonomy")

    # Legacy layout: Taxonomy, A, B, C, D, E, F where E is Genus.
    if genus_idx is None and taxonomy_idx is not None:
        genus_idx = by_name.get("e")

    aggregates: dict[str, dict[str, float]] = {}

    for row in rows:
        if otu_idx >= len(row) or not str(row[otu_idx] or "").strip():
            continue

        positive_counts: list[tuple[str, float]] = []
        for idx, sample_id in sample_columns:
            raw_value = row[idx] if idx < len(row) else None
            try:
                value = float(raw_value)
            except (TypeError, ValueError):
                continue
            if not math.isfinite(value) or value <= 0:
                continue
            positive_counts.append((sample_id, value))

        if not positive_counts:
            continue

        genus = ""
        if genus_idx is not None and genus_idx < len(row):
            genus = clean_taxon_value(row[genus_idx])

        if not genus and taxonomy_idx is not None and taxonomy_idx < len(row):
            genus = parse_taxonomy_string(row[taxonomy_idx]).get("Genus", "")
            genus = clean_taxon_value(genus)

        for sample_id, value in positive_counts:
            sample_counts = aggregates.setdefault(sample_id, {})
            sample_counts[genus] = sample_counts.get(genus, 0.0) + value

    output_rows = []
    for sample_id, genus_counts in aggregates.items():
        for genus, count in genus_counts.items():
            output_rows.append(
                {
                    "sample_id": sample_id,
                    "otu_id": f"raw-genus:{genus or 'unclassified'}",
                    "count": count,
                    "taxa": {"Genus": genus} if genus else {},
                }
            )

    return pd.DataFrame(output_rows, columns=["sample_id", "otu_id", "count", "taxa"])


def _read_fungal_rows_from_source(
    source_path: Path,
    target_samples: set[str],
) -> pd.DataFrame:
    suffix = source_path.suffix.lower()

    if suffix == ".xlsx":
        wb = load_workbook(source_path, read_only=True, data_only=True)
        try:
            sheet_name = next(
                (
                    name
                    for name in wb.sheetnames
                    if str(name).strip().lower() == "clean_phylum"
                ),
                None,
            )

            if sheet_name is None:
                for candidate in wb.sheetnames:
                    ws_candidate = wb[candidate]
                    first_row = next(
                        ws_candidate.iter_rows(
                            min_row=1,
                            max_row=1,
                            values_only=True,
                        ),
                        None,
                    )
                    if first_row and any(
                        BIODIVERSITY_SAMPLE_COLUMN_RE.fullmatch(str(v or "").strip())
                        for v in first_row
                    ):
                        sheet_name = candidate
                        break

            if sheet_name is None:
                raise ValueError(f"No biodiversity worksheet found in {source_path.name}")

            rows = wb[sheet_name].iter_rows(values_only=True)
            header = next(rows, None)
            if header is None:
                raise ValueError(f"Empty biodiversity worksheet in {source_path.name}")

            return _aggregate_fungal_rows(header, rows, target_samples)
        finally:
            wb.close()

    if suffix not in {".csv", ".tsv", ".txt"}:
        raise ValueError(f"Unsupported raw biodiversity file: {source_path.name}")

    with source_path.open("r", encoding="utf-8-sig", newline="") as stream:
        probe = stream.read(65536)
        stream.seek(0)

        if suffix == ".tsv":
            dialect = csv.excel_tab
        else:
            try:
                dialect = csv.Sniffer().sniff(probe, delimiters=",\t;")
            except csv.Error:
                dialect = csv.excel

        reader = csv.reader(stream, dialect=dialect)
        header = next(reader, None)
        if header is None:
            raise ValueError(f"Empty biodiversity file: {source_path.name}")

        return _aggregate_fungal_rows(header, reader, target_samples)


def fetch_current_fungal_data_from_archives(
    mclient,
    sample_ids: set[str] | None = None,
) -> tuple[pd.DataFrame, set[str], set[str]]:
    """
    Reconstruct current ITS genus/count data from raw MinIO archives.

    Returns: dataframe, all current sample IDs, successfully reconstructed IDs.
    """
    mappings = fetch_current_biodiversity_archives("ITS", sample_ids)
    if mappings.empty:
        return (
            pd.DataFrame(columns=["sample_id", "otu_id", "count", "taxa"]),
            set(),
            set(),
        )

    current_ids = set(mappings["sample_id"].astype(str).str.upper())
    frames: list[pd.DataFrame] = []
    reconstructed_ids: set[str] = set()

    grouped = mappings.groupby("archive_object_name", dropna=False)

    with tempfile.TemporaryDirectory(prefix="echorepo-funguild-") as tmp:
        tmp_dir = Path(tmp)

        for archive_number, (object_name, group) in enumerate(grouped, start=1):
            object_name = str(object_name or "").strip()
            archive_samples = set(group["sample_id"].astype(str).str.upper())

            if not object_name:
                print(
                    "[WARN] Current ITS samples have no archive object: "
                    f"{sorted(archive_samples)}"
                )
                continue

            archive_path = tmp_dir / f"archive-{archive_number}.zip"
            extract_dir = tmp_dir / f"archive-{archive_number}"
            extract_dir.mkdir(parents=True, exist_ok=True)

            try:
                print(
                    f"[INFO] Reading raw ITS archive {object_name} "
                    f"for {len(archive_samples)} samples"
                )
                _download_minio_object(mclient, object_name, archive_path)
                source_path = _extract_biodiversity_source(archive_path, extract_dir)
                frame = _read_fungal_rows_from_source(source_path, archive_samples)
            except Exception as exc:
                print(f"[WARN] Could not reconstruct {object_name}: {exc}")
                continue

            if frame.empty:
                print(f"[WARN] No positive ITS rows reconstructed from {object_name}")
                continue

            frame["sample_id"] = frame["sample_id"].astype(str).str.upper()
            found = set(frame["sample_id"].unique())
            reconstructed_ids.update(found)
            frames.append(frame)

            missing = archive_samples - found
            if missing:
                print(
                    f"[WARN] Archive {object_name} did not yield ITS data for "
                    f"{len(missing)} mapped samples: {sorted(missing)[:10]}"
                )

    combined = (
        pd.concat(frames, ignore_index=True)
        if frames
        else pd.DataFrame(columns=["sample_id", "otu_id", "count", "taxa"])
    )
    return combined, current_ids, reconstructed_ids


def fetch_fungal_guild_source_data(
    mclient,
    sample_ids: set[str] | None = None,
) -> pd.DataFrame:
    """
    Build the effective fungal source dataset.

    Current raw archives take precedence. The legacy sample_otu_counts table is
    retained only for samples that have no current sample_taxon_abundance
    provenance, so old data cannot override a newer upload.
    """
    archive_df, current_ids, reconstructed_ids = (
        fetch_current_fungal_data_from_archives(mclient, sample_ids)
    )

    legacy_df = fetch_otu_data(marker="ITS")
    if not legacy_df.empty:
        legacy_df["sample_id"] = legacy_df["sample_id"].astype(str).str.upper()
        if sample_ids:
            legacy_df = legacy_df[legacy_df["sample_id"].isin(sample_ids)].copy()
        legacy_df = legacy_df[~legacy_df["sample_id"].isin(current_ids)].copy()

    failed_current = current_ids - reconstructed_ids
    print(f"[INFO] Current ITS samples linked to raw archives: {len(current_ids)}")
    print(f"[INFO] Current ITS samples reconstructed: {len(reconstructed_ids)}")
    print(
        "[INFO] Legacy-only ITS samples retained: "
        f"{legacy_df['sample_id'].nunique() if not legacy_df.empty else 0}"
    )
    if failed_current:
        print(
            "[WARN] Current ITS samples not reconstructed and not replaced by "
            f"stale legacy rows: {len(failed_current)}"
        )

    frames = [df for df in (archive_df, legacy_df) if not df.empty]
    if not frames:
        return pd.DataFrame(columns=["sample_id", "otu_id", "count", "taxa"])

    combined = pd.concat(frames, ignore_index=True)
    print(f"[INFO] Effective ITS guild samples: {combined['sample_id'].nunique()}")
    return combined


def fetch_taxon_abundance(
    marker: str = "16S",
    level: str = "Phylum",
) -> pd.DataFrame:
    """
    Fetch compact taxonomic statistics produced by the current biodiversity
    importer.
    """
    sql = """
        SELECT
            sample_id,
            taxon,
            read_count AS count,
            relative_abundance_pct
        FROM sample_taxon_abundance
        WHERE marker = %s
          AND level = %s
        ORDER BY sample_id, read_count DESC, taxon
    """
    with get_pg_conn() as conn:
        return pd.read_sql(
            sql,
            conn,
            params=[marker.upper(), level],
        )


TAX_PREFIX_TO_RANK = {
    "k": "Kingdom",
    "p": "Phylum",
    "c": "Class",
    "o": "Order",
    "f": "Family",
    "g": "Genus",
    "s": "Species",
}


def clean_taxon_value(value: str) -> str:
    """
    Clean values like:
      g__Fusarium
      f__Nectriaceae
      Capnodiales_fam_Incertae_sedis

    Returns a display-friendly value, or "" if it is not useful.
    """
    if value is None:
        return ""

    s = str(value).strip()
    if not s:
        return ""

    # Remove rank prefix if still present
    s = re.sub(r"^[a-zA-Z]__", "", s)

    # Convert underscores to spaces for display / matching
    s = s.replace("_", " ").strip()

    # Drop low-information labels
    if re.search(r"incertae|unclassified|uncultured|unknown", s, flags=re.I):
        return ""

    # Drop generic species placeholders such as "Capnodiales sp"
    if re.search(r"\bsp\.?$", s, flags=re.I):
        return ""

    return s


def parse_taxonomy_string(raw: str) -> dict:
    """
    Parse a single taxonomy string like:
      k__Fungi;p__Ascomycota;c__Dothideomycetes;o__Capnodiales;f__...;g__...;s__...

    Returns:
      {
        "Kingdom": "Fungi",
        "Phylum": "Ascomycota",
        "Class": "Dothideomycetes",
        "Order": "Capnodiales",
        "Family": "...",
        "Genus": "...",
        "Species": "..."
      }
    """
    out = {}

    if raw is None:
        return out

    s = str(raw).strip()
    if not s:
        return out

    # Accept semicolon, pipe, or comma-separated taxonomy strings
    parts = re.split(r"\s*[;|]\s*", s)

    for part in parts:
        part = part.strip()
        if not part:
            continue

        m = re.match(r"^([kpcofgs])__?(.*)$", part, flags=re.I)
        if not m:
            continue

        prefix = m.group(1).lower()
        value = m.group(2).strip()
        rank = TAX_PREFIX_TO_RANK.get(prefix)
        if not rank:
            continue

        cleaned = clean_taxon_value(value)
        if cleaned:
            out[rank] = cleaned

    return out


def taxa_to_normalized_dict(taxa) -> dict:
    """
    Normalize taxa from Postgres sample_otu_counts.taxa.

    Supports:
      - dict with Taxonomy raw string
      - dict with A/B/C/D/E/F columns
      - dict with named ranks
      - JSON string
      - raw taxonomy string
    """
    d = _taxa_to_dict(taxa)

    # Case 1: taxa is a raw taxonomy string, not JSON
    if not d and isinstance(taxa, str):
        parsed = parse_taxonomy_string(taxa)
        if parsed:
            return parsed

    out = {}

    # Case 2: raw taxonomy column inside JSON/dict
    raw_tax = d.get("Taxonomy") or d.get("taxonomy") or d.get("taxon") or d.get("Taxon") or ""
    if raw_tax:
        out.update(parse_taxonomy_string(raw_tax))

    # Case 3: named rank columns already present
    for rank in ("Kingdom", "Phylum", "Class", "Order", "Family", "Genus", "Species"):
        val = d.get(rank) or d.get(rank.lower())
        cleaned = clean_taxon_value(val)
        if cleaned:
            out[rank] = cleaned

    # Case 4: old A/B/C/D/E/F style.
    # Based on your Excel: Taxonomy = kingdom, A=phylum, B=class,
    # C=order, D=family, E=genus, F=species.
    letter_map = {
        "A": "Phylum",
        "B": "Class",
        "C": "Order",
        "D": "Family",
        "E": "Genus",
        "F": "Species",
    }
    for key, rank in letter_map.items():
        cleaned = clean_taxon_value(d.get(key))
        if cleaned and rank not in out:
            out[rank] = cleaned

    return out


def extract_taxon_label(row: pd.Series, level: str) -> str:
    level = normalize_taxonomic_level(level)
    taxa = taxa_to_normalized_dict(row.get("taxa"))

    val = taxa.get(level)
    cleaned = clean_taxon_value(val)

    if cleaned:
        return cleaned

    return "Unclassified"


def make_piechart_for_sample(
    sample_df: pd.DataFrame, sample_id: str, marker: str, level: str, out_path: Path
):
    plot_df = sample_df.copy()

    # Current importer: taxon/read_count are already aggregated in
    # sample_taxon_abundance. Legacy OTU rows still need taxonomy extraction.
    if "taxon" in plot_df.columns:
        plot_df["taxon"] = (
            plot_df["taxon"]
            .fillna("Unclassified")
            .astype(str)
            .str.strip()
            .replace("", "Unclassified")
        )
    else:
        plot_df["taxon"] = plot_df.apply(
            lambda r: extract_taxon_label(r, level),
            axis=1,
        )

    plot_df["count"] = pd.to_numeric(
        plot_df["count"],
        errors="coerce",
    ).fillna(0)

    grouped = (
        plot_df.groupby("taxon", dropna=False)["count"]
        .sum()
        .reset_index()
        .sort_values("count", ascending=False)
    )

    grouped = grouped[grouped["count"] > 0].copy()
    if grouped.empty:
        return False

    # ---------- Collapse low-percentage taxa into Other ----------
    # First calculate percentages against the full sample total.
    total = grouped["count"].sum()
    grouped["pct"] = grouped["count"] / total * 100.0

    min_pct = float(os.getenv("BIODIV_MIN_TAXON_PCT", "1.0"))
    other_label = os.getenv("BIODIV_OTHER_LABEL", "Other")

    small = grouped[grouped["pct"] < min_pct].copy()
    large = grouped[grouped["pct"] >= min_pct].copy()

    other_count = small["count"].sum()

    if other_count > 0:
        other_row = pd.DataFrame(
            [
                {
                    "taxon": other_label,
                    "count": other_count,
                    "pct": other_count / total * 100.0,
                }
            ]
        )
        grouped = pd.concat([large, other_row], ignore_index=True)
    else:
        grouped = large

    # Sort again after adding Other.
    grouped = grouped.sort_values("count", ascending=False).reset_index(drop=True)

    # Optional: still limit the chart to top N visible labels.
    # If there are more than top_n categories above 1%, collapse the rest into Other too.
    top_n = int(os.getenv("BIODIV_TOP_N", "10"))

    if len(grouped) > top_n:
        existing_other = grouped[grouped["taxon"] == other_label].copy()
        main = grouped[grouped["taxon"] != other_label].copy()

        top = main.iloc[:top_n].copy()
        rest_count = main.iloc[top_n:]["count"].sum()

        if not existing_other.empty:
            rest_count += existing_other["count"].sum()

        if rest_count > 0:
            top = pd.concat(
                [
                    top,
                    pd.DataFrame(
                        [
                            {
                                "taxon": other_label,
                                "count": rest_count,
                                "pct": rest_count / total * 100.0,
                            }
                        ]
                    ),
                ],
                ignore_index=True,
            )

        grouped = top

    # Recalculate final percentages so labels always sum correctly.
    grouped["pct"] = grouped["count"] / total * 100.0

    # ---------- Figure ----------
    fig, ax = plt.subplots(figsize=(14, 10), facecolor=PIE_BG)
    ax.set_facecolor(PIE_BG)

    colors = PIE_COLORS[: len(grouped)]
    if len(colors) < len(grouped):
        # fallback if ever needed
        extra = ["#777777"] * (len(grouped) - len(colors))
        colors = colors + extra

    wedges, _ = ax.pie(
        grouped["pct"],
        startangle=90,
        colors=colors,
        labels=None,  # no labels directly on the pie
        counterclock=True,
        wedgeprops={
            "edgecolor": PIE_EDGE,
            "linewidth": 2.0,
        },
        radius=1.0,
    )

    ax.axis("equal")

    # ---------- Title ----------
    fig.suptitle(
        f"Top {min(top_n, len(grouped))} {level} — {sample_id} ({marker})",
        fontsize=22,
        fontweight="bold",
        color=PIE_TEXT,
        y=0.96,
    )

    # ---------- Legend ----------
    legend_labels = [
        f"{taxon} ({pct:.1f}%)" for taxon, pct in zip(grouped["taxon"], grouped["pct"])
    ]

    leg = ax.legend(
        wedges,
        legend_labels,
        loc="center left",
        bbox_to_anchor=(1.02, 0.5),
        frameon=False,
        fontsize=14,
        labelcolor=PIE_TEXT,
        handlelength=1.6,
        handleheight=1.6,
        borderaxespad=0.0,
    )

    # Some matplotlib versions ignore labelcolor above, so force it:
    for txt in leg.get_texts():
        txt.set_color(PIE_TEXT)

    # Remove axes junk
    ax.set_xticks([])
    ax.set_yticks([])

    out_path.parent.mkdir(parents=True, exist_ok=True)
    fig.tight_layout(rect=[0.02, 0.02, 0.82, 0.93])
    fig.savefig(
        out_path,
        dpi=300,
        bbox_inches="tight",
        facecolor=fig.get_facecolor(),
    )
    plt.close(fig)
    return True


def _current_raw_sample_sources(marker: str) -> pd.DataFrame:
    """
    Return the current structured-raw source and sample_index for every sample.

    sample_taxon_abundance.source_upload_id is the canonical pointer to the
    upload currently represented by the compact taxonomic data. If inconsistent
    historical rows ever exist, the newest biodiversity_uploads.uploaded_at wins.
    """
    sql = """
        WITH current_sample_sources AS (
            SELECT DISTINCT ON (
                UPPER(sta.sample_id),
                UPPER(sta.marker)
            )
                UPPER(sta.sample_id) AS sample_id,
                UPPER(sta.marker) AS marker,
                sta.source_upload_id AS upload_id,
                bu.uploaded_at
            FROM sample_taxon_abundance AS sta
            JOIN biodiversity_uploads AS bu
              ON bu.upload_id = sta.source_upload_id
            WHERE UPPER(sta.marker) = UPPER(%s)
              AND sta.source_upload_id IS NOT NULL
            ORDER BY
                UPPER(sta.sample_id),
                UPPER(sta.marker),
                bu.uploaded_at DESC,
                sta.source_upload_id DESC
        )
        SELECT
            css.sample_id,
            css.marker,
            css.upload_id,
            rs.sample_index
        FROM current_sample_sources AS css
        LEFT JOIN biodiversity_raw_samples AS rs
          ON rs.upload_id = css.upload_id
         AND UPPER(rs.sample_id) = css.sample_id
         AND UPPER(rs.marker) = css.marker
        ORDER BY css.sample_id
    """

    with get_pg_conn() as conn:
        return pd.read_sql(sql, conn, params=[marker])


def fetch_current_raw_faprotax_data(
    marker: str = "16S",
    min_prev: int = 2,
    min_total: int = 50,
) -> tuple[pd.DataFrame, list[str], int]:
    """
    Fetch current raw OTU counts/taxonomy for FAPROTAX.

    The raw importer stores each source file independently. Historical 16S
    uploads were split into several files containing different sample columns
    from the same OTU table, so source_feature_id is the cross-file OTU key.

    We therefore merge matching source_feature_id values across the *current*
    uploads and validate below that the same OTU ID does not carry conflicting
    taxonomy. Filtering is performed in PostgreSQL before pivoting to avoid
    constructing a huge dense unfiltered OTU x sample matrix in Python.
    """
    marker = marker.upper()

    sources = _current_raw_sample_sources(marker)
    if sources.empty:
        raise RuntimeError(
            f"No current sample_taxon_abundance sources found for marker={marker}"
        )

    missing_raw = sources[sources["sample_index"].isna()].copy()
    if not missing_raw.empty:
        missing_ids = sorted(missing_raw["sample_id"].astype(str).unique())
        preview = ", ".join(missing_ids[:20])
        raise RuntimeError(
            f"Structured raw data are incomplete for marker={marker}: "
            f"{len(missing_ids)} current samples are missing raw rows. "
            f"Examples: {preview}. Run the biodiversity raw backfill first."
        )

    current_sample_ids = sorted(
        sources["sample_id"].astype(str).str.upper().unique().tolist()
    )

    stats_sql = """
        WITH current_sample_sources AS (
            SELECT DISTINCT ON (
                UPPER(sta.sample_id),
                UPPER(sta.marker)
            )
                UPPER(sta.sample_id) AS sample_id,
                UPPER(sta.marker) AS marker,
                sta.source_upload_id AS upload_id,
                bu.uploaded_at
            FROM sample_taxon_abundance AS sta
            JOIN biodiversity_uploads AS bu
              ON bu.upload_id = sta.source_upload_id
            WHERE UPPER(sta.marker) = UPPER(%s)
              AND sta.source_upload_id IS NOT NULL
            ORDER BY
                UPPER(sta.sample_id),
                UPPER(sta.marker),
                bu.uploaded_at DESC,
                sta.source_upload_id DESC
        ),
        current_samples AS (
            SELECT
                css.sample_id,
                css.upload_id,
                rs.sample_index
            FROM current_sample_sources AS css
            JOIN biodiversity_raw_samples AS rs
              ON rs.upload_id = css.upload_id
             AND UPPER(rs.sample_id) = css.sample_id
             AND UPPER(rs.marker) = css.marker
        )
        SELECT COUNT(DISTINCT f.source_feature_id) AS otu_count
        FROM current_samples AS cs
        JOIN biodiversity_raw_abundance AS a
          ON a.upload_id = cs.upload_id
         AND a.sample_index = cs.sample_index
        JOIN biodiversity_raw_features AS f
          ON f.upload_id = a.upload_id
         AND f.feature_index = a.feature_index
        WHERE a.read_count > 0
          AND NULLIF(BTRIM(f.source_feature_id), '') IS NOT NULL
    """

    data_sql = """
        WITH current_sample_sources AS (
            SELECT DISTINCT ON (
                UPPER(sta.sample_id),
                UPPER(sta.marker)
            )
                UPPER(sta.sample_id) AS sample_id,
                UPPER(sta.marker) AS marker,
                sta.source_upload_id AS upload_id,
                bu.uploaded_at
            FROM sample_taxon_abundance AS sta
            JOIN biodiversity_uploads AS bu
              ON bu.upload_id = sta.source_upload_id
            WHERE UPPER(sta.marker) = UPPER(%s)
              AND sta.source_upload_id IS NOT NULL
            ORDER BY
                UPPER(sta.sample_id),
                UPPER(sta.marker),
                bu.uploaded_at DESC,
                sta.source_upload_id DESC
        ),
        current_samples AS (
            SELECT
                css.sample_id,
                css.upload_id,
                rs.sample_index
            FROM current_sample_sources AS css
            JOIN biodiversity_raw_samples AS rs
              ON rs.upload_id = css.upload_id
             AND UPPER(rs.sample_id) = css.sample_id
             AND UPPER(rs.marker) = css.marker
        ),
        current_counts AS (
            SELECT
                cs.sample_id,
                a.upload_id,
                a.feature_index,
                a.read_count
            FROM current_samples AS cs
            JOIN biodiversity_raw_abundance AS a
              ON a.upload_id = cs.upload_id
             AND a.sample_index = cs.sample_index
            WHERE a.read_count > 0
        ),
        eligible_otus AS (
            SELECT
                f.source_feature_id AS otu_id
            FROM current_counts AS c
            JOIN biodiversity_raw_features AS f
              ON f.upload_id = c.upload_id
             AND f.feature_index = c.feature_index
            WHERE NULLIF(BTRIM(f.source_feature_id), '') IS NOT NULL
            GROUP BY f.source_feature_id
            HAVING COUNT(DISTINCT c.sample_id) >= %s
               AND SUM(c.read_count) >= %s
        )
        SELECT
            c.sample_id,
            f.source_feature_id AS otu_id,
            c.read_count AS count,
            f.kingdom,
            f.phylum,
            f.class_name,
            f.order_name,
            f.family,
            f.genus,
            f.species
        FROM current_counts AS c
        JOIN biodiversity_raw_features AS f
          ON f.upload_id = c.upload_id
         AND f.feature_index = c.feature_index
        JOIN eligible_otus AS e
          ON e.otu_id = f.source_feature_id
        ORDER BY
            f.source_feature_id,
            c.sample_id
    """

    with get_pg_conn() as conn:
        stats = pd.read_sql(stats_sql, conn, params=[marker])
        df = pd.read_sql(
            data_sql,
            conn,
            params=[marker, int(min_prev), int(min_total)],
        )

    otu_before = int(stats.iloc[0]["otu_count"] or 0) if not stats.empty else 0

    if df.empty:
        raise RuntimeError(
            f"No current raw OTUs survive FAPROTAX filtering for marker={marker} "
            f"(min_prev={min_prev}, min_total={min_total})"
        )

    df["sample_id"] = df["sample_id"].astype(str).str.strip().str.upper()
    df["otu_id"] = df["otu_id"].astype(str).str.strip()
    df["count"] = pd.to_numeric(df["count"], errors="coerce").fillna(0)

    # The historical *_part_*.csv files are sample-column partitions of the
    # same OTU table. Reusing source_feature_id across them is intentional, but
    # it is only safe when each OTU ID has one taxonomy lineage.
    tax_cols = [
        "kingdom",
        "phylum",
        "class_name",
        "order_name",
        "family",
        "genus",
        "species",
    ]
    tax_unique = df[["otu_id", *tax_cols]].drop_duplicates()
    conflicting = (
        tax_unique.groupby("otu_id", sort=False)
        .size()
        .loc[lambda x: x > 1]
    )
    if not conflicting.empty:
        examples = ", ".join(conflicting.index.astype(str).tolist()[:20])
        raise RuntimeError(
            "Cannot safely merge raw source files for FAPROTAX: "
            f"{len(conflicting)} source_feature_id values have conflicting "
            f"taxonomy across current uploads. Examples: {examples}"
        )

    return df, current_sample_ids, otu_before


def build_clean_otu_and_taxonomy_files(
    marker: str = "16S",
    out_dir: Path | None = None,
    min_prev: int = 2,
    min_total: int = 50,
) -> tuple[Path, Path]:
    """
    Build R-compatible FAPROTAX inputs from the current structured raw tables.

    Produces:
      - 6_otu_clean_counts_no_blanks.csv
      - 7_taxonomy_clean.csv

    Only the source_upload_id currently referenced by sample_taxon_abundance is
    used for each sample. Historical/replaced uploads are therefore excluded.

    Filtering is equivalent to the previous implementation:
      - keep OTUs present in at least min_prev samples
      - keep OTUs with at least min_total total reads
    """
    marker = marker.upper()
    if marker != "16S":
        raise ValueError("FAPROTAX input generation is intended for marker=16S")

    if out_dir is None:
        out_dir = PROJECT_ROOT / "data" / "biodiversity" / "faprotax_work"

    out_dir.mkdir(parents=True, exist_ok=True)

    print(
        f"[INFO] Building FAPROTAX inputs from current structured raw marker={marker} "
        f"(min_prev={min_prev}, min_total={min_total})"
    )

    df, current_sample_ids, otu_before = fetch_current_raw_faprotax_data(
        marker=marker,
        min_prev=min_prev,
        min_total=min_total,
    )

    print(f"[INFO] Current structured raw samples: {len(current_sample_ids)}")
    print(f"[INFO] OTUs before filtering: {otu_before}")
    print(f"[INFO] OTUs after filtering : {df['otu_id'].nunique()}")
    print(f"[INFO] Non-zero values after filtering: {len(df)}")

    # ------------------------------------------------------------------
    # 1) OTU count matrix: rows = OTU IDs, columns = current sample IDs
    # ------------------------------------------------------------------
    otu_clean = df.pivot_table(
        index="otu_id",
        columns="sample_id",
        values="count",
        aggfunc="sum",
        fill_value=0,
    ).sort_index()

    # Keep the complete current sample set in a stable order. Normally every
    # sample has at least one retained OTU; zero-only samples are reported.
    otu_clean = otu_clean.reindex(
        columns=current_sample_ids,
        fill_value=0,
    )

    zero_samples = [
        str(sample_id)
        for sample_id, total in otu_clean.sum(axis=0).items()
        if float(total) <= 0
    ]
    if zero_samples:
        print(
            "[WARN] Current samples with no reads after OTU filtering: "
            f"{len(zero_samples)}; examples={zero_samples[:20]}"
        )

    # ------------------------------------------------------------------
    # 2) Taxonomy table for the retained OTUs
    # ------------------------------------------------------------------
    tax_cols = [
        "kingdom",
        "phylum",
        "class_name",
        "order_name",
        "family",
        "genus",
        "species",
    ]
    tax_source = (
        df[["otu_id", *tax_cols]]
        .drop_duplicates(subset=["otu_id"])
        .set_index("otu_id")
        .reindex(otu_clean.index)
    )

    tax_df = tax_source.rename(
        columns={
            "kingdom": "Kingdom",
            "phylum": "Phylum",
            "class_name": "Class",
            "order_name": "Order",
            "family": "Family",
            "genus": "Genus",
            "species": "Species",
        }
    ).fillna("")
    tax_df.index.name = "OTU_ID"

    # ------------------------------------------------------------------
    # 3) Write files expected by the R/FAPROTAX workflow
    # ------------------------------------------------------------------
    otu_path = out_dir / "6_otu_clean_counts_no_blanks.csv"
    tax_path = out_dir / "7_taxonomy_clean.csv"

    otu_clean.to_csv(otu_path)
    tax_df.to_csv(tax_path, sep=";")

    print(f"[OK] Wrote {otu_path}")
    print(f"[OK] Wrote {tax_path}")
    print(
        f"[OK] FAPROTAX input matrix: {otu_clean.shape[0]} OTUs x "
        f"{otu_clean.shape[1]} samples"
    )

    return otu_path, tax_path


# ---------------------------------------------------------------------------
# Fungal ecological guild plots, based on FUNGuild genus-level assignments
# ---------------------------------------------------------------------------

FUNGUILD_KEEP_CONFIDENCE = {
    "Probable",
    "Highly Probable",
    "Higly Probable",  # typo present in some FUNGuild outputs
}

FUNGUILD_CONF_SCORE = {
    "Highly Probable": 3,
    "Higly Probable": 3,
    "Probable": 2,
    "Possible": 1,
}

FUNGAL_GUILD_MACRO_MAP = {
    "Ectomycorrhizal": "Ectomycorrhizal fungi",
    "Arbuscular Mycorrhizal": "Arbuscular mycorrhizal fungi",
    "Ericoid Mycorrhizal": "Mycorrhizal fungi",
    "Orchid Mycorrhizal": "Mycorrhizal fungi",
    "Wood Saprotroph": "Wood decomposers",
    "Litter Saprotroph": "Litter decomposers",
    "Plant Saprotroph": "Plant litter decomposers",
    "Dung Saprotroph": "Dung decomposers",
    "Undefined Saprotroph": "Decomposers (unspecified)",
    "Plant Pathogen": "Plant pathogens",
    "Animal Pathogen": "Animal pathogens",
    "Animal Parasite": "Animal pathogens",
    "Endophyte": "Endophytes",
    "Fungal Parasite": "Fungal parasites",
    "Lichen Parasite": "Lichen parasites",
    "Lichenized": "Lichenized fungi",
    "Nematophagous": "Nematophagous fungi",
    "Algal Parasite": "Algal parasites",
    "Insect Pathogen": "Insect pathogens",
    "Epiphyte": "Endophytes",
    "Pollen Saprotroph": "Decomposers (unspecified)",
}

FUNGAL_GUILD_ORDER = [
    "Ectomycorrhizal fungi",
    "Arbuscular mycorrhizal fungi",
    "Mycorrhizal fungi",
    "Wood decomposers",
    "Litter decomposers",
    "Plant litter decomposers",
    "Dung decomposers",
    "Decomposers (unspecified)",
    "Plant pathogens",
    "Animal pathogens",
    "Endophytes",
    "Fungal parasites",
    "Lichen parasites",
    "Lichenized fungi",
    "Nematophagous fungi",
    "Algal parasites",
    "Insect pathogens",
]

FUNGAL_GUILD_COLORS = {
    "Ectomycorrhizal fungi": "#264653",
    "Arbuscular mycorrhizal fungi": "#2A9D8F",
    "Mycorrhizal fungi": "#457B9D",
    "Wood decomposers": "#8B5E3C",
    "Litter decomposers": "#C9A96E",
    "Plant litter decomposers": "#E9C46A",
    "Dung decomposers": "#A8DADC",
    "Decomposers (unspecified)": "#BDB2A7",
    "Plant pathogens": "#E76F51",
    "Animal pathogens": "#F4A261",
    "Endophytes": "#6A994E",
    "Fungal parasites": "#BC6C25",
    "Lichen parasites": "#8D99AE",
    "Lichenized fungi": "#CDB4DB",
    "Nematophagous fungi": "#FFAFCC",
    "Algal parasites": "#D4E09B",
    "Insect pathogens": "#F08080",
}

# ---------------------------------------------------------------------------
# Bacterial ecological guild plots from FAPROTAX output
# ---------------------------------------------------------------------------

BACTERIAL_SOIL_CORE = [
    # Nitrogen
    "nitrogen_fixation",
    "nitrification",
    "aerobic_ammonia_oxidation",
    "nitrate_reduction",
    "nitrate_respiration",
    "nitrite_respiration",
    "nitrogen_respiration",
    "ureolysis",
    # Sulfur
    "sulfate_respiration",
    "sulfur_respiration",
    "sulfite_respiration",
    "respiration_of_sulfur_compounds",
    "dark_sulfide_oxidation",
    "dark_oxidation_of_sulfur_compounds",
    # Methane / C1
    "methanotrophy",
    "methanol_oxidation",
    "methylotrophy",
    "methanogenesis",
    "hydrogenotrophic_methanogenesis",
    "methanogenesis_by_reduction_of_methyl_compounds_with_H2",
    # Carbon degradation
    "cellulolysis",
    "xylanolysis",
    "aromatic_compound_degradation",
    "aromatic_hydrocarbon_degradation",
    "hydrocarbon_degradation",
    "aliphatic_non_methane_hydrocarbon_degradation",
    # Heterotrophy
    "aerobic_chemoheterotrophy",
    "anaerobic_chemoheterotrophy",
    "fermentation",
    # Mineral cycling
    "iron_respiration",
    "dark_iron_oxidation",
    "manganese_oxidation",
    # Pathogens / parasites / predation
    "plant_pathogen",
    "animal_parasite_or_symbiont",
    "predatory_or_exoparasitic",
    "chitinolysis",
    "nitrous_oxide_denitrification",
    "ligninolysis",
    "dark_hydrogen_oxidation",
    "phototrophy",
    "photoautotrophy",
    "cyanobacteria",
]

BACTERIAL_MACRO_MAP = {
    "chitinolysis": "Chitinolytic bacteria",
    "nitrogen_fixation": "Nitrogen fixers",
    "nitrification": "Nitrifiers",
    "aerobic_ammonia_oxidation": "Nitrifiers",
    "nitrate_reduction": "Denitrifiers",
    "nitrate_respiration": "Denitrifiers",
    "nitrite_respiration": "Denitrifiers",
    "nitrogen_respiration": "Denitrifiers",
    "nitrous_oxide_denitrification": "Denitrifiers",
    "ureolysis": "Ureolytic bacteria",
    "aromatic_compound_degradation": "Hydrocarbon degraders",
    "aromatic_hydrocarbon_degradation": "Hydrocarbon degraders",
    "hydrocarbon_degradation": "Hydrocarbon degraders",
    "aliphatic_non_methane_hydrocarbon_degradation": "Hydrocarbon degraders",
    "methanotrophy": "Methanotrophs",
    "methanol_oxidation": "Methanotrophs",
    "methylotrophy": "Methanotrophs",
    "methanogenesis": "Methanogens",
    "hydrogenotrophic_methanogenesis": "Methanogens",
    "methanogenesis_by_reduction_of_methyl_compounds_with_H2": "Methanogens",
    "dark_sulfide_oxidation": "Sulfur oxidizers",
    "dark_oxidation_of_sulfur_compounds": "Sulfur oxidizers",
    "sulfate_respiration": "Sulfate reducers",
    "sulfur_respiration": "Sulfate reducers",
    "sulfite_respiration": "Sulfate reducers",
    "respiration_of_sulfur_compounds": "Sulfate reducers",
    "iron_respiration": "Iron & Manganese cyclers",
    "dark_iron_oxidation": "Iron & Manganese cyclers",
    "manganese_oxidation": "Iron & Manganese cyclers",
    "fermentation": "Anaerobic heterotrophs",
    "aerobic_chemoheterotrophy": "Aerobic heterotrophs",
    "anaerobic_chemoheterotrophy": "Anaerobic heterotrophs",
    "plant_pathogen": "Plant pathogens",
    "animal_parasite_or_symbiont": "Animal parasites",
    "predatory_or_exoparasitic": "Predatory bacteria",
    "ligninolysis": "Lignocellulose degraders",
    "cellulolysis": "Lignocellulose degraders",
    "xylanolysis": "Lignocellulose degraders",
    "dark_hydrogen_oxidation": "Hydrogen oxidizers",
    "phototrophy": "Phototrophs",
    "photoautotrophy": "Phototrophs",
    "cyanobacteria": "Phototrophs",
}

BACTERIAL_GUILD_ORDER = [
    "Aerobic heterotrophs",
    "Anaerobic heterotrophs",
    "Nitrogen fixers",
    "Nitrifiers",
    "Denitrifiers",
    "Ureolytic bacteria",
    "Hydrocarbon degraders",
    "Methanotrophs",
    "Methanogens",
    "Sulfur oxidizers",
    "Sulfate reducers",
    "Iron & Manganese cyclers",
    "Plant pathogens",
    "Animal parasites",
    "Predatory bacteria",
    "Chitinolytic bacteria",
    "Lignocellulose degraders",
    "Hydrogen oxidizers",
    "Phototrophs",
]

BACTERIAL_GUILD_COLORS = {
    "Nitrogen fixers": "#264653",
    "Nitrifiers": "#2A9D8F",
    "Denitrifiers": "#457B9D",
    "Ureolytic bacteria": "#A8DADC",
    "Hydrocarbon degraders": "#C9A96E",
    "Methanotrophs": "#6A994E",
    "Methanogens": "#386641",
    "Sulfur oxidizers": "#FBF259",
    "Sulfate reducers": "#E9C46A",
    "Iron & Manganese cyclers": "#8D99AE",
    "Aerobic heterotrophs": "#E76F51",
    "Anaerobic heterotrophs": "#F4A261",
    "Plant pathogens": "#D62828",
    "Animal parasites": "#F08080",
    "Predatory bacteria": "#BC6C25",
    "Chitinolytic bacteria": "#CDB4DB",
    "Lignocellulose degraders": "#8B5E3C",
    "Hydrogen oxidizers": "#577590",
    "Phototrophs": "#D4E09B",
}


def _taxa_to_dict(taxa) -> dict:
    """
    sample_otu_counts.taxa may arrive as dict, JSON string, or None.
    """
    if isinstance(taxa, dict):
        return taxa
    if taxa is None:
        return {}
    if isinstance(taxa, str):
        s = taxa.strip()
        if not s:
            return {}
        try:
            obj = json.loads(s)
            return obj if isinstance(obj, dict) else {}
        except Exception:
            return {}
    return {}


def _strip_tax_prefix(value: str) -> str:
    """
    Convert g__Fusarium -> Fusarium, f__Nectriaceae -> Nectriaceae.
    Also replaces underscores with spaces.
    """
    if value is None:
        return ""
    s = str(value).strip()
    s = re.sub(r"^[a-zA-Z]__", "", s)
    s = s.replace("_", " ").strip()
    if not s:
        return ""
    if re.search(r"incertae|unclassified|uncultured", s, flags=re.I):
        return ""
    return s


def extract_fungal_genus(row: pd.Series) -> str:
    """
    Extract genus for FUNGuild matching.

    Supports both:
      - raw taxonomy string: k__;p__;c__;o__;f__;g__;s__
      - split fields: A/B/C/D/E/F
      - named rank fields: Genus
    """
    taxa = taxa_to_normalized_dict(row.get("taxa"))
    return taxa.get("Genus", "").strip()


def extract_primary_guild(guild_name: str) -> str:
    """
    FUNGuild guild names may contain a primary guild between pipes:
      Something-|Plant Pathogen|-Something
    If there are no pipes, fall back to the raw value or parts split by '-'.
    """
    if guild_name is None:
        return ""

    s = str(guild_name).strip()
    if not s or s.upper() == "NULL":
        return ""

    m = re.search(r"\|([^|]+)\|", s)
    if m:
        return m.group(1).strip()

    # fallback: try direct match first
    if s in FUNGAL_GUILD_MACRO_MAP:
        return s

    # fallback: split compound guilds
    for part in re.split(r"\s*-\s*", s):
        part = part.strip().replace("|", "")
        if part in FUNGAL_GUILD_MACRO_MAP:
            return part

    return s.replace("|", "").strip()


def load_funguild_best_by_genus() -> dict[str, dict]:
    """
    Load local FUNGuild_db.json and keep one best assignment per genus.

    Returns:
      {
        "Fusarium": {
          "guild": "...",
          "primary_guild": "...",
          "macro": "Plant pathogens",
          ...
        }
      }
    """
    path = Path(FUNGUILD_DB_JSON)
    if not path.exists():
        raise FileNotFoundError(
            f"FUNGuild DB JSON not found: {path}. Set FUNGUILD_DB_JSON or place the file there."
        )

    with path.open("r", encoding="utf-8") as f:
        raw = json.load(f)

    if isinstance(raw, dict):
        # Some JSON exports are dict-like. Try values.
        records = list(raw.values())
    elif isinstance(raw, list):
        records = raw
    else:
        raise ValueError(f"Unsupported FUNGuild JSON structure in {path}")

    best: dict[str, dict] = {}

    for rec in records:
        if not isinstance(rec, dict):
            continue

        taxon = str(rec.get("taxon") or rec.get("queried_taxon") or "").strip()
        if not taxon:
            continue

        confidence = str(rec.get("confidenceRanking") or "").strip()
        if confidence not in FUNGUILD_KEEP_CONFIDENCE:
            continue

        trophic = str(rec.get("trophicMode") or "").strip()
        guild = str(rec.get("guild") or "").strip()

        if not trophic or trophic.upper() == "NULL":
            continue
        if not guild or guild.upper() == "NULL":
            continue

        primary = extract_primary_guild(guild)
        macro = FUNGAL_GUILD_MACRO_MAP.get(primary)

        # Only keep guilds that map to citizen-friendly categories.
        if not macro:
            continue

        score = FUNGUILD_CONF_SCORE.get(confidence, 0)

        prev = best.get(taxon)
        if prev is None or score > prev["score"]:
            best[taxon] = {
                "taxon": taxon,
                "confidence": confidence,
                "score": score,
                "trophicMode": trophic,
                "guild": guild,
                "primary_guild": primary,
                "macro": macro,
            }

    print(f"[INFO] Loaded FUNGuild best assignments for {len(best)} genera")
    return best


def make_bacterial_guildplot_for_sample(
    sample_id: str,
    func_row: pd.Series,
    out_path: Path,
) -> bool:
    """
    Create one citizen-friendly bacterial ecological guild plot from one
    FAPROTAX sample row.

    Input values are expected to be FAPROTAX fractions, as in the R script.
    If values look like percentages already, the function handles that too.
    """
    values = {}

    for func_name, raw_val in func_row.items():
        if func_name not in BACTERIAL_SOIL_CORE:
            continue

        guild = BACTERIAL_MACRO_MAP.get(func_name)
        if not guild:
            continue

        try:
            v = float(raw_val)
        except Exception:
            continue

        if not math.isfinite(v) or v <= 0:
            continue

        values[guild] = values.get(guild, 0.0) + v

    if not values:
        return False

    df = pd.DataFrame([{"guild": k, "value": v} for k, v in values.items()])

    # R script does Percent = 100 * sum(Value).
    # But if the CSV already contains percentages, avoid multiplying again.
    max_v = df["value"].max()
    if max_v <= 1.5:
        df["percent"] = df["value"] * 100.0
    else:
        df["percent"] = df["value"]

    df = df[df["percent"] >= 1.0].copy()
    if df.empty:
        return False

    order_index = {name: i for i, name in enumerate(BACTERIAL_GUILD_ORDER)}
    df["order"] = df["guild"].map(lambda x: order_index.get(x, 999))
    df = df.sort_values(["order", "percent"], ascending=[True, False])

    # barh draws bottom-to-top, so reverse for top-to-bottom display.
    df = df.iloc[::-1].copy()

    labels = df["guild"].tolist()
    values_pct = df["percent"].tolist()
    colors = [BACTERIAL_GUILD_COLORS.get(label, "#999999") for label in labels]

    out_path.parent.mkdir(parents=True, exist_ok=True)

    fig_height = max(6.5, 0.55 * len(df) + 2.2)
    fig, ax = plt.subplots(figsize=(10, fig_height))

    # White background + black text style
    fig.patch.set_facecolor("white")
    ax.set_facecolor("white")

    y_pos = list(range(len(labels)))

    ax.barh(
        y_pos,
        values_pct,
        color=colors,
        height=0.75,
        edgecolor="none",
    )

    ax.set_yticks(y_pos)
    ax.set_yticklabels(labels, fontsize=11, color="#333333")

    x_max = max(max(values_pct) * 1.12, 20)
    ax.set_xlim(0, x_max)

    if x_max <= 20:
        ticks = [0, 5, 10, 15, 20]
    else:
        step = 5
        ticks = list(range(0, int(x_max + step), step))

    ax.set_xticks(ticks)
    ax.set_xticklabels([f"{t}%" for t in ticks], fontsize=10, color="#333333")

    ax.set_xlabel(
        "% of bacterial community",
        fontsize=12,
        color="#111111",
        labelpad=8,
    )

    ax.xaxis.grid(True, color="#dddddd", linewidth=1)
    ax.yaxis.grid(False)
    ax.set_axisbelow(True)

    for spine in ax.spines.values():
        spine.set_visible(False)

    ax.tick_params(axis="both", length=0)

    for y, pct in zip(y_pos, values_pct, strict=False):
        ax.text(
            pct + x_max * 0.006,
            y,
            f"{pct:.1f}%",
            va="center",
            ha="left",
            fontsize=11,
            color="#222222",
        )

    ax.set_title(
        "Soil bacterial ecological guilds\n",
        loc="left",
        fontsize=15,
        fontweight="bold",
        color="#111111",
        pad=8,
    )

    ax.text(
        0,
        1.02,
        "Guild-level functional categories",
        transform=ax.transAxes,
        ha="left",
        va="bottom",
        fontsize=11,
        color="#555555",
    )

    fig.suptitle(
        f"Your soil bacteria at a glance — Sample: {sample_id}",
        x=0.02,
        y=0.98,
        ha="left",
        va="top",
        fontsize=16,
        fontweight="bold",
        color="#111111",
    )

    caption = (
        "Values indicate the estimated percentage of the bacterial community associated with each ecological guild.\n"
        "Only guilds exceeding 1% are shown; absent categories may reflect low detection rather than true absence.\n"
        "Guild assignments are based on FAPROTAX (Louca et al. 2016)."
    )

    fig.text(
        0.02,
        0.025,
        caption,
        ha="left",
        va="bottom",
        fontsize=8.5,
        color="#666666",
    )

    fig.tight_layout(rect=[0.02, 0.09, 0.98, 0.92])
    fig.savefig(out_path, dpi=300, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)

    return True


def make_fungal_guildplot_for_sample(
    sample_df: pd.DataFrame,
    sample_id: str,
    funguild_by_genus: dict[str, dict],
    out_path: Path,
) -> bool:
    """
    Create one citizen-friendly fungal guild horizontal bar plot.

    Percentages are relative to the full fungal community for the sample,
    matching the R script interpretation.

    Output style mirrors the R/ggplot example:
      - horizontal bars
      - fixed macro-category colours
      - title + subtitle
      - percentage labels at bar ends
      - explanatory caption
    """
    plot_df = sample_df.copy()
    plot_df["count"] = pd.to_numeric(plot_df["count"], errors="coerce").fillna(0)
    plot_df = plot_df[plot_df["count"] > 0].copy()

    if plot_df.empty:
        return False

    total = plot_df["count"].sum()
    if total <= 0:
        return False

    # Extract genus and map to citizen-friendly guild macro-category
    plot_df["genus"] = plot_df.apply(extract_fungal_genus, axis=1)
    plot_df["macro"] = plot_df["genus"].map(lambda g: funguild_by_genus.get(g, {}).get("macro", ""))

    annotated = plot_df[plot_df["macro"].astype(str).str.strip() != ""].copy()
    if annotated.empty:
        return False

    grouped = annotated.groupby("macro", dropna=False)["count"].sum().reset_index()
    grouped["percent"] = grouped["count"] / total * 100.0

    # Same communication threshold as the R script
    grouped = grouped[grouped["percent"] >= 1.0].copy()
    if grouped.empty:
        return False

    # Keep the same category order as the R script.
    order_index = {name: i for i, name in enumerate(FUNGAL_GUILD_ORDER)}
    grouped["order"] = grouped["macro"].map(lambda x: order_index.get(x, 999))
    grouped = grouped.sort_values(["order", "percent"], ascending=[True, False])

    # Matplotlib barh draws bottom-to-top, so reverse for top-to-bottom display.
    grouped = grouped.iloc[::-1].copy()

    labels = grouped["macro"].tolist()
    values = grouped["percent"].tolist()
    colors = [FUNGAL_GUILD_COLORS.get(label, "#999999") for label in labels]

    out_path.parent.mkdir(parents=True, exist_ok=True)

    # Similar aspect to the R output: wide, communication-oriented.
    fig_height = max(5.5, 0.55 * len(grouped) + 2.2)
    fig, ax = plt.subplots(figsize=(10, fig_height))

    # Warm, clean background like ggplot/theme_minimal
    fig.patch.set_facecolor("#f7f7f5")
    ax.set_facecolor("#f7f7f5")

    y_pos = list(range(len(labels)))

    ax.barh(
        y_pos,
        values,
        color=colors,
        height=0.75,
        edgecolor="none",
    )

    ax.set_yticks(y_pos)
    ax.set_yticklabels(labels, fontsize=11, color="#4d4d4d")

    # Axis max: at least 20%, otherwise 12% extra headroom
    x_max = max(max(values) * 1.12, 20)
    ax.set_xlim(0, x_max)

    # Use 0/5/10/15/20 style ticks where possible
    if x_max <= 20:
        ticks = [0, 5, 10, 15, 20]
    else:
        step = 5
        ticks = list(range(0, int(x_max + step), step))

    ax.set_xticks(ticks)
    ax.set_xticklabels([f"{t}%" for t in ticks], fontsize=10, color="#555555")

    ax.set_xlabel(
        "% of fungal community",
        fontsize=12,
        color="#111111",
        labelpad=8,
    )

    # Subtle vertical gridlines
    ax.xaxis.grid(True, color=PIE_GRID, linewidth=1)
    ax.yaxis.grid(False)
    ax.set_axisbelow(True)

    # Remove plot frame for ggplot-like look
    for spine in ax.spines.values():
        spine.set_visible(False)

    ax.tick_params(axis="both", length=0)

    # Percentage labels at bar ends
    for y, pct in zip(y_pos, values, strict=False):
        ax.text(
            pct + x_max * 0.006,
            y,
            f"{pct:.1f}%",
            va="center",
            ha="left",
            fontsize=11,
            color="#333333",
        )

    # Main chart title and subtitle
    ax.set_title(
        "Fungal ecological guilds\n",
        loc="left",
        fontsize=15,
        fontweight="bold",
        color="#111111",
        pad=8,
    )

    ax.text(
        0,
        1.02,
        "Guild macro-categories",
        transform=ax.transAxes,
        ha="left",
        va="bottom",
        fontsize=11,
        color="#666666",
    )

    # Figure-level top title, like the R patchwork annotation
    fig.suptitle(
        f"Your soil fungi at a glance — Sample: {sample_id}",
        x=0.02,
        y=0.98,
        ha="left",
        va="top",
        fontsize=16,
        fontweight="bold",
        color="#111111",
    )

    # Footer caption
    caption = (
        "Values indicate the estimated percentage of the fungal community associated with each ecological guild.\n"
        "Only guilds exceeding 1% are shown; absent categories may reflect low detection rather than true absence."
    )
    fig.text(
        0.02,
        0.025,
        caption,
        ha="left",
        va="bottom",
        fontsize=8.5,
        color="#777777",
    )

    # Leave room for suptitle and footer
    fig.tight_layout(rect=[0.02, 0.08, 0.98, 0.92])

    fig.savefig(out_path, dpi=300, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)
    return True


def generate_bacterial_guildplots_from_faprotax(
    mclient,
    *,
    force: bool = False,
    sample_ids: set[str] | None = None,
    dry_run: bool = False,
) -> tuple[int, int]:
    """
    Generate bacterial ecological guild plots from a FAPROTAX sample x function CSV.

    Existing MinIO objects are skipped unless force=True.

    Expected input:
      rows    = sample IDs
      columns = FAPROTAX function names
      values  = fractions or percentages

    Uploads to:
      biodiversity/guildplots/bacteria/<sample_id>.png
    """
    path = Path(FAPROTAX_FUNCTION_CSV)
    if not path.exists():
        raise FileNotFoundError(
            f"FAPROTAX function CSV not found: {path}. "
            "Set FAPROTAX_FUNCTION_CSV or place the file there."
        )

    print(f"[INFO] Loading FAPROTAX functions from {path}")

    func_sxf = pd.read_csv(path, index_col=0)

    if func_sxf.empty:
        print("[INFO] FAPROTAX function matrix is empty; skipping bacterial guild plots.")
        return 0, 0

    print(f"[INFO] FAPROTAX matrix: {func_sxf.shape[0]} samples x {func_sxf.shape[1]} functions")

    available = set(func_sxf.columns)
    selected = [c for c in BACTERIAL_SOIL_CORE if c in available]

    print(f"[INFO] Relevant FAPROTAX functions present: {len(selected)}")
    if selected:
        print("[INFO] First relevant functions:", ", ".join(selected[:20]))

    if not selected:
        print("[WARN] No expected FAPROTAX soil functions found in the CSV.")
        return 0, 0

    prefix = "biodiversity/guildplots/bacteria/"
    existing_objects = set() if force else list_existing_minio_objects(mclient, prefix)

    out_dir = PROJECT_ROOT / "data" / "bacterial_guildplots"
    out_dir.mkdir(parents=True, exist_ok=True)

    generated = 0
    uploaded = 0
    skipped_existing = 0
    selected_samples = 0

    for sample_id, row in func_sxf.iterrows():
        sample_id = str(sample_id).strip()
        if not sample_id:
            continue

        if sample_ids and sample_id.upper() not in sample_ids:
            continue

        selected_samples += 1
        safe_id = sanitize_filename(sample_id)
        object_name = f"{prefix}{safe_id}.png"

        if not force and object_name in existing_objects:
            skipped_existing += 1
            continue

        if dry_run:
            print(f"[NEW] would generate {object_name}")
            continue

        local_png = out_dir / f"{safe_id}.png"

        ok = make_bacterial_guildplot_for_sample(
            sample_id=sample_id,
            func_row=row[selected],
            out_path=local_png,
        )
        if not ok:
            continue

        generated += 1

        uploaded_url = upload_file_to_minio(
            mclient,
            local_png,
            object_name,
            content_type="image/png",
        )
        if uploaded_url:
            uploaded += 1

    print(f"[OK] Considered {selected_samples} bacterial guild samples")
    print(f"[OK] Skipped {skipped_existing} existing bacterial guild plots")
    print(f"[OK] Generated {generated} bacterial guild plots")
    print(f"[OK] Uploaded {uploaded} bacterial guild plots to MinIO")
    return generated, uploaded


def generate_fungal_guildplots(
    mclient,
    *,
    force: bool = False,
    sample_ids: set[str] | None = None,
    dry_run: bool = False,
) -> tuple[int, int]:
    """
    Generate fungal ecological guild plots from current raw ITS archives,
    with legacy OTU rows used only for samples that have no current upload.

    Existing MinIO objects are skipped unless force=True.

    Uploads to:
      biodiversity/guildplots/fungi/<sample_id>.png
    """
    marker = "ITS"
    print("[INFO] Generating fungal ecological guild plots from ITS data")

    df = fetch_fungal_guild_source_data(
        mclient,
        sample_ids=sample_ids,
    )
    if df.empty:
        print("[INFO] No ITS data available for fungal guild plots.")
        return 0, 0

    funguild_by_genus = load_funguild_best_by_genus()

    all_genera = df.apply(extract_fungal_genus, axis=1)
    nonempty = all_genera[all_genera.astype(str).str.strip() != ""]
    matched = nonempty.map(lambda g: g in funguild_by_genus)

    print(f"[DEBUG] Extracted non-empty genera: {len(nonempty)}")
    print(f"[DEBUG] Unique extracted genera: {nonempty.nunique()}")
    print(f"[DEBUG] FUNGuild genus matches: {matched.sum()} / {len(nonempty)}")

    prefix = "biodiversity/guildplots/fungi/"
    existing_objects = set() if force else list_existing_minio_objects(mclient, prefix)

    out_dir = PROJECT_ROOT / "data" / "biodiversity_guildplots" / "fungi"
    out_dir.mkdir(parents=True, exist_ok=True)

    generated = 0
    uploaded = 0
    skipped_existing = 0

    for sample_id, sample_df in df.groupby("sample_id"):
        sample_id = str(sample_id).strip()
        if not sample_id:
            continue

        safe_id = sanitize_filename(sample_id)
        object_name = f"{prefix}{safe_id}.png"

        if not force and object_name in existing_objects:
            skipped_existing += 1
            continue

        if dry_run:
            print(f"[NEW] would generate {object_name}")
            continue

        local_png = out_dir / f"{safe_id}.png"

        ok = make_fungal_guildplot_for_sample(
            sample_df=sample_df,
            sample_id=sample_id,
            funguild_by_genus=funguild_by_genus,
            out_path=local_png,
        )
        if not ok:
            continue

        generated += 1

        uploaded_url = upload_file_to_minio(
            mclient,
            local_png,
            object_name,
            content_type="image/png",
        )
        if uploaded_url:
            uploaded += 1

    print(f"[OK] Skipped {skipped_existing} existing fungal guild plots")
    print(f"[OK] Generated {generated} fungal guild plots")
    print(f"[OK] Uploaded {uploaded} fungal guild plots to MinIO")
    return generated, uploaded


def parse_args():
    parser = argparse.ArgumentParser(
        description=(
            "Generate biodiversity charts that are missing from MinIO. "
            "Existing objects are skipped by default."
        )
    )
    parser.add_argument(
        "--marker",
        default=os.getenv("BIODIV_MARKER", "16S"),
        choices=("16S", "ITS", "16s", "its"),
        help="Marker for taxonomic pie charts (default: BIODIV_MARKER or 16S).",
    )
    parser.add_argument(
        "--level",
        default=os.getenv("BIODIV_LEVEL", "Phylum"),
        help="Taxonomic level (current compact importer stores Phylum).",
    )
    parser.add_argument(
        "--sample-id",
        action="append",
        help=(
            "Generate only selected sample IDs. May be repeated or contain "
            "comma-separated IDs."
        ),
    )
    parser.add_argument(
        "--force",
        action="store_true",
        default=os.getenv("BIODIV_FORCE_REGENERATE", "0") == "1",
        help="Regenerate and overwrite charts even when they already exist.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Show missing chart objects without creating or uploading them.",
    )
    parser.add_argument(
        "--fungal-guilds",
        action="store_true",
        default=GENERATE_FUNGAL_GUILDS,
        help=(
            "Generate ITS fungal ecological guild charts. "
            "Can also be enabled with GENERATE_FUNGAL_GUILDS=1."
        ),
    )
    parser.add_argument(
        "--bacterial-guilds",
        action="store_true",
        default=GENERATE_BACTERIAL_GUILDS,
        help=(
            "Generate 16S bacterial ecological guild charts from the "
            "configured FAPROTAX sample-by-function CSV. Can also be enabled "
            "with GENERATE_BACTERIAL_GUILDS=1."
        ),
    )
    parser.add_argument(
        "--build-faprotax-inputs",
        action="store_true",
        default=BUILD_FAPROTAX_INPUTS,
        help=(
            "Build FAPROTAX OTU-count and taxonomy input files. "
            "Can also be enabled with BUILD_FAPROTAX_INPUTS=1."
        ),
    )
    return parser.parse_args()


def main():
    args = parse_args()

    marker = args.marker.upper()
    level = normalize_taxonomic_level(args.level)
    sample_ids = normalize_sample_filter(args.sample_id)

    out_dir = PROJECT_ROOT / "data" / "biodiversity_piecharts" / marker / level
    out_dir.mkdir(parents=True, exist_ok=True)

    print(
        f"[INFO] marker={marker} level={level} "
        f"force={args.force} dry_run={args.dry_run}"
    )
    if sample_ids:
        print(f"[INFO] sample filter: {sorted(sample_ids)}")

    if args.build_faprotax_inputs:
        build_clean_otu_and_taxonomy_files(
            marker=marker,
            out_dir=Path(
                os.getenv(
                    "FAPROTAX_WORK_DIR",
                    str(PROJECT_ROOT / "data" / "biodiversity" / "faprotax_work"),
                )
            ),
            min_prev=int(os.getenv("FAPROTAX_MIN_PREV", "2")),
            min_total=int(os.getenv("FAPROTAX_MIN_TOTAL", "50")),
        )

        # When this flag is used on its own, it is a data-preparation command.
        # Do not require MinIO or regenerate/inspect chart images unnecessarily.
        if not args.fungal_guilds and not args.bacterial_guilds:
            return

    # Current ingestion stores compact Phylum statistics here.
    df = fetch_taxon_abundance(
        marker=marker,
        level=level,
    )

    if sample_ids and not df.empty:
        df = df[
            df["sample_id"]
            .fillna("")
            .astype(str)
            .str.upper()
            .isin(sample_ids)
        ].copy()

    mclient = init_minio()

    generated = 0
    uploaded = 0
    skipped_existing = 0
    missing = 0

    prefix = f"biodiversity/piecharts/{marker}/{level}/"
    existing_objects = set() if args.force else list_existing_minio_objects(
        mclient,
        prefix,
    )

    if df.empty:
        print(
            "[INFO] No compact taxonomic rows found for "
            f"marker={marker}, level={level}."
        )
    else:
        for sample_id, sample_df in df.groupby("sample_id"):
            sample_id = str(sample_id).strip()
            if not sample_id:
                continue

            safe_id = sanitize_filename(sample_id)
            object_name = f"{prefix}{safe_id}.png"

            if not args.force and object_name in existing_objects:
                skipped_existing += 1
                continue

            missing += 1

            if args.dry_run:
                print(f"[NEW] would generate {object_name}")
                continue

            local_png = out_dir / f"{safe_id}.png"
            ok = make_piechart_for_sample(
                sample_df,
                sample_id,
                marker,
                level,
                local_png,
            )
            if not ok:
                print(f"[WARN] No positive chart data for {sample_id}")
                continue

            generated += 1

            uploaded_url = upload_file_to_minio(
                mclient,
                local_png,
                object_name,
                content_type="image/png",
            )
            if uploaded_url:
                uploaded += 1

    print(f"[OK] Missing taxonomic charts: {missing}")
    print(f"[OK] Skipped existing taxonomic charts: {skipped_existing}")
    print(f"[OK] Generated {generated} taxonomic charts")
    print(f"[OK] Uploaded {uploaded} taxonomic charts to MinIO")

    if args.fungal_guilds:
        generate_fungal_guildplots(
            mclient,
            force=args.force,
            sample_ids=sample_ids,
            dry_run=args.dry_run,
        )

    if args.bacterial_guilds:
        generate_bacterial_guildplots_from_faprotax(
            mclient,
            force=args.force,
            sample_ids=sample_ids,
            dry_run=args.dry_run,
        )


if __name__ == "__main__":
    main()
